from fastapi import FastAPI, APIRouter, HTTPException, Depends, Query, BackgroundTasks, Request, File, UploadFile
from fastapi.security import HTTPBearer, HTTPAuthorizationCredentials
from fastapi.responses import StreamingResponse
from fastapi.staticfiles import StaticFiles
from dotenv import load_dotenv
from starlette.middleware.cors import CORSMiddleware
from motor.motor_asyncio import AsyncIOMotorClient
import os
import logging
from pathlib import Path
from pydantic import BaseModel, Field, EmailStr
from typing import List, Optional
import uuid
from datetime import datetime, timezone, timedelta
from passlib.context import CryptContext
from jose import jwt, JWTError
import qrcode
from io import BytesIO
import base64
from openpyxl import Workbook
import random
import string
import razorpay
import httpx
import smtplib
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
import ssl

ROOT_DIR = Path(__file__).parent

# Load environment variables from .env file
env_path = ROOT_DIR / '.env'
if env_path.exists():
    load_dotenv(env_path)
    print(f"Loaded .env from {env_path}")
else:
    load_dotenv()  # Try default locations
    print("Using default .env loading")

# MongoDB connection with fallback defaults
mongo_url = os.environ.get('MONGO_URL', 'mongodb://localhost:27017')
db_name = os.environ.get('DB_NAME', 'bitz_club')
client = AsyncIOMotorClient(mongo_url)
db = client[db_name]
print(f"Connected to MongoDB: {mongo_url}, Database: {db_name}")

# Razorpay client
razorpay_client = razorpay.Client(auth=(
    os.environ.get("RAZORPAY_KEY_ID", ""),
    os.environ.get("RAZORPAY_KEY_SECRET", "")
))

# SoftSMS Configuration
SOFTSMS_API_KEY = os.environ.get("SOFTSMS_API_KEY", "")
SOFTSMS_SENDER_ID = os.environ.get("SOFTSMS_SENDER_ID", "BITZCL")
SOFTSMS_API_URL = os.environ.get("SOFTSMS_API_URL", "https://softsms.in/app/smsapi/index.php")

# SMTP Email Configuration
SMTP_HOST = os.environ.get("SMTP_HOST", "")
SMTP_PORT = int(os.environ.get("SMTP_PORT", 465))
SMTP_USERNAME = os.environ.get("SMTP_USERNAME", "")
SMTP_PASSWORD = os.environ.get("SMTP_PASSWORD", "")

# Security
pwd_context = CryptContext(schemes=["bcrypt"], deprecated="auto")
SECRET_KEY = os.environ.get("JWT_SECRET", "bitz-club-secret-key-2024")
ALGORITHM = "HS256"
ACCESS_TOKEN_EXPIRE_HOURS = 24
security = HTTPBearer()

app = FastAPI(title="BITZ Club Membership API")
api_router = APIRouter(prefix="/api")

# Configure logging
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(name)s - %(levelname)s - %(message)s')
logger = logging.getLogger(__name__)

# ==================== MODELS ====================

class UserRole:
    SUPER_ADMIN = "super_admin"
    ADMIN = "admin"
    TELECALLER = "telecaller"
    MEMBER = "member"

class AdminRole:
    MEMBERSHIP = "membership"
    TELECALLER = "telecaller_manager"
    ACCOUNTS = "accounts"
    MARKETING = "marketing"
    FULL_ACCESS = "full_access"

# Admin Management Models
class AdminCreate(BaseModel):
    name: str
    email: EmailStr  # Email is primary login for admins
    mobile: Optional[str] = None
    password: str
    department: Optional[str] = None  # e.g., Operations, Marketing, Finance
    admin_role: str = AdminRole.MEMBERSHIP
    permissions: List[str] = []  # specific permissions like 'members', 'payments', 'leads', 'reports'

class AdminUpdate(BaseModel):
    name: Optional[str] = None
    email: Optional[EmailStr] = None
    mobile: Optional[str] = None
    department: Optional[str] = None
    admin_role: Optional[str] = None
    permissions: Optional[List[str]] = None
    is_active: Optional[bool] = None

# Activity Log Model
class ActivityLog(BaseModel):
    admin_id: str
    admin_name: str
    action: str
    entity_type: str
    entity_id: Optional[str] = None
    details: Optional[str] = None
    ip_address: Optional[str] = None

class MembershipStatus:
    ACTIVE = "active"
    EXPIRED = "expired"
    PENDING = "pending"
    CANCELLED = "cancelled"

# Auth Models
class UserCreate(BaseModel):
    mobile: str
    password: str
    name: str
    email: Optional[EmailStr] = None
    date_of_birth: Optional[str] = None
    role: str = UserRole.MEMBER

class UserLogin(BaseModel):
    identifier: str  # mobile or member_id
    password: str

class TokenResponse(BaseModel):
    access_token: str
    token_type: str = "bearer"
    user: dict

# Registration with Payment Models
class RegistrationInitiate(BaseModel):
    name: str
    mobile: str
    country_code: Optional[str] = "+91"  # International dialing code
    email: Optional[EmailStr] = None
    date_of_birth: Optional[str] = None
    password: str
    plan_id: str
    referral_id: Optional[str] = None
    photo_base64: Optional[str] = None  # Base64 encoded photo
    country: Optional[str] = "India"
    state: Optional[str] = None

# Member Models
class MemberCreate(BaseModel):
    name: str
    mobile: str
    country_code: Optional[str] = "+91"  # International dialing code
    email: EmailStr  # Email is now MANDATORY for reminders and notifications
    address: Optional[str] = None
    city: Optional[str] = None
    state: Optional[str] = None
    country: Optional[str] = "India"
    pincode: Optional[str] = None
    area: Optional[str] = None
    date_of_birth: Optional[str] = None
    plan_id: str
    password: Optional[str] = None
    referral_id: Optional[str] = None  # Employee ID (BITZ-E001) or Associate ID (BITZ-A001)

class MemberUpdate(BaseModel):
    name: Optional[str] = None
    email: Optional[EmailStr] = None
    country_code: Optional[str] = None
    address: Optional[str] = None
    city: Optional[str] = None
    state: Optional[str] = None
    country: Optional[str] = None
    pincode: Optional[str] = None
    area: Optional[str] = None
    date_of_birth: Optional[str] = None
    plan_id: Optional[str] = None
    referral_id: Optional[str] = None
    photo_url: Optional[str] = None
    status: Optional[str] = None

# Family Member Models
class FamilyMemberCreate(BaseModel):
    member_id: str  # Primary member ID
    name: str
    relationship: str  # spouse, child, parent, sibling
    date_of_birth: Optional[str] = None
    mobile: Optional[str] = None
    email: Optional[EmailStr] = None
    id_proof_type: Optional[str] = None  # aadhaar, pan, passport, driving_license
    id_proof_number: Optional[str] = None
    photo_url: Optional[str] = None

class FamilyMemberUpdate(BaseModel):
    name: Optional[str] = None
    relationship: Optional[str] = None
    date_of_birth: Optional[str] = None
    mobile: Optional[str] = None
    email: Optional[EmailStr] = None
    id_proof_type: Optional[str] = None
    id_proof_number: Optional[str] = None
    photo_url: Optional[str] = None
    is_active: Optional[bool] = None

# Plan Models
class PlanCreate(BaseModel):
    name: str
    description: str
    duration_months: int
    price: float
    currency: str = "INR"  # INR, USD, AED, GBP, EUR, etc.
    price_usd: Optional[float] = None  # Price in USD for international members
    price_aed: Optional[float] = None  # Price in AED for UAE members
    features: List[str] = []
    is_active: bool = True
    # Maintenance configuration per plan
    maintenance_type: str = 'none'  # 'none', 'inclusive', 'enter_value'
    maintenance_amount: float = 0  # Monthly maintenance amount for this plan
    maintenance_gst: float = 0  # GST percentage
    maintenance_billing_cycle: str = 'monthly'  # monthly, quarterly, half_yearly, yearly
    maintenance_frequency: str = 'monthly'  # Legacy field, same as billing_cycle
    renewal_amount: float = 0  # Amount to charge on renewal

class PlanUpdate(BaseModel):
    name: Optional[str] = None
    description: Optional[str] = None
    duration_months: Optional[int] = None
    price: Optional[float] = None
    currency: Optional[str] = None
    price_usd: Optional[float] = None
    price_aed: Optional[float] = None
    features: Optional[List[str]] = None
    is_active: Optional[bool] = None
    maintenance_type: Optional[str] = None
    maintenance_amount: Optional[float] = None
    maintenance_gst: Optional[float] = None
    maintenance_billing_cycle: Optional[str] = None
    maintenance_frequency: Optional[str] = None
    renewal_amount: Optional[float] = None

# Partner Models
class FacilityDiscount(BaseModel):
    facility_name: str
    discount_percentage: float
    description: Optional[str] = None

class PartnerCreate(BaseModel):
    name: str
    description: str
    category: Optional[str] = None  # Hotel, Restaurant, Spa, Gym, etc.
    logo_url: Optional[str] = None
    image_url: Optional[str] = None  # Main image for the affiliation
    contact_email: Optional[EmailStr] = None
    contact_phone: Optional[str] = None
    # Enhanced Affiliation Fields
    contact_person_1: Optional[str] = None  # e.g., "Mr. Jerin T Jose"
    contact_person_1_phone: Optional[str] = None
    contact_person_2: Optional[str] = None
    contact_person_2_phone: Optional[str] = None
    address: Optional[str] = None
    city: Optional[str] = None
    website: Optional[str] = None
    # Offers - can be multiple
    offers: Optional[str] = None  # e.g., "Stay 20% off, Food 15% off"
    facilities: List[FacilityDiscount] = []
    is_active: bool = True
    # For booking
    booking_enabled: bool = True
    booking_instructions: Optional[str] = None

class PartnerUpdate(BaseModel):
    name: Optional[str] = None
    description: Optional[str] = None
    category: Optional[str] = None
    logo_url: Optional[str] = None
    image_url: Optional[str] = None
    contact_email: Optional[EmailStr] = None
    contact_phone: Optional[str] = None
    contact_person_1: Optional[str] = None
    contact_person_1_phone: Optional[str] = None
    contact_person_2: Optional[str] = None
    contact_person_2_phone: Optional[str] = None
    address: Optional[str] = None
    city: Optional[str] = None
    website: Optional[str] = None
    offers: Optional[str] = None
    facilities: Optional[List[FacilityDiscount]] = None
    is_active: Optional[bool] = None
    booking_enabled: Optional[bool] = None
    booking_instructions: Optional[str] = None

# Telecaller Models
class TelecallerCreate(BaseModel):
    name: str
    mobile: str
    email: Optional[EmailStr] = None
    password: str

class TelecallerUpdate(BaseModel):
    name: Optional[str] = None
    email: Optional[EmailStr] = None
    is_active: Optional[bool] = None

# Follow-up Models
class FollowUpCreate(BaseModel):
    member_id: str
    notes: str
    follow_up_date: str
    status: str = "pending"

class FollowUpUpdate(BaseModel):
    notes: Optional[str] = None
    follow_up_date: Optional[str] = None
    status: Optional[str] = None

# Booking Models (for Affiliate Bookings)
class BookingCreate(BaseModel):
    affiliate_id: str  # Partner/Affiliation ID
    booking_date: str  # Date of visit/booking
    notes: Optional[str] = None

class BookingUpdate(BaseModel):
    status: Optional[str] = None  # confirmed, cancelled, completed
    notes: Optional[str] = None

# Offers/Events Models
class OfferCreate(BaseModel):
    title: str
    description: str
    image_url: Optional[str] = None
    discount_text: Optional[str] = None  # e.g., "20% OFF"
    valid_from: Optional[str] = None
    valid_until: Optional[str] = None
    terms: Optional[str] = None
    affiliate_id: Optional[str] = None  # Link to specific affiliate
    category: Optional[str] = None  # Hotel, Restaurant, Event, etc.
    is_active: bool = True
    is_featured: bool = False

class OfferUpdate(BaseModel):
    title: Optional[str] = None
    description: Optional[str] = None
    image_url: Optional[str] = None
    discount_text: Optional[str] = None
    valid_from: Optional[str] = None
    valid_until: Optional[str] = None
    terms: Optional[str] = None
    affiliate_id: Optional[str] = None
    category: Optional[str] = None
    is_active: Optional[bool] = None
    is_featured: Optional[bool] = None

# Gallery Models
class GalleryItemCreate(BaseModel):
    title: str
    image_url: str
    description: Optional[str] = None
    category: Optional[str] = None  # events, partners, members, etc.
    order: int = 0
    is_active: bool = True

class GalleryItemUpdate(BaseModel):
    title: Optional[str] = None
    image_url: Optional[str] = None
    description: Optional[str] = None
    category: Optional[str] = None
    order: Optional[int] = None
    is_active: Optional[bool] = None

# Event Models
class EventCreate(BaseModel):
    title: str
    description: str
    event_date: str  # ISO format date
    event_time: Optional[str] = None
    venue: str
    venue_address: Optional[str] = None
    city: Optional[str] = None
    image_url: Optional[str] = None
    category: Optional[str] = None  # party, networking, workshop, etc.
    max_attendees: Optional[int] = None
    entry_fee: float = 0
    currency: str = "INR"
    is_members_only: bool = True
    is_active: bool = True
    registration_deadline: Optional[str] = None

class EventUpdate(BaseModel):
    title: Optional[str] = None
    description: Optional[str] = None
    event_date: Optional[str] = None
    event_time: Optional[str] = None
    venue: Optional[str] = None
    venue_address: Optional[str] = None
    city: Optional[str] = None
    image_url: Optional[str] = None
    category: Optional[str] = None
    max_attendees: Optional[int] = None
    entry_fee: Optional[float] = None
    currency: Optional[str] = None
    is_members_only: Optional[bool] = None
    is_active: Optional[bool] = None
    registration_deadline: Optional[str] = None

class EventRegistration(BaseModel):
    event_id: str
    member_id: str
    guests: int = 0
    notes: Optional[str] = None

# Payment Models
class PaymentCreate(BaseModel):
    member_id: str
    amount: float
    payment_type: str = "offline"  # 'online' or 'offline'
    payment_method: Optional[str] = None  # 'razorpay', 'cash', 'upi', 'card', etc.
    transaction_id: Optional[str] = None
    plan_id: Optional[str] = None
    notes: Optional[str] = None
    coupon_code: Optional[str] = None
    gst_amount: Optional[float] = None
    discount_amount: Optional[float] = None

class PaymentUpdate(BaseModel):
    amount: Optional[float] = None
    payment_method: Optional[str] = None
    status: Optional[str] = None
    transaction_id: Optional[str] = None
    notes: Optional[str] = None

# Coupon/Discount Models
class CouponCreate(BaseModel):
    code: str
    discount_type: str  # 'percentage' or 'fixed'
    discount_value: float
    min_amount: float = 0
    max_discount: Optional[float] = None
    valid_from: str
    valid_until: str
    usage_limit: int = 0  # 0 means unlimited
    applicable_plans: List[str] = []  # Empty means all plans
    is_active: bool = True

class CouponUpdate(BaseModel):
    discount_type: Optional[str] = None
    discount_value: Optional[float] = None
    min_amount: Optional[float] = None
    max_discount: Optional[float] = None
    valid_from: Optional[str] = None
    valid_until: Optional[str] = None
    usage_limit: Optional[int] = None
    applicable_plans: Optional[List[str]] = None
    is_active: Optional[bool] = None

# Maintenance Fee Models
class MaintenanceFeeCreate(BaseModel):
    member_id: str
    amount: float
    fee_type: str  # 'monthly', 'quarterly', 'yearly'
    due_date: str
    payment_method: Optional[str] = None
    transaction_id: Optional[str] = None
    notes: Optional[str] = None
    # Enhanced fields for category-based maintenance
    plan_id: Optional[str] = None
    discount_amount: float = 0  # Discount applied
    discount_reason: Optional[str] = None
    tax_rate: float = 0  # Tax rate (0-5%)
    tax_amount: float = 0  # Calculated tax

class MaintenanceFeeUpdate(BaseModel):
    amount: Optional[float] = None
    status: Optional[str] = None
    payment_method: Optional[str] = None
    paid_date: Optional[str] = None
    transaction_id: Optional[str] = None
    notes: Optional[str] = None
    discount_amount: Optional[float] = None
    discount_reason: Optional[str] = None
    tax_rate: Optional[float] = None
    tax_amount: Optional[float] = None

# Renewal Receipt Model
class RenewalReceiptCreate(BaseModel):
    member_id: str
    plan_id: str
    amount: float
    payment_method: str  # 'cash', 'upi', 'card', 'netbanking', 'razorpay'
    transaction_id: Optional[str] = None
    notes: Optional[str] = None
    discount_amount: float = 0
    tax_amount: float = 0

# GST Configuration
GST_RATE = 0.18  # 18% GST

# Lead Models
class LeadCreate(BaseModel):
    name: str
    mobile: str
    city: str
    interested_in: str  # 'membership' or 'partnership'
    source: Optional[str] = 'landing_page'

class LeadUpdate(BaseModel):
    status: Optional[str] = None
    notes: Optional[str] = None
    assigned_telecaller: Optional[str] = None  # Telecaller user ID
    follow_up_date: Optional[str] = None
    priority: Optional[str] = None  # high, medium, low

class LeadAssign(BaseModel):
    lead_ids: List[str]  # Multiple leads can be assigned at once
    telecaller_id: str

# Marketing Landing Page Models
class MarketingLeadCreate(BaseModel):
    name: str
    mobile: str
    email: Optional[EmailStr] = None
    referral_code: Optional[str] = None
    country_code: Optional[str] = '+91'
    country: Optional[str] = 'India'
    member_type: Optional[str] = 'indian'  # indian, nri, foreigner
    source: str = 'marketing_landing'

class MarketingLeadStep2(BaseModel):
    lead_id: str
    address: Optional[str] = None
    city: Optional[str] = None
    state: Optional[str] = None
    pincode: Optional[str] = None
    country: Optional[str] = 'India'
    date_of_birth: Optional[str] = None
    plan_id: str
    password: str

class EnquiryCreate(BaseModel):
    name: str
    mobile: str
    email: Optional[EmailStr] = None
    message: str
    source: str = 'marketing_landing'

class LeadStatus:
    NEW = "new"
    CONTACTED = "contacted"
    CONVERTED = "converted"
    NOT_INTERESTED = "not_interested"

# Content Management Models
class ExperienceContent(BaseModel):
    id: Optional[str] = None
    title: str
    image_url: str
    discount: str
    description: str
    icon: str  # Icon name for frontend
    order: int = 0
    is_active: bool = True

class GalleryImage(BaseModel):
    id: Optional[str] = None
    title: str
    image_url: str
    category: str  # e.g., 'hotel', 'dining', 'spa', etc.
    order: int = 0
    is_active: bool = True

class WebsiteSettings(BaseModel):
    hero_title: Optional[str] = None
    hero_subtitle: Optional[str] = None
    hero_image: Optional[str] = None
    contact_phone: Optional[str] = None
    contact_email: Optional[str] = None
    contact_address: Optional[str] = None

# ==================== UTILITY FUNCTIONS ====================

async def generate_member_id():
    """
    Generate unique 10-digit Julian member ID in format: YYDDDNNNNN
    - YY: Last 2 digits of year (26 for 2026)
    - DDD: Day of year (001-366)
    - NNNNN: Sequential 5-digit counter
    
    Example: 2609300001 (Year 2026, Day 93, Member #00001)
    """
    now = datetime.now()
    year_part = str(now.year)[2:]  # Last 2 digits: "26" for 2026
    day_of_year = now.timetuple().tm_yday  # Day of year (1-366)
    day_part = str(day_of_year).zfill(3)  # Pad to 3 digits: "093"
    
    # Get and increment counter from database (atomic operation)
    counter_doc = await db.counters.find_one_and_update(
        {"_id": "member_id"},
        {"$inc": {"sequence_value": 1}},
        upsert=True,
        return_document=True
    )
    
    sequence = counter_doc.get("sequence_value", 1)
    sequence_part = str(sequence).zfill(5)  # Pad to 5 digits: "00001"
    
    member_id = f"{year_part}{day_part}{sequence_part}"
    logger.info(f"[MEMBER ID] Generated Julian ID: {member_id}")
    
    return member_id

def generate_member_id_sync():
    """Synchronous fallback for non-async contexts (legacy)"""
    now = datetime.now()
    year_part = str(now.year)[2:]
    day_of_year = now.timetuple().tm_yday
    day_part = str(day_of_year).zfill(3)
    random_part = ''.join(random.choices(string.digits, k=5))
    return f"{year_part}{day_part}{random_part}"

def generate_qr_code(data: str) -> str:
    """Generate QR code and return base64 string"""
    qr = qrcode.QRCode(version=1, box_size=10, border=4)
    qr.add_data(data)
    qr.make(fit=True)
    img = qr.make_image(fill_color="black", back_color="white")
    buffer = BytesIO()
    img.save(buffer, format="PNG")
    return base64.b64encode(buffer.getvalue()).decode()

def create_access_token(data: dict, expires_delta: Optional[timedelta] = None):
    to_encode = data.copy()
    expire = datetime.now(timezone.utc) + (expires_delta or timedelta(hours=ACCESS_TOKEN_EXPIRE_HOURS))
    to_encode.update({"exp": expire})
    return jwt.encode(to_encode, SECRET_KEY, algorithm=ALGORITHM)

async def get_current_user(credentials: HTTPAuthorizationCredentials = Depends(security)):
    try:
        payload = jwt.decode(credentials.credentials, SECRET_KEY, algorithms=[ALGORITHM])
        user_id = payload.get("sub")
        role = payload.get("role")
        if user_id is None:
            raise HTTPException(status_code=401, detail="Invalid token")
        
        # Check if it's an admin/super_admin - they're in the admins collection
        if role in [UserRole.SUPER_ADMIN, UserRole.ADMIN]:
            user = await db.admins.find_one({"id": user_id}, {"_id": 0})
            if user:
                return user
        
        # Check regular users collection
        user = await db.users.find_one({"id": user_id}, {"_id": 0})
        if not user:
            raise HTTPException(status_code=401, detail="User not found")
        return user
    except JWTError:
        raise HTTPException(status_code=401, detail="Invalid token")

async def require_admin(current_user: dict = Depends(get_current_user)):
    """Allow Super Admin and Admin users"""
    if current_user.get("role") not in [UserRole.SUPER_ADMIN, UserRole.ADMIN]:
        raise HTTPException(status_code=403, detail="Admin access required")
    return current_user

async def require_super_admin(current_user: dict = Depends(get_current_user)):
    """Only Super Admin can access"""
    if current_user.get("role") != UserRole.SUPER_ADMIN:
        raise HTTPException(status_code=403, detail="Super Admin access required")
    return current_user

async def require_admin_or_telecaller(current_user: dict = Depends(get_current_user)):
    if current_user.get("role") not in [UserRole.SUPER_ADMIN, UserRole.ADMIN, UserRole.TELECALLER]:
        raise HTTPException(status_code=403, detail="Admin or Telecaller access required")
    return current_user

# ==================== EMAIL CONFIGURATION ====================

EMAIL_CONFIG = {
    "LEADS_EMAIL": "leads@bitzclub.com",
    "ADMIN_EMAIL": "admin@bitzclub.com",
    "FROM_EMAIL": os.environ.get("SENDER_EMAIL", "noreply@bitzclub.com")
}

# ==================== RAZORPAY PAYMENT SERVICE ====================

# Log Razorpay mode at startup
_razorpay_key = os.environ.get("RAZORPAY_KEY_ID", "")
_razorpay_mode = "LIVE" if _razorpay_key.startswith("rzp_live_") else "TEST"
logger.info(f"[RAZORPAY] Initialized in {_razorpay_mode} mode (Key: {_razorpay_key[:15]}...)")

class RazorpayService:
    """Real Razorpay Payment Integration"""
    
    @staticmethod
    async def create_order(amount: float, member_id: str, plan_id: str = None, notes: dict = None) -> dict:
        """Create a Razorpay order"""
        try:
            order_data = {
                "amount": int(amount * 100),  # Razorpay expects amount in paise
                "currency": "INR",
                "receipt": f"rcpt_{member_id[:8]}_{uuid.uuid4().hex[:8]}",
                "notes": notes or {"member_id": member_id, "plan_id": plan_id or ""}
            }
            
            order = razorpay_client.order.create(data=order_data)
            logger.info(f"[RAZORPAY] Order created: {order.get('id')} for amount: {amount}")
            
            return {
                "id": order.get("id"),
                "amount": order.get("amount"),
                "amount_due": order.get("amount_due"),
                "currency": order.get("currency"),
                "status": order.get("status"),
                "member_id": member_id,
                "razorpay_key": os.environ.get("RAZORPAY_KEY_ID")
            }
        except Exception as e:
            logger.error(f"[RAZORPAY ERROR] Failed to create order: {str(e)}")
            raise HTTPException(status_code=500, detail=f"Payment order creation failed: {str(e)}")
    
    @staticmethod
    def verify_payment_signature(payment_id: str, order_id: str, signature: str) -> bool:
        """Verify Razorpay payment signature - STRICT verification only"""
        try:
            razorpay_key = os.environ.get("RAZORPAY_KEY_ID", "")
            is_live = razorpay_key.startswith("rzp_live_")
            logger.info(f"[RAZORPAY] Verifying payment: {payment_id}, order: {order_id}, mode: {'LIVE' if is_live else 'TEST'}")
            
            params_dict = {
                'razorpay_order_id': order_id,
                'razorpay_payment_id': payment_id,
                'razorpay_signature': signature
            }
            
            # Primary verification - Razorpay signature check
            razorpay_client.utility.verify_payment_signature(params_dict)
            logger.info(f"[RAZORPAY] Payment signature verified successfully for: {payment_id}")
            
            # Additional verification - Fetch payment and check status
            try:
                payment = razorpay_client.payment.fetch(payment_id)
                payment_status = payment.get('status')
                payment_order_id = payment.get('order_id')
                payment_amount = payment.get('amount', 0) / 100  # Convert paise to rupees
                
                logger.info(f"[RAZORPAY] Payment details - Status: {payment_status}, Amount: {payment_amount}, Order: {payment_order_id}")
                
                # Verify order ID matches
                if payment_order_id != order_id:
                    logger.error(f"[RAZORPAY ERROR] Order ID mismatch: expected {order_id}, got {payment_order_id}")
                    return False
                
                # For LIVE mode, payment must be captured
                if is_live and payment_status != 'captured':
                    logger.error(f"[RAZORPAY ERROR] Payment not captured in LIVE mode - status: {payment_status}")
                    return False
                    
            except Exception as fetch_error:
                logger.warning(f"[RAZORPAY] Could not fetch payment details: {str(fetch_error)}")
                # Continue if signature was verified - this is acceptable
            
            return True
            
        except razorpay.errors.SignatureVerificationError as e:
            logger.error(f"[RAZORPAY ERROR] Signature verification failed: {str(e)}")
            return False
        except Exception as e:
            logger.error(f"[RAZORPAY ERROR] Verification error: {str(e)}")
            return False
    
    @staticmethod
    async def fetch_payment(payment_id: str) -> dict:
        """Fetch payment details from Razorpay"""
        try:
            payment = razorpay_client.payment.fetch(payment_id)
            return payment
        except Exception as e:
            logger.error(f"[RAZORPAY ERROR] Failed to fetch payment: {str(e)}")
            return None

# Keep MockedPaymentService for fallback/offline payments
class MockedPaymentService:
    """Fallback for offline payments"""
    @staticmethod
    async def create_order(amount: float, member_id: str) -> dict:
        order_id = f"offline_{uuid.uuid4().hex[:16]}"
        return {
            "id": order_id,
            "amount": int(amount * 100),
            "currency": "INR",
            "status": "created",
            "member_id": member_id,
            "payment_type": "offline"
        }
    
    @staticmethod
    async def verify_payment(payment_id: str, order_id: str, signature: str) -> bool:
        return True
    
    @staticmethod
    async def capture_payment(payment_id: str, amount: float) -> dict:
        return {
            "payment_id": payment_id,
            "status": "captured",
            "amount": amount
        }

class EmailService:
    """Real SMTP Email Service"""
    
    @staticmethod
    async def send_email(to_email: str, subject: str, html_content: str) -> bool:
        """
        Send email using SMTP.
        If SMTP credentials are not configured, it logs the email and returns True.
        """
        logger.info(f"[EMAIL] To: {to_email}")
        logger.info(f"[EMAIL] Subject: {subject}")
        
        # Check if SMTP is configured
        if not SMTP_HOST or not SMTP_USERNAME or SMTP_PASSWORD == "(webmail password)":
            logger.warning("[EMAIL] SMTP not fully configured, email logged but not sent")
            logger.info(f"[EMAIL] Content Preview: {html_content[:200]}...")
            return True
        
        try:
            # Create message
            message = MIMEMultipart("alternative")
            message["Subject"] = subject
            message["From"] = EMAIL_CONFIG['FROM_EMAIL']
            message["To"] = to_email
            
            # Attach HTML content
            html_part = MIMEText(html_content, "html")
            message.attach(html_part)
            
            # Create SSL context and send
            context = ssl.create_default_context()
            with smtplib.SMTP_SSL(SMTP_HOST, SMTP_PORT, context=context) as server:
                server.login(SMTP_USERNAME, SMTP_PASSWORD)
                server.sendmail(EMAIL_CONFIG['FROM_EMAIL'], to_email, message.as_string())
            
            logger.info(f"[EMAIL] Successfully sent to {to_email}")
            return True
        except Exception as e:
            logger.error(f"[EMAIL ERROR] Failed to send email to {to_email}: {str(e)}")
            # Return True to not break the flow, but log the error
            return True
    
    @staticmethod
    async def send_lead_notification(lead: dict) -> bool:
        """Send lead notification to leads@bitzclub.com"""
        subject = f"New Lead Enquiry: {lead['name']} - {lead['interested_in'].title()}"
        
        html_content = f"""
        <html>
        <body style="font-family: Arial, sans-serif; padding: 20px; background-color: #f5f5f5;">
            <div style="max-width: 600px; margin: 0 auto; background: white; padding: 30px; border-radius: 10px;">
                <div style="text-align: center; margin-bottom: 30px;">
                    <h1 style="color: #D4AF37; margin: 0;">BITZ Club</h1>
                    <p style="color: #666; margin-top: 5px;">New Lead Enquiry</p>
                </div>
                
                <div style="background: #f9f9f9; padding: 20px; border-radius: 8px; margin-bottom: 20px;">
                    <h2 style="color: #333; margin-top: 0; font-size: 18px;">Lead Details</h2>
                    <table style="width: 100%; border-collapse: collapse;">
                        <tr>
                            <td style="padding: 10px 0; border-bottom: 1px solid #eee; color: #666; width: 40%;">Name</td>
                            <td style="padding: 10px 0; border-bottom: 1px solid #eee; color: #333; font-weight: bold;">{lead['name']}</td>
                        </tr>
                        <tr>
                            <td style="padding: 10px 0; border-bottom: 1px solid #eee; color: #666;">Mobile Number</td>
                            <td style="padding: 10px 0; border-bottom: 1px solid #eee; color: #333; font-weight: bold;">{lead['mobile']}</td>
                        </tr>
                        <tr>
                            <td style="padding: 10px 0; border-bottom: 1px solid #eee; color: #666;">City</td>
                            <td style="padding: 10px 0; border-bottom: 1px solid #eee; color: #333; font-weight: bold;">{lead['city']}</td>
                        </tr>
                        <tr>
                            <td style="padding: 10px 0; color: #666;">Interested In</td>
                            <td style="padding: 10px 0; color: #D4AF37; font-weight: bold; text-transform: uppercase;">{lead['interested_in']}</td>
                        </tr>
                    </table>
                </div>
                
                <div style="background: #D4AF37; color: #000; padding: 15px; border-radius: 8px; text-align: center;">
                    <p style="margin: 0; font-weight: bold;">Source: {lead.get('source', 'Landing Page')}</p>
                    <p style="margin: 5px 0 0 0; font-size: 12px;">Submitted on {datetime.now().strftime('%d %B %Y at %I:%M %p')}</p>
                </div>
                
                <div style="margin-top: 20px; text-align: center; color: #999; font-size: 12px;">
                    <p>This lead has been saved in your Admin Dashboard</p>
                    <p>Please follow up within 24 hours for best conversion</p>
                </div>
            </div>
        </body>
        </html>
        """
        
        return await EmailService.send_email(
            EMAIL_CONFIG['LEADS_EMAIL'],
            subject,
            html_content
        )
    
    @staticmethod
    async def send_membership_notification(member: dict, plan: dict) -> bool:
        """Send membership registration notification to admin@bitzclub.com"""
        subject = f"New Membership Registration: {member['name']} - {plan['name']} Plan"
        
        join_date = datetime.now().strftime('%d %B %Y')
        referral_id = member.get('referral_id', 'N/A') or 'N/A'
        
        html_content = f"""
        <html>
        <body style="font-family: Arial, sans-serif; padding: 20px; background-color: #f5f5f5;">
            <div style="max-width: 600px; margin: 0 auto; background: white; padding: 30px; border-radius: 10px;">
                <div style="text-align: center; margin-bottom: 30px;">
                    <h1 style="color: #D4AF37; margin: 0;">BITZ Club</h1>
                    <p style="color: #666; margin-top: 5px;">New Membership Registration</p>
                </div>
                
                <div style="background: linear-gradient(135deg, #D4AF37 0%, #8F741F 100%); color: white; padding: 20px; border-radius: 8px; margin-bottom: 20px; text-align: center;">
                    <h2 style="margin: 0; font-size: 24px;">{plan['name']} Membership</h2>
                    <p style="margin: 5px 0 0 0; opacity: 0.9;">₹{plan['price']:,.0f} / {plan['duration_months']} months</p>
                </div>
                
                <div style="background: #f9f9f9; padding: 20px; border-radius: 8px; margin-bottom: 20px;">
                    <h3 style="color: #333; margin-top: 0; font-size: 16px;">Member Details</h3>
                    <table style="width: 100%; border-collapse: collapse;">
                        <tr>
                            <td style="padding: 10px 0; border-bottom: 1px solid #eee; color: #666; width: 40%;">Member Name</td>
                            <td style="padding: 10px 0; border-bottom: 1px solid #eee; color: #333; font-weight: bold;">{member['name']}</td>
                        </tr>
                        <tr>
                            <td style="padding: 10px 0; border-bottom: 1px solid #eee; color: #666;">Membership ID</td>
                            <td style="padding: 10px 0; border-bottom: 1px solid #eee; color: #D4AF37; font-weight: bold; font-family: monospace;">{member['member_id']}</td>
                        </tr>
                        <tr>
                            <td style="padding: 10px 0; border-bottom: 1px solid #eee; color: #666;">Membership Plan</td>
                            <td style="padding: 10px 0; border-bottom: 1px solid #eee; color: #333; font-weight: bold;">{plan['name']}</td>
                        </tr>
                        <tr>
                            <td style="padding: 10px 0; border-bottom: 1px solid #eee; color: #666;">Join Date</td>
                            <td style="padding: 10px 0; border-bottom: 1px solid #eee; color: #333; font-weight: bold;">{join_date}</td>
                        </tr>
                        <tr>
                            <td style="padding: 10px 0; border-bottom: 1px solid #eee; color: #666;">Referral ID</td>
                            <td style="padding: 10px 0; border-bottom: 1px solid #eee; color: #333; font-weight: bold;">{referral_id}</td>
                        </tr>
                        <tr>
                            <td style="padding: 10px 0; border-bottom: 1px solid #eee; color: #666;">Mobile</td>
                            <td style="padding: 10px 0; border-bottom: 1px solid #eee; color: #333;">{member['mobile']}</td>
                        </tr>
                        <tr>
                            <td style="padding: 10px 0; color: #666;">Email</td>
                            <td style="padding: 10px 0; color: #333;">{member.get('email', 'N/A') or 'N/A'}</td>
                        </tr>
                    </table>
                </div>
                
                <div style="background: #e8f5e9; border: 1px solid #4caf50; padding: 15px; border-radius: 8px; text-align: center;">
                    <p style="margin: 0; color: #2e7d32; font-weight: bold;">Status: Pending Payment</p>
                    <p style="margin: 5px 0 0 0; color: #666; font-size: 12px;">Member details saved in Admin Dashboard</p>
                </div>
                
                <div style="margin-top: 20px; text-align: center; color: #999; font-size: 12px;">
                    <p>Registration completed on {datetime.now().strftime('%d %B %Y at %I:%M %p')}</p>
                </div>
            </div>
        </body>
        </html>
        """
        
        return await EmailService.send_email(
            EMAIL_CONFIG['ADMIN_EMAIL'],
            subject,
            html_content
        )
    
    @staticmethod
    async def send_welcome_email(member: dict) -> bool:
        """Send welcome email to the new member"""
        subject = f"Welcome to BITZ Club - Your Membership ID: {member['member_id']}"
        
        html_content = f"""
        <html>
        <body style="font-family: Arial, sans-serif; padding: 20px; background-color: #f5f5f5;">
            <div style="max-width: 600px; margin: 0 auto; background: #0F0F10; padding: 30px; border-radius: 10px; color: white;">
                <div style="text-align: center; margin-bottom: 30px;">
                    <h1 style="color: #D4AF37; margin: 0;">Welcome to BITZ Club</h1>
                    <p style="color: #999; margin-top: 5px;">Your Premium Lifestyle Membership</p>
                </div>
                
                <div style="text-align: center; margin-bottom: 30px;">
                    <p style="font-size: 18px; margin: 0;">Dear <strong>{member['name']}</strong>,</p>
                    <p style="color: #999;">Thank you for joining the BITZ Club family!</p>
                </div>
                
                <div style="background: linear-gradient(135deg, rgba(212,175,55,0.2) 0%, rgba(0,0,0,0.5) 100%); border: 2px solid #D4AF37; padding: 20px; border-radius: 8px; text-align: center; margin-bottom: 20px;">
                    <p style="color: #D4AF37; margin: 0; font-size: 12px; text-transform: uppercase; letter-spacing: 2px;">Your Membership ID</p>
                    <p style="font-size: 28px; font-weight: bold; margin: 10px 0; font-family: monospace; color: #D4AF37;">{member['member_id']}</p>
                </div>
                
                <div style="text-align: center; padding: 20px 0; border-top: 1px solid #333;">
                    <p style="color: #999; font-size: 14px;">Login to your member dashboard to:</p>
                    <ul style="list-style: none; padding: 0; color: #ccc;">
                        <li style="padding: 5px 0;">✓ View your digital membership card</li>
                        <li style="padding: 5px 0;">✓ Download your QR code</li>
                        <li style="padding: 5px 0;">✓ Explore partner benefits</li>
                    </ul>
                </div>
                
                <div style="text-align: center; margin-top: 20px;">
                    <p style="color: #666; font-size: 12px;">For support, contact us at hello@bitzclub.com</p>
                </div>
            </div>
        </body>
        </html>
        """
        
        member_email = member.get('email')
        if member_email:
            return await EmailService.send_email(member_email, subject, html_content)
        return True
    
    @staticmethod
    async def send_payment_receipt(member: dict, payment: dict) -> bool:
        """Send payment receipt to admin and member"""
        # This would send payment confirmation emails
        return await EmailService.send_email(
            member.get("email", EMAIL_CONFIG['ADMIN_EMAIL']),
            "Payment Receipt - BITZ Club",
            f"Payment of ₹{payment['amount']} received successfully for {member['name']}"
        )
    
    @staticmethod
    async def send_maintenance_reminder(member: dict, fee: dict) -> bool:
        """Send maintenance fee reminder email"""
        subject = "Maintenance Fee Reminder - BITZ Club"
        
        html_content = f"""
        <html>
        <body style="font-family: Arial, sans-serif; padding: 20px; background-color: #f5f5f5;">
            <div style="max-width: 600px; margin: 0 auto; background: #0F0F10; padding: 30px; border-radius: 10px; color: white;">
                <div style="text-align: center; margin-bottom: 30px;">
                    <h1 style="color: #D4AF37; margin: 0;">BITZ Club</h1>
                    <p style="color: #999; margin-top: 5px;">Maintenance Fee Reminder</p>
                </div>
                
                <p style="font-size: 16px;">Dear <strong>{member['name']}</strong>,</p>
                
                <p style="color: #ccc;">This is a reminder that your maintenance fee is due.</p>
                
                <div style="background: rgba(212,175,55,0.1); border: 1px solid #D4AF37; padding: 20px; border-radius: 8px; margin: 20px 0;">
                    <p style="margin: 5px 0;"><strong>Member ID:</strong> {member['member_id']}</p>
                    <p style="margin: 5px 0;"><strong>Fee Type:</strong> {fee.get('fee_type', 'Monthly')}</p>
                    <p style="margin: 5px 0;"><strong>Amount:</strong> ₹{fee.get('amount', 0)}</p>
                    <p style="margin: 5px 0;"><strong>Due Date:</strong> {fee.get('due_date', '')}</p>
                </div>
                
                <p style="color: #ccc;">Please make the payment at your earliest convenience to continue enjoying uninterrupted BITZ Club benefits.</p>
                
                <div style="text-align: center; margin-top: 20px;">
                    <p style="color: #666; font-size: 12px;">For support, contact us at hello@bitzclub.com</p>
                </div>
            </div>
        </body>
        </html>
        """
        
        member_email = member.get('email')
        if member_email:
            return await EmailService.send_email(member_email, subject, html_content)
        return False
    
    @staticmethod
    async def send_renewal_reminder(member: dict, days_left: int) -> bool:
        """Send membership renewal reminder email"""
        subject = "Membership Renewal Reminder - BITZ Club"
        
        html_content = f"""
        <html>
        <body style="font-family: Arial, sans-serif; padding: 20px; background-color: #f5f5f5;">
            <div style="max-width: 600px; margin: 0 auto; background: #0F0F10; padding: 30px; border-radius: 10px; color: white;">
                <div style="text-align: center; margin-bottom: 30px;">
                    <h1 style="color: #D4AF37; margin: 0;">BITZ Club</h1>
                    <p style="color: #999; margin-top: 5px;">Membership Renewal Reminder</p>
                </div>
                
                <p style="font-size: 16px;">Dear <strong>{member['name']}</strong>,</p>
                
                <div style="background: #ff6b6b; color: white; padding: 15px; border-radius: 8px; text-align: center; margin: 20px 0;">
                    <p style="margin: 0; font-size: 18px; font-weight: bold;">Your membership expires in {days_left} days!</p>
                </div>
                
                <div style="background: rgba(212,175,55,0.1); border: 1px solid #D4AF37; padding: 20px; border-radius: 8px; margin: 20px 0;">
                    <p style="margin: 5px 0;"><strong>Member ID:</strong> {member['member_id']}</p>
                    <p style="margin: 5px 0;"><strong>Current Plan:</strong> {member.get('plan_name', 'N/A')}</p>
                    <p style="margin: 5px 0;"><strong>Expiry Date:</strong> {member.get('membership_end', '')[:10]}</p>
                </div>
                
                <p style="color: #ccc;">Renew now to continue enjoying exclusive BITZ Club privileges at luxury hotels, fine dining, spas, and more!</p>
                
                <div style="text-align: center; margin: 30px 0;">
                    <a href="https://thebitzclub.com/login" style="background: #D4AF37; color: black; padding: 15px 30px; text-decoration: none; border-radius: 5px; font-weight: bold;">Renew Now</a>
                </div>
                
                <div style="text-align: center; margin-top: 20px;">
                    <p style="color: #666; font-size: 12px;">For support, contact us at hello@bitzclub.com</p>
                </div>
            </div>
        </body>
        </html>
        """
        
        member_email = member.get('email')
        if member_email:
            return await EmailService.send_email(member_email, subject, html_content)
        return False

# ==================== SMS SERVICE (SoftSMS) ====================

class SMSService:
    """Real SoftSMS Integration"""
    
    @staticmethod
    async def send_sms(phone: str, message: str) -> bool:
        """Send SMS via SoftSMS API"""
        try:
            # Format phone number (remove +91 prefix if present)
            phone = phone.replace("+91", "").replace(" ", "").strip()
            if len(phone) == 10:
                phone = "91" + phone
            
            params = {
                "key": SOFTSMS_API_KEY,
                "campaign": "0",
                "routeid": "9",
                "type": "text",
                "contacts": phone,
                "senderid": SOFTSMS_SENDER_ID,
                "msg": message
            }
            
            async with httpx.AsyncClient() as client:
                response = await client.get(SOFTSMS_API_URL, params=params, timeout=30)
                logger.info(f"[SMS] To: {phone}, Status: {response.status_code}, Response: {response.text[:100]}")
                return response.status_code == 200
        except Exception as e:
            logger.error(f"[SMS ERROR] Failed to send SMS to {phone}: {str(e)}")
            return False
    
    @staticmethod
    async def send_registration_sms(member: dict) -> bool:
        """Send registration confirmation SMS"""
        message = f"Welcome to BITZ Club! Your Member ID: {member.get('member_id', '')}. Download the app & show your digital card at partner venues. -BITZ Club"
        return await SMSService.send_sms(member.get("mobile", ""), message)
    
    @staticmethod
    async def send_payment_success_sms(member: dict, amount: float, plan_name: str = "") -> bool:
        """Send payment success SMS"""
        message = f"BITZ Club: Payment of Rs.{int(amount)} received successfully for {plan_name or 'membership'}. Your membership is now active. Thank you! -BITZ Club"
        return await SMSService.send_sms(member.get("mobile", ""), message)
    
    @staticmethod
    async def send_membership_activation_sms(member: dict, plan_name: str, validity_end: str) -> bool:
        """Send membership activation SMS"""
        message = f"Congratulations! Your BITZ Club {plan_name} membership is now ACTIVE. Valid till {validity_end}. Enjoy exclusive benefits! -BITZ Club"
        return await SMSService.send_sms(member.get("mobile", ""), message)
    
    @staticmethod
    async def send_renewal_reminder_sms(member: dict, days_left: int) -> bool:
        """Send membership renewal reminder SMS"""
        message = f"BITZ Club: Your membership expires in {days_left} days. Renew now to continue enjoying exclusive benefits. Visit our app or contact us. -BITZ Club"
        return await SMSService.send_sms(member.get("mobile", ""), message)
    
    @staticmethod
    async def send_telecaller_followup_sms(lead: dict, telecaller_name: str) -> bool:
        """Send follow-up alert SMS"""
        message = f"Hi {lead.get('name', '')}, {telecaller_name} from BITZ Club will contact you shortly regarding your inquiry. Thank you for your interest! -BITZ Club"
        return await SMSService.send_sms(lead.get("mobile", ""), message)

# Keep MockedSMSService for fallback if needed
class MockedSMSService:
    """Fallback Mocked SMS Service"""
    @staticmethod
    async def send_sms(phone: str, message: str) -> bool:
        logger.info(f"[MOCKED SMS] To: {phone}, Message: {message}")
        return True
    
    @staticmethod
    async def send_welcome_sms(member: dict) -> bool:
        return await SMSService.send_registration_sms(member)
    
    @staticmethod
    async def send_payment_sms(member: dict, amount: float) -> bool:
        return await SMSService.send_payment_success_sms(member, amount)

# ==================== AUTH ENDPOINTS ====================

@api_router.post("/auth/register", response_model=TokenResponse)
async def register_user(user_data: UserCreate, background_tasks: BackgroundTasks):
    # Check if mobile already exists
    existing = await db.users.find_one({"mobile": user_data.mobile})
    if existing:
        raise HTTPException(status_code=400, detail="Mobile number already registered")
    
    user_id = str(uuid.uuid4())
    member_id = await generate_member_id()
    hashed_password = pwd_context.hash(user_data.password)
    
    user = {
        "id": user_id,
        "member_id": member_id,
        "mobile": user_data.mobile,
        "password": hashed_password,
        "name": user_data.name,
        "email": user_data.email,
        "date_of_birth": user_data.date_of_birth,
        "role": user_data.role,
        "is_active": True,
        "created_at": datetime.now(timezone.utc).isoformat()
    }
    
    await db.users.insert_one(user)
    
    # Remove _id that MongoDB adds after insert
    user.pop('_id', None)
    
    # Send welcome SMS to user
    background_tasks.add_task(SMSService.send_registration_sms, {"mobile": user_data.mobile, "member_id": member_id})
    
    # Send welcome email to user
    if user_data.email:
        background_tasks.add_task(EmailService.send_welcome_email, user)
    
    # If this is a member registration, send notification to admin@bitzclub.com
    if user_data.role == UserRole.MEMBER:
        # Get default plan for notification (or create a basic notification)
        default_plan = {"name": "Self-Registered", "price": 0, "duration_months": 0}
        plans = await db.plans.find({}, {"_id": 0}).to_list(1)
        if plans:
            default_plan = plans[0]
        
        member_notification = {
            "member_id": member_id,
            "name": user_data.name,
            "mobile": user_data.mobile,
            "email": user_data.email,
            "referral_id": None
        }
        background_tasks.add_task(EmailService.send_membership_notification, member_notification, default_plan)
    
    token = create_access_token({"sub": user_id, "role": user_data.role})
    user_response = {k: v for k, v in user.items() if k != "password"}
    
    return TokenResponse(access_token=token, user=user_response)

@api_router.post("/auth/login", response_model=TokenResponse)
async def login(credentials: UserLogin):
    # First check if it's an admin login (by email)
    if '@' in credentials.identifier:
        # Admin login by email
        admin = await db.admins.find_one({"email": credentials.identifier.lower()}, {"_id": 0})
        if admin and pwd_context.verify(credentials.password, admin["password"]):
            if not admin.get("is_active", True):
                raise HTTPException(status_code=401, detail="Account is disabled")
            
            token = create_access_token({"sub": admin["id"], "role": admin.get("role", UserRole.ADMIN)})
            admin_response = {k: v for k, v in admin.items() if k != "password"}
            return TokenResponse(access_token=token, user=admin_response)
    
    # Regular user login by mobile or member_id
    user = await db.users.find_one({
        "$or": [
            {"mobile": credentials.identifier},
            {"member_id": credentials.identifier},
            {"email": credentials.identifier.lower()}  # Also check email for regular users
        ]
    }, {"_id": 0})
    
    if not user or not pwd_context.verify(credentials.password, user["password"]):
        raise HTTPException(status_code=401, detail="Invalid credentials")
    
    if not user.get("is_active", True):
        raise HTTPException(status_code=401, detail="Account is disabled")
    
    token = create_access_token({"sub": user["id"], "role": user["role"]})
    user_response = {k: v for k, v in user.items() if k != "password"}
    
    return TokenResponse(access_token=token, user=user_response)

@api_router.get("/auth/me")
async def get_current_user_info(current_user: dict = Depends(get_current_user)):
    return {k: v for k, v in current_user.items() if k != "password"}

# ==================== ADMIN USER MANAGEMENT (Super Admin Only) ====================

def generate_admin_id():
    """Generate unique admin ID like BITZ-ADM-001"""
    import random
    return f"BITZ-ADM-{random.randint(100, 999)}"

@api_router.get("/superadmin/admins")
async def get_all_admins(super_admin: dict = Depends(require_super_admin)):
    """Get all admin users (Super Admin only)"""
    admins = await db.admins.find({}, {"_id": 0, "password": 0}).sort("created_at", -1).to_list(100)
    return admins

@api_router.post("/superadmin/admins")
async def create_admin(admin: AdminCreate, super_admin: dict = Depends(require_super_admin)):
    """Create a new admin user (Super Admin only)"""
    # Check if email already exists
    existing = await db.admins.find_one({"email": admin.email.lower()})
    if existing:
        raise HTTPException(status_code=400, detail="Email already registered")
    
    admin_id = generate_admin_id()
    user_id = str(uuid.uuid4())
    
    admin_doc = {
        "id": user_id,
        "admin_id": admin_id,
        "name": admin.name,
        "email": admin.email.lower(),
        "mobile": admin.mobile,
        "password": pwd_context.hash(admin.password),
        "department": admin.department,
        "role": UserRole.ADMIN,
        "admin_role": admin.admin_role,
        "permissions": admin.permissions or ["members", "leads"],  # Default permissions
        "is_active": True,
        "created_at": datetime.now(timezone.utc).isoformat(),
        "created_by": super_admin["id"]
    }
    await db.admins.insert_one(admin_doc)
    
    # Log activity
    await db.activity_logs.insert_one({
        "id": str(uuid.uuid4()),
        "admin_id": super_admin["id"],
        "admin_name": super_admin.get("name", "Super Admin"),
        "action": "create_admin",
        "entity_type": "admin",
        "entity_id": admin_id,
        "details": f"Created admin: {admin.name} ({admin.email})",
        "created_at": datetime.now(timezone.utc).isoformat()
    })
    
    logger.info(f"[ADMIN] Created new admin user: {admin_id} - {admin.name}")
    
    return {
        "message": "Admin created successfully",
        "admin_id": admin_id,
        "email": admin.email.lower(),
        "name": admin.name
    }

@api_router.get("/superadmin/admins/{admin_id}")
async def get_admin(admin_id: str, super_admin: dict = Depends(require_super_admin)):
    """Get admin details by ID"""
    admin = await db.admins.find_one(
        {"$or": [{"id": admin_id}, {"admin_id": admin_id}]},
        {"_id": 0, "password": 0}
    )
    if not admin:
        raise HTTPException(status_code=404, detail="Admin not found")
    return admin

@api_router.put("/superadmin/admins/{admin_id}")
async def update_admin(admin_id: str, update: AdminUpdate, super_admin: dict = Depends(require_super_admin)):
    """Update admin details"""
    update_data = {k: v for k, v in update.dict().items() if v is not None}
    if not update_data:
        raise HTTPException(status_code=400, detail="No update data provided")
    
    update_data["updated_at"] = datetime.now(timezone.utc).isoformat()
    update_data["updated_by"] = super_admin["id"]
    
    result = await db.admins.update_one(
        {"$or": [{"id": admin_id}, {"admin_id": admin_id}]},
        {"$set": update_data}
    )
    
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Admin not found")
    
    # Log activity
    await db.activity_logs.insert_one({
        "id": str(uuid.uuid4()),
        "admin_id": super_admin["id"],
        "admin_name": super_admin.get("name", "Super Admin"),
        "action": "update_admin",
        "entity_type": "admin",
        "entity_id": admin_id,
        "details": f"Updated admin: {admin_id}",
        "created_at": datetime.now(timezone.utc).isoformat()
    })
    
    return {"message": "Admin updated successfully"}

@api_router.delete("/superadmin/admins/{admin_id}")
async def delete_admin(admin_id: str, super_admin: dict = Depends(require_super_admin)):
    """Disable admin (soft delete)"""
    result = await db.admins.update_one(
        {"$or": [{"id": admin_id}, {"admin_id": admin_id}]},
        {"$set": {"is_active": False, "deleted_at": datetime.now(timezone.utc).isoformat()}}
    )
    
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Admin not found")
    
    # Log activity
    await db.activity_logs.insert_one({
        "id": str(uuid.uuid4()),
        "admin_id": super_admin["id"],
        "admin_name": super_admin.get("name", "Super Admin"),
        "action": "delete_admin",
        "entity_type": "admin",
        "entity_id": admin_id,
        "details": f"Disabled admin: {admin_id}",
        "created_at": datetime.now(timezone.utc).isoformat()
    })
    
    return {"message": "Admin disabled successfully"}

@api_router.put("/superadmin/admins/{admin_id}/reset-password")
async def reset_admin_password(admin_id: str, new_password: str, super_admin: dict = Depends(require_super_admin)):
    """Reset admin password"""
    result = await db.admins.update_one(
        {"$or": [{"id": admin_id}, {"admin_id": admin_id}]},
        {"$set": {
            "password": pwd_context.hash(new_password),
            "password_reset_at": datetime.now(timezone.utc).isoformat()
        }}
    )
    
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Admin not found")
    
    return {"message": "Password reset successfully"}

@api_router.get("/superadmin/activity-logs")
async def get_activity_logs(
    admin_id: Optional[str] = None,
    action: Optional[str] = None,
    page: int = Query(1, ge=1),
    limit: int = Query(50, ge=1, le=200),
    super_admin: dict = Depends(require_super_admin)
):
    """Get activity logs"""
    query = {}
    if admin_id:
        query["admin_id"] = admin_id
    if action:
        query["action"] = action
    
    skip = (page - 1) * limit
    total = await db.activity_logs.count_documents(query)
    logs = await db.activity_logs.find(query, {"_id": 0}).sort("created_at", -1).skip(skip).limit(limit).to_list(limit)
    
    return {
        "logs": logs,
        "total": total,
        "page": page,
        "limit": limit,
        "pages": (total + limit - 1) // limit
    }

# ==================== REGISTRATION WITH PAYMENT FLOW ====================

@api_router.post("/registration/initiate")
async def initiate_registration(data: RegistrationInitiate):
    """
    Step 1: Initiate registration - validate data, store temporarily, create Razorpay order
    """
    # Check if mobile already exists
    existing = await db.users.find_one({"mobile": data.mobile})
    if existing:
        raise HTTPException(status_code=400, detail="Mobile number already registered")
    
    # Get plan details
    plan = await db.plans.find_one({"id": data.plan_id}, {"_id": 0})
    if not plan:
        raise HTTPException(status_code=404, detail="Selected plan not found")
    
    # Generate a temporary registration ID
    temp_reg_id = f"REG-{uuid.uuid4().hex[:12].upper()}"
    
    # Store temporary registration data
    temp_registration = {
        "id": temp_reg_id,
        "name": data.name,
        "mobile": data.mobile,
        "email": data.email,
        "date_of_birth": data.date_of_birth,
        "password": pwd_context.hash(data.password),
        "plan_id": data.plan_id,
        "plan_name": plan["name"],
        "plan_price": plan["price"],
        "plan_duration_months": plan["duration_months"],
        "referral_id": data.referral_id,
        "photo_base64": data.photo_base64,
        "status": "pending_payment",
        "created_at": datetime.now(timezone.utc).isoformat(),
        "expires_at": (datetime.now(timezone.utc) + timedelta(hours=24)).isoformat()
    }
    await db.temp_registrations.insert_one(temp_registration)
    
    # Create Razorpay order
    try:
        order = await RazorpayService.create_order(
            amount=plan["price"],
            member_id=temp_reg_id,
            plan_id=data.plan_id,
            notes={
                "registration_id": temp_reg_id,
                "name": data.name,
                "mobile": data.mobile,
                "plan_name": plan["name"]
            }
        )
        
        # Update temp registration with order ID
        await db.temp_registrations.update_one(
            {"id": temp_reg_id},
            {"$set": {"razorpay_order_id": order["id"]}}
        )
        
        return {
            "registration_id": temp_reg_id,
            "order_id": order["id"],
            "amount": plan["price"],
            "currency": "INR",
            "razorpay_key": order["razorpay_key"],
            "plan_name": plan["name"],
            "name": data.name,
            "email": data.email,
            "mobile": data.mobile
        }
    except Exception as e:
        # Clean up temp registration if order creation fails
        await db.temp_registrations.delete_one({"id": temp_reg_id})
        raise HTTPException(status_code=500, detail=f"Failed to create payment order: {str(e)}")

@api_router.post("/registration/complete")
async def complete_registration(
    registration_id: str,
    razorpay_payment_id: str,
    razorpay_order_id: str,
    razorpay_signature: str,
    background_tasks: BackgroundTasks
):
    """
    Step 2: Complete registration after successful payment
    """
    # Get temp registration
    temp_reg = await db.temp_registrations.find_one({"id": registration_id})
    if not temp_reg:
        raise HTTPException(status_code=404, detail="Registration not found or expired")
    
    if temp_reg.get("status") == "completed":
        raise HTTPException(status_code=400, detail="Registration already completed")
    
    # Verify payment signature
    is_valid = RazorpayService.verify_payment_signature(
        razorpay_payment_id, razorpay_order_id, razorpay_signature
    )
    
    if not is_valid:
        await db.temp_registrations.update_one(
            {"id": registration_id},
            {"$set": {"status": "payment_failed", "failed_at": datetime.now(timezone.utc).isoformat()}}
        )
        raise HTTPException(status_code=400, detail="Payment verification failed")
    
    # Generate member ID
    member_id = await generate_member_id()
    user_id = str(uuid.uuid4())
    
    # Create user account
    user = {
        "id": user_id,
        "member_id": member_id,
        "mobile": temp_reg["mobile"],
        "password": temp_reg["password"],  # Already hashed
        "name": temp_reg["name"],
        "email": temp_reg.get("email"),
        "date_of_birth": temp_reg.get("date_of_birth"),
        "role": UserRole.MEMBER,
        "is_active": True,
        "created_at": datetime.now(timezone.utc).isoformat()
    }
    await db.users.insert_one(user)
    
    # Calculate membership dates
    start_date = datetime.now(timezone.utc)
    duration_months = temp_reg.get("plan_duration_months", 12)
    end_date = start_date + timedelta(days=duration_months * 30)
    
    # Handle photo if provided
    photo_url = None
    if temp_reg.get("photo_base64"):
        try:
            # Decode and save photo
            photo_data = base64.b64decode(temp_reg["photo_base64"].split(",")[-1] if "," in temp_reg["photo_base64"] else temp_reg["photo_base64"])
            filename = f"{member_id}_{uuid.uuid4().hex[:8]}.jpg"
            file_path = UPLOAD_DIR / filename
            with open(file_path, "wb") as f:
                f.write(photo_data)
            photo_url = f"/api/uploads/photos/{filename}"
        except Exception as e:
            logger.error(f"Failed to save photo for {member_id}: {str(e)}")
    
    # Create member record
    referred_by = temp_reg.get("referral_id")  # The member who referred this new member
    member_doc = {
        "id": str(uuid.uuid4()),
        "user_id": user_id,
        "member_id": member_id,
        "name": temp_reg["name"],
        "mobile": temp_reg["mobile"],
        "email": temp_reg.get("email"),
        "date_of_birth": temp_reg.get("date_of_birth"),
        "plan_id": temp_reg["plan_id"],
        "plan_name": temp_reg["plan_name"],
        "membership_start": start_date.isoformat(),
        "membership_end": end_date.isoformat(),
        "status": MembershipStatus.ACTIVE,
        "qr_code": generate_qr_code(member_id),
        "photo_url": photo_url,
        "referral_id": referred_by,  # Legacy field - who referred this member
        "referred_by": referred_by,  # New field - who referred this member
        "created_at": datetime.now(timezone.utc).isoformat()
    }
    await db.members.insert_one(member_doc)
    
    # Create referral reward if referred by a member
    if referred_by and (referred_by.startswith("26") or referred_by.startswith("BITZ-")):
        # Create a pending referral reward
        reward_doc = {
            "id": str(uuid.uuid4()),
            "referrer_member_id": referred_by,
            "referred_member_id": member_id,
            "referred_name": temp_reg["name"],
            "reward_amount": 0,  # Admin can set this later
            "reward_type": "pending",
            "status": "pending",
            "notes": f"Auto-created: {temp_reg['name']} joined via referral",
            "created_at": datetime.now(timezone.utc).isoformat()
        }
        await db.referral_rewards.insert_one(reward_doc)
        logger.info(f"[REFERRAL] Created referral reward for {referred_by} (referred {member_id})")
    
    # Create payment record
    payment_doc = {
        "id": str(uuid.uuid4()),
        "order_id": razorpay_order_id,
        "member_id": member_id,
        "member_name": temp_reg["name"],
        "plan_id": temp_reg["plan_id"],
        "plan_name": temp_reg["plan_name"],
        "amount": temp_reg["plan_price"],
        "payment_type": "online",
        "payment_method": "razorpay",
        "razorpay_payment_id": razorpay_payment_id,
        "razorpay_order_id": razorpay_order_id,
        "status": "completed",
        "created_at": datetime.now(timezone.utc).isoformat(),
        "completed_at": datetime.now(timezone.utc).isoformat()
    }
    await db.payments.insert_one(payment_doc)
    
    # Update temp registration status
    await db.temp_registrations.update_one(
        {"id": registration_id},
        {"$set": {"status": "completed", "member_id": member_id, "completed_at": datetime.now(timezone.utc).isoformat()}}
    )
    
    # Send notifications
    validity_end = end_date.strftime("%d %b %Y")
    background_tasks.add_task(SMSService.send_registration_sms, {"mobile": temp_reg["mobile"], "member_id": member_id})
    background_tasks.add_task(SMSService.send_payment_success_sms, {"mobile": temp_reg["mobile"]}, temp_reg["plan_price"], temp_reg["plan_name"])
    background_tasks.add_task(SMSService.send_membership_activation_sms, {"mobile": temp_reg["mobile"]}, temp_reg["plan_name"], validity_end)
    
    if temp_reg.get("email"):
        background_tasks.add_task(EmailService.send_welcome_email, user)
    
    background_tasks.add_task(EmailService.send_membership_notification, member_doc, {"name": temp_reg["plan_name"], "price": temp_reg["plan_price"], "duration_months": duration_months})
    
    # Remove password and _id before creating token
    user.pop('_id', None)
    
    # Create access token
    token = create_access_token({"sub": user_id, "role": UserRole.MEMBER})
    user_response = {k: v for k, v in user.items() if k != "password"}
    
    logger.info(f"[REGISTRATION] Completed registration for {member_id} with plan {temp_reg['plan_name']}")
    
    return {
        "success": True,
        "message": "Registration successful! Welcome to BITZ Club!",
        "member_id": member_id,
        "access_token": token,
        "token_type": "bearer",
        "user": user_response,
        "membership": {
            "plan_name": temp_reg["plan_name"],
            "start_date": start_date.isoformat(),
            "end_date": end_date.isoformat(),
            "status": "active"
        }
    }

# ==================== MEMBERSHIP PLANS ENDPOINTS ====================

@api_router.post("/plans")
async def create_plan(plan: PlanCreate, admin: dict = Depends(require_admin)):
    plan_id = str(uuid.uuid4())
    plan_doc = {
        "id": plan_id,
        **plan.model_dump(),
        "created_at": datetime.now(timezone.utc).isoformat(),
        "created_by": admin["id"]
    }
    await db.plans.insert_one(plan_doc)
    return {"id": plan_id, **plan.model_dump()}

@api_router.get("/plans")
async def get_plans(is_active: Optional[bool] = None):
    query = {}
    if is_active is not None:
        query["is_active"] = is_active
    plans = await db.plans.find(query, {"_id": 0}).to_list(100)
    return plans

@api_router.get("/plans/{plan_id}")
async def get_plan(plan_id: str):
    plan = await db.plans.find_one({"id": plan_id}, {"_id": 0})
    if not plan:
        raise HTTPException(status_code=404, detail="Plan not found")
    return plan

@api_router.put("/plans/{plan_id}")
async def update_plan(plan_id: str, plan: PlanUpdate, admin: dict = Depends(require_admin)):
    update_data = {k: v for k, v in plan.model_dump().items() if v is not None}
    if not update_data:
        raise HTTPException(status_code=400, detail="No data to update")
    
    update_data["updated_at"] = datetime.now(timezone.utc).isoformat()
    result = await db.plans.update_one({"id": plan_id}, {"$set": update_data})
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Plan not found")
    
    return await db.plans.find_one({"id": plan_id}, {"_id": 0})

@api_router.delete("/plans/{plan_id}")
async def delete_plan(plan_id: str, admin: dict = Depends(require_admin)):
    result = await db.plans.delete_one({"id": plan_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Plan not found")
    return {"message": "Plan deleted"}

# ==================== MEMBERS ENDPOINTS ====================

@api_router.post("/members")
async def create_member(member: MemberCreate, background_tasks: BackgroundTasks, user: dict = Depends(require_admin_or_telecaller)):
    # Check if mobile already exists
    existing = await db.members.find_one({"mobile": member.mobile})
    if existing:
        raise HTTPException(status_code=400, detail="Mobile number already registered")
    
    # Get plan details
    plan = await db.plans.find_one({"id": member.plan_id}, {"_id": 0})
    if not plan:
        raise HTTPException(status_code=404, detail="Plan not found")
    
    member_id = await generate_member_id()
    user_id = str(uuid.uuid4())
    
    # Create user account
    password = member.password or ''.join(random.choices(string.ascii_letters + string.digits, k=8))
    hashed_password = pwd_context.hash(password)
    
    user_doc = {
        "id": user_id,
        "member_id": member_id,
        "mobile": member.mobile,
        "password": hashed_password,
        "name": member.name,
        "email": member.email,
        "role": UserRole.MEMBER,
        "is_active": True,
        "created_at": datetime.now(timezone.utc).isoformat()
    }
    await db.users.insert_one(user_doc)
    
    # Calculate membership dates
    start_date = datetime.now(timezone.utc)
    end_date = start_date + timedelta(days=plan["duration_months"] * 30)
    
    member_doc = {
        "id": str(uuid.uuid4()),
        "user_id": user_id,
        "member_id": member_id,
        "name": member.name,
        "mobile": member.mobile,
        "country_code": member.country_code or "+91",
        "email": member.email,
        "address": member.address,
        "city": member.city,
        "state": member.state,
        "country": member.country or "India",
        "pincode": member.pincode,
        "area": member.area,
        "date_of_birth": member.date_of_birth,
        "plan_id": member.plan_id,
        "plan_name": plan["name"],
        "membership_start": start_date.isoformat(),
        "membership_end": end_date.isoformat(),
        "status": MembershipStatus.PENDING,
        "qr_code": generate_qr_code(member_id),
        "referral_id": member.referral_id,
        "assigned_telecaller": user["id"] if user["role"] == UserRole.TELECALLER else None,
        "created_by": user["id"],
        "created_at": datetime.now(timezone.utc).isoformat()
    }
    await db.members.insert_one(member_doc)
    
    # Send welcome SMS to member
    background_tasks.add_task(SMSService.send_registration_sms, {"mobile": member.mobile, "member_id": member_id})
    
    # Send welcome email to member
    if member.email:
        background_tasks.add_task(EmailService.send_welcome_email, {"email": member.email, "name": member.name, "member_id": member_id})
    
    # Send membership notification to admin@bitzclub.com
    background_tasks.add_task(EmailService.send_membership_notification, member_doc, plan)
    
    return {**member_doc, "temporary_password": password, "_id": None}

class OfflineMemberCreate(BaseModel):
    """Model for admin-created offline members with extended fields"""
    name: str
    mobile: str
    country_code: str = "+91"
    email: str
    password: Optional[str] = None
    date_of_birth: Optional[str] = None
    address: Optional[str] = None
    area: Optional[str] = None
    city: Optional[str] = None
    state: Optional[str] = None
    country: str = "India"
    pincode: Optional[str] = None
    plan_id: str
    referral_id: Optional[str] = None
    source: str = "offline_admin"
    gender: Optional[str] = None
    phone_residence: Optional[str] = None

@api_router.post("/admin/members/offline")
async def create_offline_member(
    member: OfflineMemberCreate, 
    background_tasks: BackgroundTasks, 
    admin: dict = Depends(require_admin)
):
    """Admin endpoint to create members via offline registration with full details"""
    logger.info(f"[OFFLINE MEMBER] Creating offline member: {member.mobile}")
    
    # Check if mobile already exists
    existing = await db.members.find_one({"mobile": member.mobile})
    if existing:
        raise HTTPException(status_code=400, detail="Mobile number already registered")
    
    # Check email uniqueness
    existing_email = await db.members.find_one({"email": member.email})
    if existing_email:
        raise HTTPException(status_code=400, detail="Email already registered")
    
    # Get plan details
    plan = await db.plans.find_one({"id": member.plan_id}, {"_id": 0})
    if not plan:
        raise HTTPException(status_code=404, detail="Plan not found")
    
    # Generate Julian member ID
    member_id = await generate_member_id()
    user_id = str(uuid.uuid4())
    
    # Create user account
    password = member.password or ''.join(random.choices(string.ascii_letters + string.digits, k=8))
    hashed_password = pwd_context.hash(password)
    
    user_doc = {
        "id": user_id,
        "member_id": member_id,
        "mobile": member.mobile,
        "password": hashed_password,
        "name": member.name,
        "email": member.email,
        "role": UserRole.MEMBER,
        "is_active": True,
        "created_at": datetime.now(timezone.utc).isoformat()
    }
    await db.users.insert_one(user_doc)
    
    # Calculate membership dates
    start_date = datetime.now(timezone.utc)
    end_date = start_date + timedelta(days=plan["duration_months"] * 30)
    
    member_doc = {
        "id": str(uuid.uuid4()),
        "user_id": user_id,
        "member_id": member_id,
        "name": member.name,
        "mobile": member.mobile,
        "country_code": member.country_code,
        "email": member.email,
        "address": member.address,
        "area": member.area,
        "city": member.city,
        "state": member.state,
        "country": member.country,
        "pincode": member.pincode,
        "date_of_birth": member.date_of_birth,
        "gender": member.gender,
        "phone_residence": member.phone_residence,
        "plan_id": member.plan_id,
        "plan_name": plan["name"],
        "membership_start": start_date.isoformat(),
        "membership_end": end_date.isoformat(),
        "status": MembershipStatus.PENDING,  # Will be activated after payment confirmation
        "qr_code": generate_qr_code(member_id),
        "referral_id": member.referral_id,
        "source": member.source,
        "created_at": datetime.now(timezone.utc).isoformat(),
        "created_by": admin.get("id"),
        "created_by_role": "admin"
    }
    
    await db.members.insert_one(member_doc)
    logger.info(f"[OFFLINE MEMBER] Created member {member_id} by admin {admin.get('email')}")
    
    return {
        "success": True,
        "member_id": member_id,
        "user_id": user_id,
        "temporary_password": password,
        "message": f"Member {member_id} created successfully"
    }

@api_router.get("/members")
async def get_members(
    status: Optional[str] = None,
    plan_id: Optional[str] = None,
    search: Optional[str] = None,
    telecaller_id: Optional[str] = None,
    referral_id: Optional[str] = None,
    page: int = Query(1, ge=1),
    limit: int = Query(20, ge=1, le=100),
    user: dict = Depends(require_admin_or_telecaller)
):
    query = {}
    
    # Telecallers can only see their assigned members
    if user["role"] == UserRole.TELECALLER:
        query["assigned_telecaller"] = user["id"]
    elif telecaller_id:
        query["assigned_telecaller"] = telecaller_id
    
    if status:
        query["status"] = status
    if plan_id:
        query["plan_id"] = plan_id
    if referral_id:
        query["referral_id"] = {"$regex": referral_id, "$options": "i"}
    if search:
        query["$or"] = [
            {"name": {"$regex": search, "$options": "i"}},
            {"mobile": {"$regex": search, "$options": "i"}},
            {"member_id": {"$regex": search, "$options": "i"}},
            {"email": {"$regex": search, "$options": "i"}},
            {"referral_id": {"$regex": search, "$options": "i"}}
        ]
    
    skip = (page - 1) * limit
    total = await db.members.count_documents(query)
    members = await db.members.find(query, {"_id": 0}).skip(skip).limit(limit).to_list(limit)
    
    return {
        "members": members,
        "total": total,
        "page": page,
        "limit": limit,
        "pages": (total + limit - 1) // limit
    }

@api_router.get("/members/{member_id}")
async def get_member(member_id: str, user: dict = Depends(get_current_user)):
    member = await db.members.find_one(
        {"$or": [{"id": member_id}, {"member_id": member_id}]},
        {"_id": 0}
    )
    if not member:
        raise HTTPException(status_code=404, detail="Member not found")
    
    # Members can only view their own profile
    if user["role"] == UserRole.MEMBER and member.get("user_id") != user["id"]:
        raise HTTPException(status_code=403, detail="Access denied")
    
    # Telecallers can only view assigned members
    if user["role"] == UserRole.TELECALLER and member.get("assigned_telecaller") != user["id"]:
        raise HTTPException(status_code=403, detail="Access denied")
    
    return member

@api_router.put("/members/{member_id}")
async def update_member(member_id: str, member: MemberUpdate, user: dict = Depends(require_admin_or_telecaller)):
    update_data = {k: v for k, v in member.model_dump().items() if v is not None}
    if not update_data:
        raise HTTPException(status_code=400, detail="No data to update")
    
    # If plan is updated, update plan_name too
    if "plan_id" in update_data:
        plan = await db.plans.find_one({"id": update_data["plan_id"]}, {"_id": 0})
        if plan:
            update_data["plan_name"] = plan["name"]
    
    update_data["updated_at"] = datetime.now(timezone.utc).isoformat()
    result = await db.members.update_one({"id": member_id}, {"$set": update_data})
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Member not found")
    
    return await db.members.find_one({"id": member_id}, {"_id": 0})

@api_router.delete("/members/{member_id}")
async def delete_member(member_id: str, admin: dict = Depends(require_admin)):
    member = await db.members.find_one({"id": member_id})
    if not member:
        raise HTTPException(status_code=404, detail="Member not found")
    
    # Delete user account
    await db.users.delete_one({"id": member["user_id"]})
    await db.members.delete_one({"id": member_id})
    
    return {"message": "Member deleted"}

@api_router.post("/members/{member_id}/assign-telecaller")
async def assign_telecaller(member_id: str, telecaller_id: str, admin: dict = Depends(require_admin)):
    result = await db.members.update_one(
        {"id": member_id},
        {"$set": {"assigned_telecaller": telecaller_id, "updated_at": datetime.now(timezone.utc).isoformat()}}
    )
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Member not found")
    return {"message": "Telecaller assigned"}


@api_router.get("/members/{member_id}/payments")
async def get_member_payments(member_id: str, user: dict = Depends(get_current_user)):
    """Get payment history for a specific member. Members can see their own payments, admins/telecallers can see any."""
    # Check access - members can only see their own payments
    if user.get("role") == UserRole.MEMBER:
        if user.get("member_id") != member_id:
            raise HTTPException(status_code=403, detail="You can only view your own payments")
    elif user.get("role") not in [UserRole.ADMIN, UserRole.SUPER_ADMIN, UserRole.TELECALLER]:
        raise HTTPException(status_code=403, detail="Access denied")
    
    payments = await db.payments.find(
        {"member_id": member_id},
        {"_id": 0}
    ).sort("created_at", -1).to_list(100)
    return payments

class AssignTelecallerRequest(BaseModel):
    telecaller_id: Optional[str] = None

@api_router.put("/members/{member_id}/assign")
async def assign_member_telecaller(member_id: str, request: AssignTelecallerRequest, admin: dict = Depends(require_admin)):
    result = await db.members.update_one(
        {"member_id": member_id},
        {"$set": {"assigned_telecaller": request.telecaller_id, "updated_at": datetime.now(timezone.utc).isoformat()}}
    )
    if result.matched_count == 0:
        # Try with id field
        result = await db.members.update_one(
            {"id": member_id},
            {"$set": {"assigned_telecaller": request.telecaller_id, "updated_at": datetime.now(timezone.utc).isoformat()}}
        )
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Member not found")
    return {"message": "Telecaller assigned successfully"}


# ==================== MEMBER PHOTO UPLOAD ====================

UPLOAD_DIR = Path(__file__).parent / "uploads" / "photos"
UPLOAD_DIR.mkdir(parents=True, exist_ok=True)

@api_router.post("/members/{member_id}/photo")
async def upload_member_photo(member_id: str, file: UploadFile = File(...), user: dict = Depends(get_current_user)):
    """Upload a photo for a member"""
    # Validate file type
    allowed_types = ["image/jpeg", "image/png", "image/webp", "image/jpg"]
    if file.content_type not in allowed_types:
        raise HTTPException(status_code=400, detail="Invalid file type. Only JPEG, PNG, and WebP images are allowed.")
    
    # Check file size (max 5MB)
    contents = await file.read()
    if len(contents) > 5 * 1024 * 1024:
        raise HTTPException(status_code=400, detail="File size exceeds 5MB limit")
    
    # Find member
    member = await db.members.find_one({"$or": [{"id": member_id}, {"member_id": member_id}]})
    if not member:
        raise HTTPException(status_code=404, detail="Member not found")
    
    # Members can only upload their own photo
    if user["role"] == UserRole.MEMBER and member.get("user_id") != user["id"]:
        raise HTTPException(status_code=403, detail="Access denied")
    
    # Generate unique filename
    file_ext = file.filename.split(".")[-1] if "." in file.filename else "jpg"
    filename = f"{member['member_id']}_{uuid.uuid4().hex[:8]}.{file_ext}"
    file_path = UPLOAD_DIR / filename
    
    # Save file
    with open(file_path, "wb") as f:
        f.write(contents)
    
    # Generate photo URL
    photo_url = f"/api/uploads/photos/{filename}"
    
    # Update member record
    await db.members.update_one(
        {"id": member["id"]},
        {"$set": {"photo_url": photo_url, "updated_at": datetime.now(timezone.utc).isoformat()}}
    )
    
    logger.info(f"[PHOTO] Uploaded photo for member {member['member_id']}: {filename}")
    
    return {"photo_url": photo_url, "message": "Photo uploaded successfully"}

@api_router.get("/uploads/photos/{filename}")
async def get_member_photo(filename: str):
    """Serve uploaded member photos"""
    file_path = UPLOAD_DIR / filename
    if not file_path.exists():
        raise HTTPException(status_code=404, detail="Photo not found")
    
    # Determine content type
    ext = filename.split(".")[-1].lower()
    content_types = {
        "jpg": "image/jpeg",
        "jpeg": "image/jpeg",
        "png": "image/png",
        "webp": "image/webp"
    }
    content_type = content_types.get(ext, "image/jpeg")
    
    return StreamingResponse(
        open(file_path, "rb"),
        media_type=content_type,
        headers={"Cache-Control": "public, max-age=31536000"}
    )


# ==================== MEMBER VERIFICATION (PUBLIC) ====================

@api_router.get("/members/{member_id}/card")
async def get_member_card(member_id: str, user: dict = Depends(get_current_user)):
    """Get membership card data with QR code for download"""
    member = await db.members.find_one(
        {"$or": [{"id": member_id}, {"member_id": member_id}]}, 
        {"_id": 0}
    )
    if not member:
        raise HTTPException(status_code=404, detail="Member not found")
    
    # Members can only access their own card
    if user["role"] == UserRole.MEMBER and member.get("user_id") != user["id"]:
        raise HTTPException(status_code=403, detail="Access denied")
    
    # Generate QR code with verification URL
    qr_data = f"https://thebitzclub.com/verify/{member['member_id']}"
    qr_base64 = generate_qr_code(qr_data)
    
    return {
        "member_id": member["member_id"],
        "name": member["name"],
        "email": member.get("email", ""),
        "mobile": member.get("mobile", ""),
        "plan_name": member.get("plan_name", ""),
        "status": member["status"],
        "photo_url": member.get("photo_url"),
        "membership_start": member.get("membership_start"),
        "membership_end": member.get("membership_end"),
        "qr_code": f"data:image/png;base64,{qr_base64}",
        "verification_url": qr_data
    }

@api_router.get("/members/{member_id}/qr")
async def get_member_qr(member_id: str, user: dict = Depends(get_current_user)):
    """Get QR code for member verification"""
    member = await db.members.find_one(
        {"$or": [{"id": member_id}, {"member_id": member_id}]}, 
        {"_id": 0}
    )
    if not member:
        raise HTTPException(status_code=404, detail="Member not found")
    
    # Members can only access their own QR
    if user["role"] == UserRole.MEMBER and member.get("user_id") != user["id"]:
        raise HTTPException(status_code=403, detail="Access denied")
    
    # Generate QR code
    qr_data = f"https://thebitzclub.com/verify/{member['member_id']}"
    qr_base64 = generate_qr_code(qr_data)
    
    return {
        "member_id": member["member_id"],
        "qr_code": f"data:image/png;base64,{qr_base64}",
        "verification_url": qr_data
    }

@api_router.get("/verify/{member_id}")
async def verify_member(member_id: str):
    """Public endpoint for QR code verification"""
    member = await db.members.find_one({"member_id": member_id}, {"_id": 0})
    if not member:
        raise HTTPException(status_code=404, detail="Member not found")
    
    # Check validity
    end_date = datetime.fromisoformat(member["membership_end"].replace("Z", "+00:00"))
    is_valid = end_date > datetime.now(timezone.utc) and member["status"] == MembershipStatus.ACTIVE
    
    return {
        "member_id": member["member_id"],
        "name": member["name"],
        "plan_name": member["plan_name"],
        "membership_end": member["membership_end"],
        "status": member["status"],
        "is_valid": is_valid
    }

# ==================== FAMILY MEMBERS ENDPOINTS ====================

@api_router.post("/members/{member_id}/family")
async def add_family_member(member_id: str, family: FamilyMemberCreate, admin: dict = Depends(require_admin)):
    """Add a family member to a primary member's account"""
    # Verify primary member exists
    member = await db.members.find_one({"member_id": member_id}, {"_id": 0})
    if not member:
        raise HTTPException(status_code=404, detail="Primary member not found")
    
    family_id = str(uuid.uuid4())
    family_doc = {
        "id": family_id,
        "member_id": member_id,  # Link to primary member
        "primary_member_name": member["name"],
        "name": family.name,
        "relationship": family.relationship,
        "date_of_birth": family.date_of_birth,
        "mobile": family.mobile,
        "email": family.email,
        "id_proof_type": family.id_proof_type,
        "id_proof_number": family.id_proof_number,
        "photo_url": family.photo_url,
        "is_active": True,
        "created_at": datetime.now(timezone.utc).isoformat(),
        "created_by": admin["id"]
    }
    await db.family_members.insert_one(family_doc)
    
    # Update member's family count
    family_count = await db.family_members.count_documents({"member_id": member_id, "is_active": True})
    await db.members.update_one(
        {"member_id": member_id},
        {"$set": {"family_count": family_count, "updated_at": datetime.now(timezone.utc).isoformat()}}
    )
    
    logger.info(f"[FAMILY] Added family member {family.name} ({family.relationship}) for {member_id}")
    return {"message": "Family member added successfully", "id": family_id, "family_count": family_count}

@api_router.get("/members/{member_id}/family")
async def get_family_members(member_id: str, user: dict = Depends(require_admin_or_telecaller)):
    """Get all family members for a primary member"""
    members = await db.family_members.find(
        {"member_id": member_id},
        {"_id": 0}
    ).sort("created_at", 1).to_list(50)
    return {"member_id": member_id, "family_members": members, "count": len(members)}

@api_router.put("/family/{family_id}")
async def update_family_member(family_id: str, update: FamilyMemberUpdate, admin: dict = Depends(require_admin)):
    """Update a family member's details"""
    update_data = {k: v for k, v in update.dict().items() if v is not None}
    if not update_data:
        raise HTTPException(status_code=400, detail="No update data provided")
    
    update_data["updated_at"] = datetime.now(timezone.utc).isoformat()
    update_data["updated_by"] = admin["id"]
    
    result = await db.family_members.update_one({"id": family_id}, {"$set": update_data})
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Family member not found")
    
    return {"message": "Family member updated successfully"}

@api_router.delete("/family/{family_id}")
async def delete_family_member(family_id: str, admin: dict = Depends(require_admin)):
    """Soft delete a family member"""
    family = await db.family_members.find_one({"id": family_id})
    if not family:
        raise HTTPException(status_code=404, detail="Family member not found")
    
    # Soft delete
    await db.family_members.update_one(
        {"id": family_id},
        {"$set": {"is_active": False, "deleted_at": datetime.now(timezone.utc).isoformat(), "deleted_by": admin["id"]}}
    )
    
    # Update member's family count
    family_count = await db.family_members.count_documents({"member_id": family["member_id"], "is_active": True})
    await db.members.update_one(
        {"member_id": family["member_id"]},
        {"$set": {"family_count": family_count}}
    )
    
    return {"message": "Family member removed successfully"}

@api_router.get("/family")
async def get_all_family_members(
    page: int = Query(1, ge=1),
    limit: int = Query(20, ge=1, le=100),
    member_id: Optional[str] = None,
    admin: dict = Depends(require_admin)
):
    """Get all family members with pagination (Admin only)"""
    query = {"is_active": True}
    if member_id:
        query["member_id"] = member_id
    
    skip = (page - 1) * limit
    total = await db.family_members.count_documents(query)
    members = await db.family_members.find(query, {"_id": 0}).sort("created_at", -1).skip(skip).limit(limit).to_list(limit)
    
    return {
        "family_members": members,
        "total": total,
        "page": page,
        "limit": limit,
        "pages": (total + limit - 1) // limit
    }

# ==================== PARTNERS ENDPOINTS ====================

@api_router.post("/partners")
async def create_partner(partner: PartnerCreate, admin: dict = Depends(require_admin)):
    partner_id = str(uuid.uuid4())
    partner_doc = {
        "id": partner_id,
        "name": partner.name,
        "description": partner.description,
        "logo_url": partner.logo_url,
        "contact_email": partner.contact_email,
        "contact_phone": partner.contact_phone,
        "address": partner.address,
        "facilities": [f.model_dump() for f in partner.facilities],
        "is_active": partner.is_active,
        "created_at": datetime.now(timezone.utc).isoformat(),
        "created_by": admin["id"]
    }
    await db.partners.insert_one(partner_doc)
    return {k: v for k, v in partner_doc.items() if k != "_id"}

@api_router.get("/partners")
async def get_partners(is_active: Optional[bool] = None):
    query = {}
    if is_active is not None:
        query["is_active"] = is_active
    partners = await db.partners.find(query, {"_id": 0}).to_list(100)
    return partners

@api_router.get("/partners/{partner_id}")
async def get_partner(partner_id: str):
    partner = await db.partners.find_one({"id": partner_id}, {"_id": 0})
    if not partner:
        raise HTTPException(status_code=404, detail="Partner not found")
    return partner

@api_router.put("/partners/{partner_id}")
async def update_partner(partner_id: str, partner: PartnerUpdate, admin: dict = Depends(require_admin)):
    update_data = {}
    for k, v in partner.model_dump().items():
        if v is not None:
            if k == "facilities":
                update_data[k] = [f if isinstance(f, dict) else f.model_dump() for f in v]
            else:
                update_data[k] = v
    
    if not update_data:
        raise HTTPException(status_code=400, detail="No data to update")
    
    update_data["updated_at"] = datetime.now(timezone.utc).isoformat()
    result = await db.partners.update_one({"id": partner_id}, {"$set": update_data})
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Partner not found")
    
    return await db.partners.find_one({"id": partner_id}, {"_id": 0})

@api_router.delete("/partners/{partner_id}")
async def delete_partner(partner_id: str, admin: dict = Depends(require_admin)):
    result = await db.partners.delete_one({"id": partner_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Partner not found")
    return {"message": "Partner deleted"}

# ==================== AFFILIATE BOOKINGS ENDPOINTS ====================

@api_router.post("/bookings")
async def create_booking(booking: BookingCreate, user: dict = Depends(get_current_user)):
    """Create a booking at an affiliate (for members)"""
    # Get member details
    member = await db.members.find_one({"user_id": user["id"]}, {"_id": 0})
    if not member:
        raise HTTPException(status_code=404, detail="Member not found")
    
    # Get affiliate/partner details
    affiliate = await db.partners.find_one({"id": booking.affiliate_id}, {"_id": 0})
    if not affiliate:
        raise HTTPException(status_code=404, detail="Affiliate not found")
    
    if not affiliate.get("booking_enabled", True):
        raise HTTPException(status_code=400, detail="Bookings not available for this affiliate")
    
    booking_id = str(uuid.uuid4())
    booking_doc = {
        "id": booking_id,
        "member_id": member["member_id"],
        "member_name": member["name"],
        "member_email": member.get("email"),
        "member_mobile": member.get("mobile"),
        "user_id": user["id"],
        "affiliate_id": booking.affiliate_id,
        "affiliate_name": affiliate["name"],
        "affiliate_offers": affiliate.get("offers"),
        "affiliate_contact_1": affiliate.get("contact_person_1"),
        "affiliate_contact_1_phone": affiliate.get("contact_person_1_phone"),
        "affiliate_phone": affiliate.get("contact_phone"),
        "affiliate_address": affiliate.get("address"),
        "affiliate_website": affiliate.get("website"),
        "booking_date": booking.booking_date,
        "notes": booking.notes,
        "status": "confirmed",
        "created_at": datetime.now(timezone.utc).isoformat()
    }
    await db.bookings.insert_one(booking_doc)
    
    logger.info(f"[BOOKING] Created booking {booking_id} for {member['member_id']} at {affiliate['name']}")
    
    return {
        "message": "Booking created successfully!",
        "booking_id": booking_id,
        "affiliate_name": affiliate["name"],
        "booking_date": booking.booking_date
    }

@api_router.get("/bookings")
async def get_my_bookings(user: dict = Depends(get_current_user)):
    """Get member's booking history"""
    bookings = await db.bookings.find(
        {"user_id": user["id"]},
        {"_id": 0}
    ).sort("created_at", -1).to_list(100)
    return bookings

@api_router.get("/bookings/{booking_id}")
async def get_booking_details(booking_id: str, user: dict = Depends(get_current_user)):
    """Get booking details"""
    booking = await db.bookings.find_one({"id": booking_id}, {"_id": 0})
    if not booking:
        raise HTTPException(status_code=404, detail="Booking not found")
    
    # Verify ownership or admin access
    if booking["user_id"] != user["id"] and user.get("role") not in [UserRole.ADMIN, UserRole.SUPER_ADMIN]:
        raise HTTPException(status_code=403, detail="Access denied")
    
    return booking

@api_router.get("/admin/bookings")
async def get_all_bookings(
    affiliate_id: Optional[str] = None,
    member_id: Optional[str] = None,
    status: Optional[str] = None,
    page: int = Query(1, ge=1),
    limit: int = Query(20, ge=1, le=100),
    admin: dict = Depends(require_admin)
):
    """Get all bookings for admin"""
    query = {}
    if affiliate_id:
        query["affiliate_id"] = affiliate_id
    if member_id:
        query["member_id"] = member_id
    if status:
        query["status"] = status
    
    skip = (page - 1) * limit
    total = await db.bookings.count_documents(query)
    bookings = await db.bookings.find(query, {"_id": 0}).sort("created_at", -1).skip(skip).limit(limit).to_list(limit)
    
    return {
        "bookings": bookings,
        "total": total,
        "page": page,
        "limit": limit,
        "pages": (total + limit - 1) // limit
    }

@api_router.put("/bookings/{booking_id}/status")
async def update_booking_status(booking_id: str, status: str, admin: dict = Depends(require_admin)):
    """Update booking status (admin only)"""
    result = await db.bookings.update_one(
        {"id": booking_id},
        {"$set": {"status": status, "updated_at": datetime.now(timezone.utc).isoformat()}}
    )
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Booking not found")
    return {"message": "Booking status updated"}

# ==================== TELECALLERS ENDPOINTS ====================

@api_router.post("/telecallers")
async def create_telecaller(telecaller: TelecallerCreate, admin: dict = Depends(require_admin)):
    # Check if mobile already exists
    existing = await db.users.find_one({"mobile": telecaller.mobile})
    if existing:
        raise HTTPException(status_code=400, detail="Mobile number already registered")
    
    user_id = str(uuid.uuid4())
    hashed_password = pwd_context.hash(telecaller.password)
    
    user_doc = {
        "id": user_id,
        "mobile": telecaller.mobile,
        "password": hashed_password,
        "name": telecaller.name,
        "email": telecaller.email,
        "role": UserRole.TELECALLER,
        "is_active": True,
        "created_at": datetime.now(timezone.utc).isoformat(),
        "created_by": admin["id"]
    }
    await db.users.insert_one(user_doc)
    
    return {k: v for k, v in user_doc.items() if k not in ["password", "_id"]}

@api_router.get("/telecallers")
async def get_telecallers(admin: dict = Depends(require_admin)):
    telecallers = await db.users.find(
        {"role": UserRole.TELECALLER},
        {"_id": 0, "password": 0}
    ).to_list(100)
    
    # Get assigned member counts using aggregation (optimized)
    pipeline = [
        {"$group": {"_id": "$assigned_telecaller", "count": {"$sum": 1}}}
    ]
    counts_cursor = db.members.aggregate(pipeline)
    counts = {doc["_id"]: doc["count"] async for doc in counts_cursor}
    
    for tc in telecallers:
        tc["assigned_count"] = counts.get(tc["id"], 0)
    
    return telecallers

@api_router.get("/telecallers/{telecaller_id}")
async def get_telecaller(telecaller_id: str, admin: dict = Depends(require_admin)):
    telecaller = await db.users.find_one(
        {"id": telecaller_id, "role": UserRole.TELECALLER},
        {"_id": 0, "password": 0}
    )
    if not telecaller:
        raise HTTPException(status_code=404, detail="Telecaller not found")
    
    telecaller["assigned_count"] = await db.members.count_documents({"assigned_telecaller": telecaller_id})
    return telecaller

@api_router.put("/telecallers/{telecaller_id}")
async def update_telecaller(telecaller_id: str, telecaller: TelecallerUpdate, admin: dict = Depends(require_admin)):
    update_data = {k: v for k, v in telecaller.model_dump().items() if v is not None}
    if not update_data:
        raise HTTPException(status_code=400, detail="No data to update")
    
    update_data["updated_at"] = datetime.now(timezone.utc).isoformat()
    result = await db.users.update_one(
        {"id": telecaller_id, "role": UserRole.TELECALLER},
        {"$set": update_data}
    )
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Telecaller not found")
    
    return await db.users.find_one({"id": telecaller_id}, {"_id": 0, "password": 0})

@api_router.delete("/telecallers/{telecaller_id}")
async def delete_telecaller(telecaller_id: str, admin: dict = Depends(require_admin)):
    # Unassign all members first
    await db.members.update_many(
        {"assigned_telecaller": telecaller_id},
        {"$set": {"assigned_telecaller": None}}
    )
    
    result = await db.users.delete_one({"id": telecaller_id, "role": UserRole.TELECALLER})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Telecaller not found")
    return {"message": "Telecaller deleted"}

# ==================== TELECALLER DASHBOARD ENDPOINTS ====================

@api_router.get("/telecaller/dashboard")
async def get_telecaller_dashboard(user: dict = Depends(require_admin_or_telecaller)):
    """Get telecaller dashboard stats"""
    if user["role"] != UserRole.TELECALLER:
        raise HTTPException(status_code=403, detail="This endpoint is for telecallers only")
    
    telecaller_id = user["id"]
    
    # Get assigned members count
    assigned_members = await db.members.count_documents({"assigned_telecaller": telecaller_id})
    
    # Get active members count
    active_members = await db.members.count_documents({
        "assigned_telecaller": telecaller_id,
        "status": MembershipStatus.ACTIVE
    })
    
    # Get assigned leads count
    assigned_leads = await db.leads.count_documents({"assigned_telecaller": telecaller_id})
    new_leads = await db.leads.count_documents({"assigned_telecaller": telecaller_id, "status": LeadStatus.NEW})
    
    # Get pending follow-ups
    pending_followups = await db.follow_ups.count_documents({
        "telecaller_id": telecaller_id,
        "status": {"$in": ["pending", "scheduled"]}
    })
    
    return {
        "assigned_members": assigned_members,
        "active_members": active_members,
        "assigned_leads": assigned_leads,
        "new_leads": new_leads,
        "pending_followups": pending_followups,
        "telecaller_name": user["name"]
    }

@api_router.get("/telecaller/members")
async def get_telecaller_members(
    status: Optional[str] = None,
    search: Optional[str] = None,
    page: int = Query(1, ge=1),
    limit: int = Query(20, ge=1, le=100),
    user: dict = Depends(require_admin_or_telecaller)
):
    """Get members assigned to telecaller with essential details"""
    if user["role"] != UserRole.TELECALLER:
        raise HTTPException(status_code=403, detail="This endpoint is for telecallers only")
    
    query = {"assigned_telecaller": user["id"]}
    if status:
        query["status"] = status
    if search:
        query["$or"] = [
            {"name": {"$regex": search, "$options": "i"}},
            {"mobile": {"$regex": search, "$options": "i"}},
            {"member_id": {"$regex": search, "$options": "i"}}
        ]
    
    skip = (page - 1) * limit
    total = await db.members.count_documents(query)
    
    # Get members with only essential fields for telecaller
    members = await db.members.find(
        query,
        {
            "_id": 0,
            "member_id": 1,
            "name": 1,
            "mobile": 1,
            "email": 1,
            "plan_name": 1,
            "membership_start": 1,
            "membership_end": 1,
            "status": 1,
            "city": 1
        }
    ).sort("created_at", -1).skip(skip).limit(limit).to_list(limit)
    
    return {
        "members": members,
        "total": total,
        "page": page,
        "limit": limit,
        "pages": (total + limit - 1) // limit
    }

@api_router.get("/telecaller/members/{member_id}/details")
async def get_telecaller_member_details(member_id: str, user: dict = Depends(require_admin_or_telecaller)):
    """Get detailed member info including payment history (for telecaller)"""
    if user["role"] == UserRole.TELECALLER:
        # Verify member is assigned to this telecaller
        member = await db.members.find_one({
            "member_id": member_id,
            "assigned_telecaller": user["id"]
        }, {"_id": 0})
        if not member:
            raise HTTPException(status_code=404, detail="Member not found or not assigned to you")
    else:
        member = await db.members.find_one({"member_id": member_id}, {"_id": 0})
        if not member:
            raise HTTPException(status_code=404, detail="Member not found")
    
    # Get payment history
    payments = await db.payments.find(
        {"member_id": member_id},
        {"_id": 0}
    ).sort("created_at", -1).limit(20).to_list(20)
    
    # Get maintenance fees
    maintenance = await db.maintenance_fees.find(
        {"member_id": member_id},
        {"_id": 0}
    ).sort("due_date", -1).limit(10).to_list(10)
    
    # Get family members count
    family_count = await db.family_members.count_documents({"member_id": member_id, "is_active": True})
    
    # Get follow-ups
    follow_ups = await db.follow_ups.find(
        {"member_id": member_id},
        {"_id": 0}
    ).sort("created_at", -1).limit(10).to_list(10)
    
    return {
        "member": member,
        "payment_history": payments,
        "maintenance_fees": maintenance,
        "family_count": family_count,
        "follow_ups": follow_ups
    }

# ==================== FOLLOW-UPS ENDPOINTS ====================

@api_router.post("/follow-ups")
async def create_follow_up(follow_up: FollowUpCreate, user: dict = Depends(require_admin_or_telecaller)):
    follow_up_id = str(uuid.uuid4())
    follow_up_doc = {
        "id": follow_up_id,
        **follow_up.model_dump(),
        "telecaller_id": user["id"],
        "created_at": datetime.now(timezone.utc).isoformat()
    }
    await db.follow_ups.insert_one(follow_up_doc)
    return {k: v for k, v in follow_up_doc.items() if k != "_id"}

@api_router.get("/follow-ups")
async def get_follow_ups(
    member_id: Optional[str] = None,
    status: Optional[str] = None,
    user: dict = Depends(require_admin_or_telecaller)
):
    query = {}
    if user["role"] == UserRole.TELECALLER:
        query["telecaller_id"] = user["id"]
    if member_id:
        query["member_id"] = member_id
    if status:
        query["status"] = status
    
    follow_ups = await db.follow_ups.find(query, {"_id": 0}).to_list(100)
    return follow_ups

@api_router.put("/follow-ups/{follow_up_id}")
async def update_follow_up(follow_up_id: str, follow_up: FollowUpUpdate, user: dict = Depends(require_admin_or_telecaller)):
    update_data = {k: v for k, v in follow_up.model_dump().items() if v is not None}
    if not update_data:
        raise HTTPException(status_code=400, detail="No data to update")
    
    update_data["updated_at"] = datetime.now(timezone.utc).isoformat()
    result = await db.follow_ups.update_one({"id": follow_up_id}, {"$set": update_data})
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Follow-up not found")
    
    return await db.follow_ups.find_one({"id": follow_up_id}, {"_id": 0})

# ==================== PAYMENTS ENDPOINTS ====================

@api_router.post("/payments/create-order")
async def create_payment_order(payment: PaymentCreate, user: dict = Depends(get_current_user)):
    """Create a Razorpay order for online payment"""
    # Get member info
    member = await db.members.find_one({"member_id": payment.member_id}, {"_id": 0})
    if not member:
        # Try finding by id
        member = await db.members.find_one({"id": payment.member_id}, {"_id": 0})
    
    # Get plan info
    plan = await db.plans.find_one({"id": payment.plan_id}, {"_id": 0}) if payment.plan_id else None
    
    # Create Razorpay order
    notes = {
        "member_id": payment.member_id,
        "member_name": member.get("name", "") if member else "",
        "plan_id": payment.plan_id or "",
        "plan_name": plan.get("name", "") if plan else ""
    }
    
    order = await RazorpayService.create_order(
        amount=payment.amount,
        member_id=payment.member_id,
        plan_id=payment.plan_id,
        notes=notes
    )
    
    # Store payment record
    payment_doc = {
        "id": str(uuid.uuid4()),
        "order_id": order["id"],
        "member_id": payment.member_id,
        "member_name": member.get("name", "") if member else "",
        "plan_id": payment.plan_id,
        "plan_name": plan.get("name", "") if plan else "",
        "amount": payment.amount,
        "payment_type": "online",
        "payment_method": "razorpay",
        "status": "pending",
        "created_at": datetime.now(timezone.utc).isoformat()
    }
    await db.payments.insert_one(payment_doc)
    
    return {
        "order_id": order["id"],
        "amount": order["amount"],
        "currency": order["currency"],
        "razorpay_key": order["razorpay_key"],
        "payment_id": payment_doc["id"],
        "member_name": member.get("name", "") if member else "",
        "member_email": member.get("email", "") if member else "",
        "member_mobile": member.get("mobile", "") if member else ""
    }

@api_router.post("/payments/verify")
async def verify_payment(
    payment_id: str,
    razorpay_payment_id: str,
    razorpay_order_id: str,
    razorpay_signature: str,
    background_tasks: BackgroundTasks,
    user: dict = Depends(get_current_user)
):
    """Verify Razorpay payment signature and complete the payment"""
    # Verify payment signature
    is_valid = RazorpayService.verify_payment_signature(
        razorpay_payment_id, razorpay_order_id, razorpay_signature
    )
    
    if not is_valid:
        # Update payment status to failed
        await db.payments.update_one(
            {"id": payment_id},
            {"$set": {"status": "failed", "failed_at": datetime.now(timezone.utc).isoformat()}}
        )
        raise HTTPException(status_code=400, detail="Payment verification failed - Invalid signature")
    
    # Get payment record
    payment = await db.payments.find_one({"id": payment_id})
    if not payment:
        raise HTTPException(status_code=404, detail="Payment not found")
    
    # Fetch payment details from Razorpay for additional verification
    razorpay_payment = await RazorpayService.fetch_payment(razorpay_payment_id)
    
    # Update payment status
    await db.payments.update_one(
        {"id": payment_id},
        {"$set": {
            "status": "completed",
            "razorpay_payment_id": razorpay_payment_id,
            "razorpay_order_id": razorpay_order_id,
            "transaction_id": razorpay_payment_id,
            "razorpay_details": {
                "method": razorpay_payment.get("method") if razorpay_payment else None,
                "bank": razorpay_payment.get("bank") if razorpay_payment else None,
                "wallet": razorpay_payment.get("wallet") if razorpay_payment else None,
                "vpa": razorpay_payment.get("vpa") if razorpay_payment else None
            },
            "completed_at": datetime.now(timezone.utc).isoformat()
        }}
    )
    
    # Get member and update membership status
    member = await db.members.find_one({"member_id": payment["member_id"]})
    if not member:
        member = await db.members.find_one({"id": payment["member_id"]})
    
    if member:
        plan = await db.plans.find_one({"id": payment.get("plan_id")})
        duration_days = (plan.get("duration_months", 1) * 30) if plan else 365
        end_date = datetime.now(timezone.utc) + timedelta(days=duration_days)
        
        await db.members.update_one(
            {"id": member["id"]},
            {"$set": {
                "status": MembershipStatus.ACTIVE,
                "plan_id": payment.get("plan_id"),
                "plan_name": plan.get("name") if plan else payment.get("plan_name"),
                "membership_start": datetime.now(timezone.utc).isoformat(),
                "membership_end": end_date.isoformat()
            }}
        )
        
        # Send SMS notifications
        plan_name = plan.get("name") if plan else "Membership"
        validity_end = end_date.strftime("%d %b %Y")
        
        # Payment success SMS
        background_tasks.add_task(
            SMSService.send_payment_success_sms, 
            member, 
            payment["amount"],
            plan_name
        )
        
        # Membership activation SMS
        background_tasks.add_task(
            SMSService.send_membership_activation_sms,
            member,
            plan_name,
            validity_end
        )
        
        # Email notification
        if member.get("email"):
            background_tasks.add_task(EmailService.send_payment_receipt, member, payment)
    
    return {
        "message": "Payment verified successfully",
        "status": "completed",
        "razorpay_payment_id": razorpay_payment_id,
        "member_id": payment["member_id"]
    }

@api_router.post("/payments/offline")
async def record_offline_payment(
    payment: PaymentCreate,
    background_tasks: BackgroundTasks,
    admin: dict = Depends(require_admin)
):
    """Record an offline/cash payment (Admin only)"""
    # Get member info
    member = await db.members.find_one({"member_id": payment.member_id}, {"_id": 0})
    if not member:
        member = await db.members.find_one({"id": payment.member_id}, {"_id": 0})
    
    if not member:
        raise HTTPException(status_code=404, detail="Member not found")
    
    # Get plan info
    plan = await db.plans.find_one({"id": payment.plan_id}, {"_id": 0}) if payment.plan_id else None
    
    # Create offline payment record
    payment_doc = {
        "id": str(uuid.uuid4()),
        "order_id": f"OFFLINE-{uuid.uuid4().hex[:12].upper()}",
        "member_id": payment.member_id,
        "member_name": member.get("name", ""),
        "plan_id": payment.plan_id,
        "plan_name": plan.get("name", "") if plan else "",
        "amount": payment.amount,
        "payment_type": "offline",
        "payment_method": payment.payment_method or "cash",
        "transaction_id": payment.transaction_id,
        "notes": payment.notes,
        "status": "completed",
        "created_at": datetime.now(timezone.utc).isoformat(),
        "completed_at": datetime.now(timezone.utc).isoformat(),
        "created_by": admin.get("id")
    }
    await db.payments.insert_one(payment_doc)
    
    # Activate membership
    if plan:
        duration_days = plan.get("duration_months", 1) * 30
        end_date = datetime.now(timezone.utc) + timedelta(days=duration_days)
        
        await db.members.update_one(
            {"id": member["id"]},
            {"$set": {
                "status": MembershipStatus.ACTIVE,
                "plan_id": payment.plan_id,
                "plan_name": plan.get("name"),
                "membership_start": datetime.now(timezone.utc).isoformat(),
                "membership_end": end_date.isoformat()
            }}
        )
        
        # Send SMS notifications
        validity_end = end_date.strftime("%d %b %Y")
        background_tasks.add_task(
            SMSService.send_payment_success_sms,
            member,
            payment.amount,
            plan.get("name", "Membership")
        )
        background_tasks.add_task(
            SMSService.send_membership_activation_sms,
            member,
            plan.get("name", "Membership"),
            validity_end
        )
    
    payment_doc.pop("_id", None)
    return {"message": "Offline payment recorded successfully", "payment": payment_doc}

@api_router.get("/payments")
async def get_payments(
    member_id: Optional[str] = None,
    status: Optional[str] = None,
    page: int = Query(1, ge=1),
    limit: int = Query(20, ge=1, le=100),
    user: dict = Depends(require_admin_or_telecaller)
):
    query = {}
    if member_id:
        query["member_id"] = member_id
    if status:
        query["status"] = status
    
    skip = (page - 1) * limit
    total = await db.payments.count_documents(query)
    payments = await db.payments.find(query, {"_id": 0}).skip(skip).limit(limit).to_list(limit)
    
    return {
        "payments": payments,
        "total": total,
        "page": page,
        "limit": limit
    }

# ==================== REPORTS ENDPOINTS ====================

@api_router.get("/reports/dashboard-stats")
async def get_dashboard_stats(admin: dict = Depends(require_admin)):
    """Dashboard summary stats"""
    today = datetime.now(timezone.utc).date()
    today_start = datetime.combine(today, datetime.min.time()).isoformat()
    
    total_members = await db.members.count_documents({})
    active_members = await db.members.count_documents({"status": MembershipStatus.ACTIVE})
    pending_members = await db.members.count_documents({"status": MembershipStatus.PENDING})
    expired_members = await db.members.count_documents({"status": MembershipStatus.EXPIRED})
    
    # Today's registrations
    today_registrations = await db.members.count_documents({
        "created_at": {"$gte": today_start}
    })
    
    # Revenue stats
    payments = await db.payments.find({"status": "completed"}, {"_id": 0}).to_list(10000)
    total_revenue = sum(p.get("amount", 0) for p in payments)
    
    # This month stats
    month_start = datetime.now(timezone.utc).replace(day=1, hour=0, minute=0, second=0, microsecond=0)
    month_members = await db.members.count_documents({
        "created_at": {"$gte": month_start.isoformat()}
    })
    month_payments = await db.payments.find({
        "status": "completed",
        "created_at": {"$gte": month_start.isoformat()}
    }, {"_id": 0}).to_list(10000)
    month_revenue = sum(p.get("amount", 0) for p in month_payments)
    
    # Plan distribution
    plan_stats = await db.members.aggregate([
        {"$group": {"_id": "$plan_name", "count": {"$sum": 1}}}
    ]).to_list(100)
    
    return {
        "total_members": total_members,
        "active_members": active_members,
        "pending_members": pending_members,
        "expired_members": expired_members,
        "today_registrations": today_registrations,
        "total_revenue": total_revenue,
        "month_members": month_members,
        "month_revenue": month_revenue,
        "plan_distribution": [{"plan": p["_id"], "count": p["count"]} for p in plan_stats if p["_id"]],
        "telecallers_count": await db.users.count_documents({"role": UserRole.TELECALLER}),
        "partners_count": await db.partners.count_documents({}),
        "leads_count": await db.leads.count_documents({})
    }

@api_router.get("/reports/members")
async def get_members_report(
    search: Optional[str] = None,
    name: Optional[str] = None,
    mobile: Optional[str] = None,
    member_id: Optional[str] = None,
    plan_id: Optional[str] = None,
    status: Optional[str] = None,
    city: Optional[str] = None,
    pincode: Optional[str] = None,
    area: Optional[str] = None,
    referral_id: Optional[str] = None,
    start_date: Optional[str] = None,
    end_date: Optional[str] = None,
    expiry_start: Optional[str] = None,
    expiry_end: Optional[str] = None,
    admin: dict = Depends(require_admin)
):
    """Member reports with comprehensive filters"""
    query = {}
    
    # Search across name, mobile, member_id
    if search:
        query["$or"] = [
            {"name": {"$regex": search, "$options": "i"}},
            {"mobile": {"$regex": search, "$options": "i"}},
            {"member_id": {"$regex": search, "$options": "i"}}
        ]
    
    # Individual filters
    if name:
        query["name"] = {"$regex": name, "$options": "i"}
    if mobile:
        query["mobile"] = {"$regex": mobile, "$options": "i"}
    if member_id:
        query["member_id"] = {"$regex": member_id, "$options": "i"}
    if plan_id:
        query["plan_id"] = plan_id
    if status:
        query["status"] = status
    if city:
        query["city"] = {"$regex": city, "$options": "i"}
    if pincode:
        query["pincode"] = pincode
    if area:
        query["area"] = {"$regex": area, "$options": "i"}
    if referral_id:
        query["referral_id"] = {"$regex": referral_id, "$options": "i"}
    
    # Registration date range
    if start_date or end_date:
        query["created_at"] = {}
        if start_date:
            query["created_at"]["$gte"] = start_date
        if end_date:
            query["created_at"]["$lte"] = end_date + "T23:59:59"
    
    # Expiry date range
    if expiry_start or expiry_end:
        query["membership_end"] = {}
        if expiry_start:
            query["membership_end"]["$gte"] = expiry_start
        if expiry_end:
            query["membership_end"]["$lte"] = expiry_end + "T23:59:59"
    
    members = await db.members.find(query, {"_id": 0, "qr_code": 0}).sort("created_at", -1).to_list(10000)
    return members

@api_router.get("/reports/payments")
async def get_payments_report(
    payment_type: Optional[str] = None,
    payment_method: Optional[str] = None,
    plan_id: Optional[str] = None,
    min_amount: Optional[float] = None,
    max_amount: Optional[float] = None,
    start_date: Optional[str] = None,
    end_date: Optional[str] = None,
    month: Optional[str] = None,
    year: Optional[str] = None,
    admin: dict = Depends(require_admin)
):
    """Payment reports with filters"""
    query = {"status": "completed"}
    
    if payment_type:
        query["payment_type"] = payment_type
    if payment_method:
        query["payment_method"] = payment_method
    if plan_id:
        query["plan_id"] = plan_id
    
    if min_amount is not None or max_amount is not None:
        query["amount"] = {}
        if min_amount is not None:
            query["amount"]["$gte"] = min_amount
        if max_amount is not None:
            query["amount"]["$lte"] = max_amount
    
    if start_date or end_date:
        query["created_at"] = {}
        if start_date:
            query["created_at"]["$gte"] = start_date
        if end_date:
            query["created_at"]["$lte"] = end_date + "T23:59:59"
    
    if month and year:
        month_start = f"{year}-{month.zfill(2)}-01"
        if int(month) == 12:
            month_end = f"{int(year)+1}-01-01"
        else:
            month_end = f"{year}-{str(int(month)+1).zfill(2)}-01"
        query["created_at"] = {"$gte": month_start, "$lt": month_end}
    elif year:
        query["created_at"] = {"$gte": f"{year}-01-01", "$lt": f"{int(year)+1}-01-01"}
    
    payments = await db.payments.find(query, {"_id": 0}).sort("created_at", -1).to_list(10000)
    
    # Calculate summary
    total_amount = sum(p.get("amount", 0) for p in payments)
    online_amount = sum(p.get("amount", 0) for p in payments if p.get("payment_type") == "online")
    offline_amount = sum(p.get("amount", 0) for p in payments if p.get("payment_type") == "offline")
    
    return {
        "payments": payments,
        "summary": {
            "total_count": len(payments),
            "total_amount": total_amount,
            "online_amount": online_amount,
            "offline_amount": offline_amount
        }
    }

@api_router.get("/reports/leads-dashboard")
async def get_leads_dashboard(
    start_date: Optional[str] = None,
    end_date: Optional[str] = None,
    admin: dict = Depends(require_admin)
):
    """Leads dashboard for marketing - source, city, conversions"""
    query = {}
    if start_date or end_date:
        date_filter = {}
        if start_date:
            date_filter["$gte"] = start_date
        if end_date:
            date_filter["$lte"] = end_date + "T23:59:59"
        query["created_at"] = date_filter
    
    # Total leads
    total_leads = await db.leads.count_documents(query)
    
    # Leads by source
    source_stats = await db.leads.aggregate([
        {"$match": query},
        {"$group": {"_id": "$source", "count": {"$sum": 1}}},
        {"$sort": {"count": -1}}
    ]).to_list(50)
    
    # Leads by city
    city_stats = await db.leads.aggregate([
        {"$match": {**query, "city": {"$exists": True, "$nin": [None, ""]}}},
        {"$group": {"_id": "$city", "count": {"$sum": 1}}},
        {"$sort": {"count": -1}},
        {"$limit": 20}
    ]).to_list(20)
    
    # Leads by member type
    member_type_stats = await db.leads.aggregate([
        {"$match": query},
        {"$group": {"_id": {"$ifNull": ["$member_type", "indian"]}, "count": {"$sum": 1}}},
        {"$sort": {"count": -1}}
    ]).to_list(10)
    
    # Leads by status (conversion tracking)
    status_stats = await db.leads.aggregate([
        {"$match": query},
        {"$group": {"_id": "$status", "count": {"$sum": 1}}},
        {"$sort": {"count": -1}}
    ]).to_list(20)
    
    # Calculate conversion rate
    converted = await db.leads.count_documents({**query, "status": "converted"})
    conversion_rate = (converted / total_leads * 100) if total_leads > 0 else 0
    
    # Daily leads trend (last 30 days)
    thirty_days_ago = (datetime.now(timezone.utc) - timedelta(days=30)).isoformat()
    daily_trend = await db.leads.aggregate([
        {"$match": {"created_at": {"$gte": thirty_days_ago}}},
        {"$project": {
            "date": {"$substr": ["$created_at", 0, 10]}
        }},
        {"$group": {"_id": "$date", "count": {"$sum": 1}}},
        {"$sort": {"_id": 1}}
    ]).to_list(30)
    
    return {
        "total_leads": total_leads,
        "converted_leads": converted,
        "conversion_rate": round(conversion_rate, 2),
        "by_source": [{"source": s["_id"] or "unknown", "count": s["count"]} for s in source_stats],
        "by_city": [{"city": c["_id"], "count": c["count"]} for c in city_stats],
        "by_member_type": [{"type": m["_id"], "count": m["count"]} for m in member_type_stats],
        "by_status": [{"status": s["_id"] or "new", "count": s["count"]} for s in status_stats],
        "daily_trend": [{"date": d["_id"], "count": d["count"]} for d in daily_trend]
    }

@api_router.get("/reports/location")
async def get_location_report(
    group_by: str = "city",
    admin: dict = Depends(require_admin)
):
    """Location-wise member count"""
    if group_by == "pincode":
        stats = await db.members.aggregate([
            {"$match": {"pincode": {"$exists": True, "$nin": [None, ""]}}},
            {"$group": {"_id": "$pincode", "count": {"$sum": 1}}},
            {"$sort": {"count": -1}}
        ]).to_list(1000)
    elif group_by == "area":
        stats = await db.members.aggregate([
            {"$match": {"area": {"$exists": True, "$nin": [None, ""]}}},
            {"$group": {"_id": "$area", "count": {"$sum": 1}}},
            {"$sort": {"count": -1}}
        ]).to_list(1000)
    else:  # city
        stats = await db.members.aggregate([
            {"$match": {"city": {"$exists": True, "$nin": [None, ""]}}},
            {"$group": {"_id": "$city", "count": {"$sum": 1}}},
            {"$sort": {"count": -1}}
        ]).to_list(1000)
    
    return [{"location": s["_id"], "count": s["count"]} for s in stats]

@api_router.get("/reports/telecaller-performance")
async def get_telecaller_performance(
    telecaller_id: Optional[str] = None,
    start_date: Optional[str] = None,
    end_date: Optional[str] = None,
    admin: dict = Depends(require_admin)
):
    """Telecaller performance report"""
    telecallers = await db.users.find({"role": UserRole.TELECALLER}, {"_id": 0, "password": 0}).to_list(100)
    
    result = []
    for tc in telecallers:
        if telecaller_id and tc.get("id") != telecaller_id:
            continue
        
        tc_id = tc.get("id")
        
        # Build query for date range
        lead_query = {"assigned_to": tc_id}
        member_query = {"created_by": tc_id}
        payment_query = {"created_by": tc_id}
        
        if start_date or end_date:
            date_filter = {}
            if start_date:
                date_filter["$gte"] = start_date
            if end_date:
                date_filter["$lte"] = end_date + "T23:59:59"
            lead_query["created_at"] = date_filter
            member_query["created_at"] = date_filter
            payment_query["created_at"] = date_filter
        
        # Leads stats
        leads_assigned = await db.leads.count_documents(lead_query)
        leads_converted = await db.leads.count_documents({**lead_query, "status": "converted"})
        leads_contacted = await db.leads.count_documents({**lead_query, "status": {"$in": ["contacted", "interested"]}})
        leads_pending = await db.leads.count_documents({**lead_query, "status": "new"})
        
        # Members created by telecaller
        members_created = await db.members.count_documents(member_query)
        
        # Payments collected by telecaller
        payments = await db.payments.find(payment_query, {"amount": 1}).to_list(10000)
        payments_collected = sum(p.get("amount", 0) for p in payments)
        payments_count = len(payments)
        
        conversion_rate = (leads_converted / leads_assigned * 100) if leads_assigned > 0 else 0
        
        result.append({
            "telecaller_id": tc_id,
            "telecaller_name": tc.get("name", "Unknown"),
            "telecaller_mobile": tc.get("mobile", ""),
            "leads_assigned": leads_assigned,
            "leads_contacted": leads_contacted,
            "leads_converted": leads_converted,
            "leads_pending": leads_pending,
            "members_created": members_created,
            "payments_count": payments_count,
            "payments_collected": payments_collected,
            "conversion_rate": round(conversion_rate, 2)
        })
    
    return result

@api_router.get("/reports/referral")
async def get_referral_report(
    referral_type: Optional[str] = None,
    start_date: Optional[str] = None,
    end_date: Optional[str] = None,
    admin: dict = Depends(require_admin)
):
    """Referral performance report"""
    query = {"referral_id": {"$exists": True, "$nin": [None, ""]}}
    
    if start_date or end_date:
        query["created_at"] = {}
        if start_date:
            query["created_at"]["$gte"] = start_date
        if end_date:
            query["created_at"]["$lte"] = end_date + "T23:59:59"
    
    # Get all members with referral IDs
    members = await db.members.find(query, {"_id": 0, "referral_id": 1, "name": 1, "member_id": 1, "created_at": 1, "plan_name": 1}).to_list(10000)
    
    # Group by referral ID
    referral_stats = {}
    for m in members:
        ref_id = m.get("referral_id", "")
        if not ref_id:
            continue
        
        # Determine type based on prefix
        ref_type = "unknown"
        if ref_id.startswith("BITZ-E"):
            ref_type = "employee"
        elif ref_id.startswith("BITZ-A"):
            ref_type = "associate"
        elif ref_id.startswith("BITZ-"):
            ref_type = "member"
        
        if referral_type and ref_type != referral_type:
            continue
        
        if ref_id not in referral_stats:
            referral_stats[ref_id] = {
                "referral_id": ref_id,
                "referral_type": ref_type,
                "members_referred": [],
                "count": 0
            }
        
        referral_stats[ref_id]["members_referred"].append({
            "member_id": m.get("member_id"),
            "name": m.get("name"),
            "plan": m.get("plan_name"),
            "joined_at": m.get("created_at")
        })
        referral_stats[ref_id]["count"] += 1
    
    # Sort by count
    result = sorted(referral_stats.values(), key=lambda x: x["count"], reverse=True)
    
    summary = {
        "total_referrals": len(members),
        "employee_referrals": sum(1 for m in members if m.get("referral_id", "").startswith("BITZ-E")),
        "associate_referrals": sum(1 for m in members if m.get("referral_id", "").startswith("BITZ-A")),
        "member_referrals": sum(1 for m in members if m.get("referral_id", "").startswith("BITZ-") and not m.get("referral_id", "").startswith("BITZ-E") and not m.get("referral_id", "").startswith("BITZ-A"))
    }
    
    return {"referrals": result, "summary": summary}

# ==================== MEMBER REFERRAL SYSTEM ====================

@api_router.get("/members/{member_id}/referral-stats")
async def get_member_referral_stats(member_id: str, user: dict = Depends(get_current_user)):
    """Get referral statistics for a specific member"""
    # Get the member
    member = await db.members.find_one(
        {"$or": [{"id": member_id}, {"member_id": member_id}]},
        {"_id": 0}
    )
    if not member:
        raise HTTPException(status_code=404, detail="Member not found")
    
    # Members can only view their own stats
    if user["role"] == UserRole.MEMBER and member.get("user_id") != user["id"]:
        raise HTTPException(status_code=403, detail="Access denied")
    
    referral_code = member.get("member_id")
    
    # Get all members referred by this member
    referred_members = await db.members.find(
        {"referred_by": referral_code},
        {"_id": 0, "member_id": 1, "name": 1, "status": 1, "plan_name": 1, "created_at": 1}
    ).sort("created_at", -1).to_list(100)
    
    # Calculate stats
    total_referred = len(referred_members)
    active_referrals = sum(1 for m in referred_members if m.get("status") == "active")
    pending_referrals = sum(1 for m in referred_members if m.get("status") == "pending")
    
    # Get referral rewards (if any)
    rewards = await db.referral_rewards.find(
        {"referrer_member_id": referral_code},
        {"_id": 0}
    ).sort("created_at", -1).to_list(50)
    
    total_rewards = sum(r.get("reward_amount", 0) for r in rewards)
    pending_rewards = sum(r.get("reward_amount", 0) for r in rewards if r.get("status") == "pending")
    claimed_rewards = sum(r.get("reward_amount", 0) for r in rewards if r.get("status") == "claimed")
    
    return {
        "referral_code": referral_code,
        "referral_link": f"https://thebitzclub.com/join?ref={referral_code}",
        "total_referred": total_referred,
        "active_referrals": active_referrals,
        "pending_referrals": pending_referrals,
        "referred_members": referred_members,
        "rewards": {
            "total": total_rewards,
            "pending": pending_rewards,
            "claimed": claimed_rewards,
            "history": rewards
        }
    }

@api_router.get("/referrals/validate")
async def validate_referral_code(referral_code: str = Query(...)):
    """Validate if a referral code exists and is valid"""
    # Check if it's a valid member ID
    member = await db.members.find_one(
        {"member_id": referral_code, "status": "active"},
        {"_id": 0, "member_id": 1, "name": 1}
    )
    
    if member:
        return {
            "valid": True,
            "type": "member",
            "referrer_name": member.get("name", "").split()[0] if member.get("name") else "Member",
            "message": f"Referral code valid! Referred by {member.get('name', '').split()[0]}"
        }
    
    # Check if it's an employee or associate code
    if referral_code.startswith("BITZ-E") or referral_code.startswith("BITZ-A"):
        return {
            "valid": True,
            "type": "employee" if referral_code.startswith("BITZ-E") else "associate",
            "referrer_name": "BITZ Team",
            "message": "Referral code valid!"
        }
    
    return {
        "valid": False,
        "type": None,
        "referrer_name": None,
        "message": "Invalid referral code"
    }

@api_router.get("/admin/referrals")
async def get_admin_referral_dashboard(
    page: int = Query(1, ge=1),
    limit: int = Query(20, ge=1, le=100),
    referral_type: Optional[str] = None,
    search: Optional[str] = None,
    admin: dict = Depends(require_admin)
):
    """Admin dashboard for referral management"""
    
    # Get all members who have referred someone
    pipeline = [
        {"$match": {"referred_by": {"$exists": True, "$nin": [None, ""]}}},
        {"$group": {
            "_id": "$referred_by",
            "count": {"$sum": 1},
            "active_count": {"$sum": {"$cond": [{"$eq": ["$status", "active"]}, 1, 0]}},
            "members": {"$push": {
                "member_id": "$member_id",
                "name": "$name",
                "status": "$status",
                "plan_name": "$plan_name",
                "created_at": "$created_at"
            }}
        }},
        {"$sort": {"count": -1}}
    ]
    
    referral_groups = await db.members.aggregate(pipeline).to_list(1000)
    
    # Enrich with referrer details
    result = []
    for group in referral_groups:
        referral_code = group["_id"]
        
        # Determine referral type
        ref_type = "member"
        if referral_code.startswith("BITZ-E"):
            ref_type = "employee"
        elif referral_code.startswith("BITZ-A"):
            ref_type = "associate"
        
        if referral_type and ref_type != referral_type:
            continue
        
        # Get referrer details if it's a member
        referrer = None
        if ref_type == "member":
            referrer = await db.members.find_one(
                {"member_id": referral_code},
                {"_id": 0, "name": 1, "mobile": 1, "email": 1, "member_id": 1}
            )
        
        if search:
            search_lower = search.lower()
            if referral_code.lower().find(search_lower) == -1:
                if not referrer or referrer.get("name", "").lower().find(search_lower) == -1:
                    continue
        
        result.append({
            "referral_code": referral_code,
            "referral_type": ref_type,
            "referrer": referrer,
            "total_referrals": group["count"],
            "active_referrals": group["active_count"],
            "recent_referrals": group["members"][:5]  # Last 5
        })
    
    # Pagination
    total = len(result)
    start = (page - 1) * limit
    end = start + limit
    paginated = result[start:end]
    
    # Summary stats
    summary = {
        "total_referrers": len(result),
        "total_referred_members": sum(r["total_referrals"] for r in result),
        "by_type": {
            "employee": sum(1 for r in result if r["referral_type"] == "employee"),
            "associate": sum(1 for r in result if r["referral_type"] == "associate"),
            "member": sum(1 for r in result if r["referral_type"] == "member")
        }
    }
    
    return {
        "referrals": paginated,
        "summary": summary,
        "total": total,
        "page": page,
        "limit": limit,
        "pages": (total + limit - 1) // limit
    }

@api_router.post("/admin/referral-rewards")
async def create_referral_reward(
    referrer_member_id: str,
    referred_member_id: str,
    reward_amount: float,
    reward_type: str = "cashback",  # cashback, discount, points
    notes: Optional[str] = None,
    admin: dict = Depends(require_admin)
):
    """Create a referral reward for a referrer"""
    reward = {
        "id": str(uuid.uuid4()),
        "referrer_member_id": referrer_member_id,
        "referred_member_id": referred_member_id,
        "reward_amount": reward_amount,
        "reward_type": reward_type,
        "status": "pending",  # pending, claimed, expired
        "notes": notes,
        "created_at": datetime.now(timezone.utc).isoformat(),
        "created_by": admin.get("id")
    }
    
    await db.referral_rewards.insert_one(reward)
    
    return {"message": "Reward created successfully", "reward_id": reward["id"]}

@api_router.get("/reports/birthday")
async def get_birthday_report(
    period: str = "today",
    admin: dict = Depends(require_admin)
):
    """Birthday report - today, 7 days, 30 days"""
    today = datetime.now(timezone.utc).date()
    
    # Get all members with DOB
    members = await db.members.find(
        {"date_of_birth": {"$exists": True, "$nin": [None, ""]}},
        {"_id": 0, "qr_code": 0}
    ).to_list(10000)
    
    result = []
    for m in members:
        dob_str = m.get("date_of_birth", "")
        if not dob_str:
            continue
        
        try:
            dob = datetime.fromisoformat(dob_str.replace("Z", "+00:00")).date() if "T" in dob_str else datetime.strptime(dob_str, "%Y-%m-%d").date()
            # Create birthday this year
            birthday_this_year = dob.replace(year=today.year)
            
            # If birthday already passed this year, check next year
            if birthday_this_year < today:
                birthday_this_year = dob.replace(year=today.year + 1)
            
            days_until = (birthday_this_year - today).days
            
            if period == "today" and days_until == 0:
                m["days_until_birthday"] = 0
                result.append(m)
            elif period == "7days" and 0 <= days_until <= 7:
                m["days_until_birthday"] = days_until
                result.append(m)
            elif period == "30days" and 0 <= days_until <= 30:
                m["days_until_birthday"] = days_until
                result.append(m)
        except Exception:
            continue
    
    # Sort by days until birthday
    result.sort(key=lambda x: x.get("days_until_birthday", 999))
    
    return result

@api_router.get("/reports/expiry")
async def get_expiry_report(
    period: str = "7days",
    admin: dict = Depends(require_admin)
):
    """Membership expiry report"""
    today = datetime.now(timezone.utc).date()
    
    if period == "7days":
        end_date = today + timedelta(days=7)
        query = {
            "membership_end": {"$gte": today.isoformat(), "$lte": end_date.isoformat()},
            "status": {"$ne": MembershipStatus.EXPIRED}
        }
    elif period == "30days":
        end_date = today + timedelta(days=30)
        query = {
            "membership_end": {"$gte": today.isoformat(), "$lte": end_date.isoformat()},
            "status": {"$ne": MembershipStatus.EXPIRED}
        }
    else:  # expired
        query = {"status": MembershipStatus.EXPIRED}
    
    members = await db.members.find(query, {"_id": 0, "qr_code": 0}).sort("membership_end", 1).to_list(10000)
    
    # Calculate days until expiry
    for m in members:
        end_str = m.get("membership_end", "")
        if end_str:
            try:
                end_date = datetime.fromisoformat(end_str.replace("Z", "+00:00")).date() if "T" in end_str else datetime.strptime(end_str[:10], "%Y-%m-%d").date()
                m["days_until_expiry"] = (end_date - today).days
            except Exception:
                m["days_until_expiry"] = None
    
    return members

# Payment CRUD
@api_router.post("/payments")
async def create_payment(payment: PaymentCreate, admin: dict = Depends(require_admin)):
    """Record a payment"""
    payment_id = f"PAY-{uuid.uuid4().hex[:8].upper()}"
    
    # Get member and plan info
    member = await db.members.find_one({"member_id": payment.member_id})
    plan = await db.plans.find_one({"id": payment.plan_id}) if payment.plan_id else None
    
    payment_doc = {
        "id": payment_id,
        "member_id": payment.member_id,
        "member_name": member.get("name") if member else None,
        "amount": payment.amount,
        "payment_type": payment.payment_type,
        "payment_method": payment.payment_method,
        "transaction_id": payment.transaction_id,
        "plan_id": payment.plan_id,
        "plan_name": plan.get("name") if plan else None,
        "notes": payment.notes,
        "status": "completed",
        "created_at": datetime.now(timezone.utc).isoformat(),
        "created_by": admin.get("id")
    }
    
    await db.payments.insert_one(payment_doc)
    payment_doc.pop("_id", None)
    return payment_doc

@api_router.get("/admin/payments")
async def get_all_payments(
    member_id: Optional[str] = None,
    admin: dict = Depends(require_admin)
):
    """Get all payments for admin"""
    query = {}
    if member_id:
        query["member_id"] = member_id
    
    payments = await db.payments.find(query, {"_id": 0}).sort("created_at", -1).to_list(10000)
    return payments

# Export endpoints
@api_router.get("/reports/export-excel")
async def export_members_excel(
    report_type: str = "members",
    search: Optional[str] = None,
    status: Optional[str] = None,
    plan_id: Optional[str] = None,
    city: Optional[str] = None,
    pincode: Optional[str] = None,
    referral_id: Optional[str] = None,
    start_date: Optional[str] = None,
    end_date: Optional[str] = None,
    expiry_start: Optional[str] = None,
    expiry_end: Optional[str] = None,
    admin: dict = Depends(require_admin)
):
    """Export report to Excel"""
    wb = Workbook()
    ws = wb.active
    
    if report_type == "members":
        ws.title = "Members Report"
        
        # Build query
        query = {}
        if search:
            query["$or"] = [
                {"name": {"$regex": search, "$options": "i"}},
                {"mobile": {"$regex": search, "$options": "i"}},
                {"member_id": {"$regex": search, "$options": "i"}}
            ]
        if status:
            query["status"] = status
        if plan_id:
            query["plan_id"] = plan_id
        if city:
            query["city"] = {"$regex": city, "$options": "i"}
        if pincode:
            query["pincode"] = pincode
        if referral_id:
            query["referral_id"] = {"$regex": referral_id, "$options": "i"}
        if start_date or end_date:
            query["created_at"] = {}
            if start_date:
                query["created_at"]["$gte"] = start_date
            if end_date:
                query["created_at"]["$lte"] = end_date + "T23:59:59"
        if expiry_start or expiry_end:
            query["membership_end"] = {}
            if expiry_start:
                query["membership_end"]["$gte"] = expiry_start
            if expiry_end:
                query["membership_end"]["$lte"] = expiry_end + "T23:59:59"
        
        members = await db.members.find(query, {"_id": 0, "qr_code": 0}).to_list(10000)
        
        headers = ["Member ID", "Name", "Mobile", "Email", "DOB", "City", "Pincode", "Area", "Plan", "Status", "Referral ID", "Start Date", "End Date", "Created At"]
        ws.append(headers)
        
        for m in members:
            ws.append([
                m.get("member_id", ""),
                m.get("name", ""),
                m.get("mobile", ""),
                m.get("email", ""),
                m.get("date_of_birth", ""),
                m.get("city", ""),
                m.get("pincode", ""),
                m.get("area", ""),
                m.get("plan_name", ""),
                m.get("status", ""),
                m.get("referral_id", ""),
                m.get("membership_start", ""),
                m.get("membership_end", ""),
                m.get("created_at", "")
            ])
    
    elif report_type == "payments":
        ws.title = "Payments Report"
        payments = await db.payments.find({"status": "completed"}, {"_id": 0}).to_list(10000)
        
        headers = ["Payment ID", "Member ID", "Member Name", "Amount", "Type", "Method", "Transaction ID", "Plan", "Date"]
        ws.append(headers)
        
        for p in payments:
            ws.append([
                p.get("id", ""),
                p.get("member_id", ""),
                p.get("member_name", ""),
                p.get("amount", 0),
                p.get("payment_type", ""),
                p.get("payment_method", ""),
                p.get("transaction_id", ""),
                p.get("plan_name", ""),
                p.get("created_at", "")
            ])
    
    elif report_type == "birthday":
        ws.title = "Birthday Report"
        # Get members with upcoming birthdays (30 days)
        members = await db.members.find(
            {"date_of_birth": {"$exists": True, "$nin": [None, ""]}},
            {"_id": 0, "qr_code": 0}
        ).to_list(10000)
        
        headers = ["Member ID", "Name", "Mobile", "DOB", "Plan", "Status"]
        ws.append(headers)
        
        for m in members:
            ws.append([
                m.get("member_id", ""),
                m.get("name", ""),
                m.get("mobile", ""),
                m.get("date_of_birth", ""),
                m.get("plan_name", ""),
                m.get("status", "")
            ])
    
    elif report_type == "expiry":
        ws.title = "Expiry Report"
        today = datetime.now(timezone.utc).date()
        end_date_calc = today + timedelta(days=30)
        
        members = await db.members.find(
            {"membership_end": {"$lte": end_date_calc.isoformat()}},
            {"_id": 0, "qr_code": 0}
        ).to_list(10000)
        
        headers = ["Member ID", "Name", "Mobile", "Plan", "Status", "Expiry Date"]
        ws.append(headers)
        
        for m in members:
            ws.append([
                m.get("member_id", ""),
                m.get("name", ""),
                m.get("mobile", ""),
                m.get("plan_name", ""),
                m.get("status", ""),
                m.get("membership_end", "")
            ])
    
    elif report_type == "referral":
        ws.title = "Referral Report"
        members = await db.members.find(
            {"referral_id": {"$exists": True, "$nin": [None, ""]}},
            {"_id": 0, "qr_code": 0}
        ).to_list(10000)
        
        headers = ["Member ID", "Name", "Mobile", "Plan", "Referral ID", "Joined Date"]
        ws.append(headers)
        
        for m in members:
            ws.append([
                m.get("member_id", ""),
                m.get("name", ""),
                m.get("mobile", ""),
                m.get("plan_name", ""),
                m.get("referral_id", ""),
                m.get("created_at", "")
            ])
    
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={report_type}_report_{datetime.now().strftime('%Y%m%d')}.xlsx"}
    )

@api_router.get("/reports/export-csv")
async def export_csv(
    report_type: str = "members",
    status: Optional[str] = None,
    admin: dict = Depends(require_admin)
):
    """Export report to CSV"""
    import csv
    from io import StringIO
    
    output = StringIO()
    writer = csv.writer(output)
    
    if report_type == "members":
        query = {}
        if status:
            query["status"] = status
        members = await db.members.find(query, {"_id": 0, "qr_code": 0}).to_list(10000)
        
        writer.writerow(["Member ID", "Name", "Mobile", "Email", "DOB", "City", "Plan", "Status", "Referral ID", "Start Date", "End Date"])
        for m in members:
            writer.writerow([
                m.get("member_id", ""),
                m.get("name", ""),
                m.get("mobile", ""),
                m.get("email", ""),
                m.get("date_of_birth", ""),
                m.get("city", ""),
                m.get("plan_name", ""),
                m.get("status", ""),
                m.get("referral_id", ""),
                m.get("membership_start", ""),
                m.get("membership_end", "")
            ])
    
    csv_content = output.getvalue()
    output.close()
    
    return StreamingResponse(
        iter([csv_content]),
        media_type="text/csv",
        headers={"Content-Disposition": f"attachment; filename={report_type}_report_{datetime.now().strftime('%Y%m%d')}.csv"}
    )

# ==================== LEADS ENDPOINTS ====================

@api_router.post("/leads")
async def create_lead(lead: LeadCreate, background_tasks: BackgroundTasks):
    """Public endpoint for lead capture from landing page"""
    lead_id = str(uuid.uuid4())
    lead_doc = {
        "id": lead_id,
        "name": lead.name,
        "mobile": lead.mobile,
        "city": lead.city,
        "interested_in": lead.interested_in,
        "source": lead.source,
        "status": LeadStatus.NEW,
        "notes": "",
        "created_at": datetime.now(timezone.utc).isoformat()
    }
    await db.leads.insert_one(lead_doc)
    
    # Send notification email to leads@bitzclub.com
    background_tasks.add_task(
        EmailService.send_lead_notification,
        lead_doc
    )
    
    return {"message": "Thank you! We will contact you soon.", "id": lead_id}

@api_router.get("/leads")
async def get_leads(
    status: Optional[str] = None,
    interested_in: Optional[str] = None,
    search: Optional[str] = None,
    page: int = Query(1, ge=1),
    limit: int = Query(20, ge=1, le=100),
    admin: dict = Depends(require_admin)
):
    query = {}
    if status:
        query["status"] = status
    if interested_in:
        query["interested_in"] = interested_in
    if search:
        query["$or"] = [
            {"name": {"$regex": search, "$options": "i"}},
            {"mobile": {"$regex": search, "$options": "i"}},
            {"city": {"$regex": search, "$options": "i"}}
        ]
    
    skip = (page - 1) * limit
    total = await db.leads.count_documents(query)
    leads = await db.leads.find(query, {"_id": 0}).sort("created_at", -1).skip(skip).limit(limit).to_list(limit)
    
    return {
        "leads": leads,
        "total": total,
        "page": page,
        "limit": limit,
        "pages": (total + limit - 1) // limit
    }

@api_router.get("/leads/stats")
async def get_leads_stats(admin: dict = Depends(require_admin)):
    total = await db.leads.count_documents({})
    new_leads = await db.leads.count_documents({"status": LeadStatus.NEW})
    contacted = await db.leads.count_documents({"status": LeadStatus.CONTACTED})
    converted = await db.leads.count_documents({"status": LeadStatus.CONVERTED})
    membership_leads = await db.leads.count_documents({"interested_in": "membership"})
    partnership_leads = await db.leads.count_documents({"interested_in": "partnership"})
    
    return {
        "total": total,
        "new": new_leads,
        "contacted": contacted,
        "converted": converted,
        "membership_leads": membership_leads,
        "partnership_leads": partnership_leads
    }

@api_router.put("/leads/{lead_id}")
async def update_lead(lead_id: str, lead: LeadUpdate, admin: dict = Depends(require_admin)):
    update_data = {k: v for k, v in lead.model_dump().items() if v is not None}
    if not update_data:
        raise HTTPException(status_code=400, detail="No data to update")
    
    update_data["updated_at"] = datetime.now(timezone.utc).isoformat()
    result = await db.leads.update_one({"id": lead_id}, {"$set": update_data})
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Lead not found")
    
    return await db.leads.find_one({"id": lead_id}, {"_id": 0})

@api_router.delete("/leads/{lead_id}")
async def delete_lead(lead_id: str, admin: dict = Depends(require_admin)):
    result = await db.leads.delete_one({"id": lead_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Lead not found")
    return {"message": "Lead deleted"}

@api_router.get("/leads/export-excel")
async def export_leads_excel(
    status: Optional[str] = None,
    interested_in: Optional[str] = None,
    admin: dict = Depends(require_admin)
):
    query = {}
    if status:
        query["status"] = status
    if interested_in:
        query["interested_in"] = interested_in
    
    leads = await db.leads.find(query, {"_id": 0}).to_list(10000)
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Leads Report"
    
    headers = ["Name", "Mobile", "City", "Interested In", "Status", "Source", "Created At"]
    ws.append(headers)
    
    for lead in leads:
        ws.append([
            lead.get("name", ""),
            lead.get("mobile", ""),
            lead.get("city", ""),
            lead.get("interested_in", ""),
            lead.get("status", ""),
            lead.get("source", ""),
            lead.get("created_at", "")
        ])
    
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename=leads_report_{datetime.now().strftime('%Y%m%d')}.xlsx"}
    )

@api_router.post("/leads/assign")
async def assign_leads_to_telecaller(assignment: LeadAssign, admin: dict = Depends(require_admin)):
    """Assign multiple leads to a telecaller"""
    # Verify telecaller exists
    telecaller = await db.users.find_one({"id": assignment.telecaller_id, "role": UserRole.TELECALLER})
    if not telecaller:
        raise HTTPException(status_code=404, detail="Telecaller not found")
    
    # Update all leads
    result = await db.leads.update_many(
        {"id": {"$in": assignment.lead_ids}},
        {"$set": {
            "assigned_telecaller": assignment.telecaller_id,
            "assigned_telecaller_name": telecaller["name"],
            "assigned_at": datetime.now(timezone.utc).isoformat(),
            "assigned_by": admin["id"]
        }}
    )
    
    logger.info(f"[LEADS] Assigned {result.modified_count} leads to telecaller {telecaller['name']}")
    return {
        "message": f"Successfully assigned {result.modified_count} leads to {telecaller['name']}",
        "assigned_count": result.modified_count
    }

@api_router.get("/leads/by-telecaller")
async def get_leads_by_telecaller(
    telecaller_id: Optional[str] = None,
    status: Optional[str] = None,
    page: int = Query(1, ge=1),
    limit: int = Query(20, ge=1, le=100),
    user: dict = Depends(require_admin_or_telecaller)
):
    """Get leads assigned to a telecaller (or self if telecaller)"""
    query = {}
    
    # If telecaller, only show their assigned leads
    if user["role"] == UserRole.TELECALLER:
        query["assigned_telecaller"] = user["id"]
    elif telecaller_id:
        query["assigned_telecaller"] = telecaller_id
    
    if status:
        query["status"] = status
    
    skip = (page - 1) * limit
    total = await db.leads.count_documents(query)
    leads = await db.leads.find(query, {"_id": 0}).sort("created_at", -1).skip(skip).limit(limit).to_list(limit)
    
    return {
        "leads": leads,
        "total": total,
        "page": page,
        "limit": limit,
        "pages": (total + limit - 1) // limit
    }

@api_router.get("/leads/unassigned")
async def get_unassigned_leads(
    page: int = Query(1, ge=1),
    limit: int = Query(20, ge=1, le=100),
    admin: dict = Depends(require_admin)
):
    """Get leads not assigned to any telecaller"""
    query = {"$or": [{"assigned_telecaller": None}, {"assigned_telecaller": {"$exists": False}}]}
    
    skip = (page - 1) * limit
    total = await db.leads.count_documents(query)
    leads = await db.leads.find(query, {"_id": 0}).sort("created_at", -1).skip(skip).limit(limit).to_list(limit)
    
    return {
        "leads": leads,
        "total": total,
        "page": page,
        "limit": limit,
        "pages": (total + limit - 1) // limit
    }

# ==================== MARKETING LANDING PAGE ENDPOINTS ====================

@api_router.post("/marketing/lead")
async def create_marketing_lead(lead: MarketingLeadCreate, background_tasks: BackgroundTasks):
    """Step 1: Capture lead from marketing landing page"""
    # Normalize mobile number - strip country code for comparison if present
    mobile_to_check = lead.mobile
    if mobile_to_check.startswith('+'):
        # Extract just the number part for India (+91) or other countries
        for prefix in ['+91', '+1', '+44', '+971', '+65', '+61', '+49', '+33', '+81', '+86', '+966', '+974', '+968', '+973', '+965']:
            if mobile_to_check.startswith(prefix):
                mobile_to_check = mobile_to_check[len(prefix):]
                break
    
    # Check if mobile already registered as member (check both with and without country code)
    existing_member = await db.users.find_one({
        "$or": [
            {"mobile": lead.mobile},
            {"mobile": mobile_to_check},
            {"mobile": f"+91{mobile_to_check}"}  # Check common Indian format
        ]
    })
    if existing_member:
        raise HTTPException(status_code=400, detail="Mobile number already registered. Please login instead.")
    
    # Check if lead already exists with this mobile
    existing_lead = await db.marketing_leads.find_one({
        "$or": [
            {"mobile": lead.mobile},
            {"mobile": mobile_to_check}
        ],
        "status": {"$ne": "converted"}
    })
    if existing_lead:
        # Return existing lead for continuation
        return {
            "message": "Lead already exists",
            "lead_id": existing_lead["id"],
            "name": existing_lead["name"],
            "mobile": existing_lead["mobile"],
            "email": existing_lead.get("email"),
            "step": existing_lead.get("step", 1)
        }
    
    lead_id = str(uuid.uuid4())
    lead_doc = {
        "id": lead_id,
        "name": lead.name,
        "mobile": lead.mobile,
        "mobile_raw": mobile_to_check,  # Store without country code too
        "email": lead.email,
        "referral_code": lead.referral_code,
        "country_code": lead.country_code,
        "country": lead.country,
        "source": lead.source,
        "status": "step1_complete",
        "step": 1,
        "created_at": datetime.now(timezone.utc).isoformat()
    }
    await db.marketing_leads.insert_one(lead_doc)
    
    # Also create in regular leads for admin visibility
    regular_lead = {
        "id": str(uuid.uuid4()),
        "name": lead.name,
        "mobile": lead.mobile,
        "city": lead.country or "",
        "interested_in": "membership",
        "source": lead.source,
        "status": LeadStatus.NEW,
        "notes": f"Referral: {lead.referral_code or 'None'}, Email: {lead.email or 'None'}, Country: {lead.country or 'India'}",
        "marketing_lead_id": lead_id,
        "created_at": datetime.now(timezone.utc).isoformat()
    }
    await db.leads.insert_one(regular_lead)
    
    # Send notification
    background_tasks.add_task(
        EmailService.send_lead_notification,
        {**regular_lead, "city": f"Marketing - {lead.country or 'India'}"}
    )
    
    logger.info(f"[MARKETING] New lead captured: {lead.name} ({lead.mobile}), Country: {lead.country}, Referral: {lead.referral_code}")
    
    return {
        "message": "Lead captured successfully",
        "lead_id": lead_id,
        "name": lead.name,
        "mobile": lead.mobile,
        "email": lead.email
    }

@api_router.post("/marketing/lead/{lead_id}/step2")
async def marketing_lead_step2(lead_id: str, data: MarketingLeadStep2):
    """Step 2: Capture additional details and create payment order"""
    # Get lead
    lead = await db.marketing_leads.find_one({"id": lead_id})
    if not lead:
        raise HTTPException(status_code=404, detail="Lead not found")
    
    # Get plan details
    plan = await db.plans.find_one({"id": data.plan_id}, {"_id": 0})
    if not plan:
        raise HTTPException(status_code=404, detail="Selected plan not found")
    
    # Update lead with step 2 data (including state and country)
    await db.marketing_leads.update_one(
        {"id": lead_id},
        {"$set": {
            "address": data.address,
            "city": data.city,
            "state": data.state,
            "pincode": data.pincode,
            "country": data.country or lead.get("country", "India"),
            "date_of_birth": data.date_of_birth,
            "plan_id": data.plan_id,
            "plan_name": plan["name"],
            "plan_price": plan["price"],
            "password": pwd_context.hash(data.password),
            "status": "step2_complete",
            "step": 2,
            "updated_at": datetime.now(timezone.utc).isoformat()
        }}
    )
    
    # Determine currency based on country (default INR for now, Razorpay handles conversion)
    currency = "INR"
    
    # Create Razorpay order
    try:
        order = await RazorpayService.create_order(
            amount=plan["price"],
            member_id=lead_id,
            plan_id=data.plan_id,
            notes={
                "marketing_lead_id": lead_id,
                "name": lead["name"],
                "mobile": lead["mobile"],
                "country": data.country or lead.get("country", "India"),
                "plan_name": plan["name"]
            }
        )
        
        # Store order ID
        await db.marketing_leads.update_one(
            {"id": lead_id},
            {"$set": {"razorpay_order_id": order["id"]}}
        )
        
        return {
            "order_id": order["id"],
            "amount": plan["price"],
            "currency": currency,
            "razorpay_key": order["razorpay_key"],
            "plan_name": plan["name"],
            "lead_id": lead_id,
            "name": lead["name"],
            "email": lead.get("email"),
            "mobile": lead["mobile"],
            "country": data.country or lead.get("country", "India")
        }
    except Exception as e:
        logger.error(f"[MARKETING] Failed to create payment order: {str(e)}")
        raise HTTPException(status_code=500, detail=f"Failed to create payment order: {str(e)}")

@api_router.post("/marketing/lead/{lead_id}/complete")
async def complete_marketing_registration(
    lead_id: str,
    razorpay_payment_id: str,
    razorpay_order_id: str,
    razorpay_signature: str,
    background_tasks: BackgroundTasks
):
    """Step 3: Complete registration after payment"""
    # Get marketing lead
    lead = await db.marketing_leads.find_one({"id": lead_id})
    if not lead:
        raise HTTPException(status_code=404, detail="Lead not found")
    
    if lead.get("status") == "converted":
        raise HTTPException(status_code=400, detail="Registration already completed")
    
    # Verify payment signature
    is_valid = RazorpayService.verify_payment_signature(
        razorpay_payment_id, razorpay_order_id, razorpay_signature
    )
    
    if not is_valid:
        await db.marketing_leads.update_one(
            {"id": lead_id},
            {"$set": {"status": "payment_failed", "failed_at": datetime.now(timezone.utc).isoformat()}}
        )
        raise HTTPException(status_code=400, detail="Payment verification failed")
    
    # Generate member ID
    member_id = await generate_member_id()
    user_id = str(uuid.uuid4())
    
    # Create user account
    user = {
        "id": user_id,
        "member_id": member_id,
        "mobile": lead["mobile"],
        "password": lead["password"],  # Already hashed
        "name": lead["name"],
        "email": lead.get("email"),
        "date_of_birth": lead.get("date_of_birth"),
        "role": UserRole.MEMBER,
        "is_active": True,
        "created_at": datetime.now(timezone.utc).isoformat()
    }
    await db.users.insert_one(user)
    
    # Calculate membership dates
    start_date = datetime.now(timezone.utc)
    plan = await db.plans.find_one({"id": lead["plan_id"]}, {"_id": 0})
    duration_months = plan["duration_months"] if plan else 12
    end_date = start_date + timedelta(days=duration_months * 30)
    
    # Create member record
    member_doc = {
        "id": str(uuid.uuid4()),
        "user_id": user_id,
        "member_id": member_id,
        "name": lead["name"],
        "mobile": lead["mobile"],
        "email": lead.get("email"),
        "address": lead.get("address"),
        "city": lead.get("city"),
        "pincode": lead.get("pincode"),
        "date_of_birth": lead.get("date_of_birth"),
        "plan_id": lead["plan_id"],
        "plan_name": lead["plan_name"],
        "membership_start": start_date.isoformat(),
        "membership_end": end_date.isoformat(),
        "status": MembershipStatus.ACTIVE,
        "qr_code": generate_qr_code(member_id),
        "referral_id": lead.get("referral_code"),
        "source": lead.get("source", "marketing_landing"),
        "created_at": datetime.now(timezone.utc).isoformat()
    }
    await db.members.insert_one(member_doc)
    
    # Create payment record
    payment_doc = {
        "id": str(uuid.uuid4()),
        "order_id": razorpay_order_id,
        "member_id": member_id,
        "member_name": lead["name"],
        "plan_id": lead["plan_id"],
        "plan_name": lead["plan_name"],
        "amount": lead["plan_price"],
        "payment_type": "online",
        "payment_method": "razorpay",
        "razorpay_payment_id": razorpay_payment_id,
        "razorpay_order_id": razorpay_order_id,
        "status": "completed",
        "source": "marketing_landing",
        "created_at": datetime.now(timezone.utc).isoformat(),
        "completed_at": datetime.now(timezone.utc).isoformat()
    }
    await db.payments.insert_one(payment_doc)
    
    # Update marketing lead status
    await db.marketing_leads.update_one(
        {"id": lead_id},
        {"$set": {
            "status": "converted",
            "member_id": member_id,
            "converted_at": datetime.now(timezone.utc).isoformat()
        }}
    )
    
    # Update regular lead status
    await db.leads.update_one(
        {"marketing_lead_id": lead_id},
        {"$set": {"status": LeadStatus.CONVERTED, "notes": f"Converted to member: {member_id}"}}
    )
    
    # Send notifications
    validity_end = end_date.strftime("%d %b %Y")
    background_tasks.add_task(SMSService.send_registration_sms, {"mobile": lead["mobile"], "member_id": member_id})
    background_tasks.add_task(SMSService.send_payment_success_sms, {"mobile": lead["mobile"]}, lead["plan_price"], lead["plan_name"])
    background_tasks.add_task(SMSService.send_membership_activation_sms, {"mobile": lead["mobile"]}, lead["plan_name"], validity_end)
    
    if lead.get("email"):
        background_tasks.add_task(EmailService.send_welcome_email, user)
    
    background_tasks.add_task(EmailService.send_membership_notification, member_doc, {"name": lead["plan_name"], "price": lead["plan_price"], "duration_months": duration_months})
    
    # Remove password and _id before creating token
    user.pop('_id', None)
    
    # Create access token
    token = create_access_token({"sub": user_id, "role": UserRole.MEMBER})
    user_response = {k: v for k, v in user.items() if k != "password"}
    
    logger.info(f"[MARKETING] Registration completed for {member_id} from marketing landing")
    
    return {
        "success": True,
        "message": "Registration successful! Welcome to BITZ Club!",
        "member_id": member_id,
        "access_token": token,
        "token_type": "bearer",
        "user": user_response,
        "membership": {
            "plan_name": lead["plan_name"],
            "start_date": start_date.isoformat(),
            "end_date": end_date.isoformat(),
            "status": "active"
        }
    }

@api_router.post("/marketing/enquiry")
async def create_enquiry(enquiry: EnquiryCreate, background_tasks: BackgroundTasks):
    """Create an enquiry from the chat/contact form"""
    enquiry_id = str(uuid.uuid4())
    enquiry_doc = {
        "id": enquiry_id,
        "name": enquiry.name,
        "mobile": enquiry.mobile,
        "email": enquiry.email,
        "message": enquiry.message,
        "source": enquiry.source,
        "status": "new",
        "created_at": datetime.now(timezone.utc).isoformat()
    }
    await db.enquiries.insert_one(enquiry_doc)
    
    # Also add to leads
    lead_doc = {
        "id": str(uuid.uuid4()),
        "name": enquiry.name,
        "mobile": enquiry.mobile,
        "city": "Chat Enquiry",
        "interested_in": "membership",
        "source": f"chat_{enquiry.source}",
        "status": LeadStatus.NEW,
        "notes": enquiry.message,
        "created_at": datetime.now(timezone.utc).isoformat()
    }
    await db.leads.insert_one(lead_doc)
    
    # Send notification
    background_tasks.add_task(
        EmailService.send_lead_notification,
        {**lead_doc, "city": f"Chat: {enquiry.message[:50]}..."}
    )
    
    logger.info(f"[ENQUIRY] New enquiry from {enquiry.name} ({enquiry.mobile})")
    
    return {"message": "Thank you! We will get back to you soon.", "id": enquiry_id}

@api_router.get("/marketing/leads")
async def get_marketing_leads(
    status: Optional[str] = None,
    page: int = Query(1, ge=1),
    limit: int = Query(20, ge=1, le=100),
    admin: dict = Depends(require_admin)
):
    """Get marketing leads for admin dashboard"""
    query = {}
    if status:
        query["status"] = status
    
    skip = (page - 1) * limit
    total = await db.marketing_leads.count_documents(query)
    leads = await db.marketing_leads.find(query, {"_id": 0, "password": 0}).sort("created_at", -1).skip(skip).limit(limit).to_list(limit)
    
    return {
        "leads": leads,
        "total": total,
        "page": page,
        "limit": limit,
        "pages": (total + limit - 1) // limit
    }

@api_router.get("/enquiries")
async def get_enquiries(
    status: Optional[str] = None,
    page: int = Query(1, ge=1),
    limit: int = Query(20, ge=1, le=100),
    admin: dict = Depends(require_admin)
):
    """Get enquiries for admin dashboard"""
    query = {}
    if status:
        query["status"] = status
    
    skip = (page - 1) * limit
    total = await db.enquiries.count_documents(query)
    enquiries = await db.enquiries.find(query, {"_id": 0}).sort("created_at", -1).skip(skip).limit(limit).to_list(limit)
    
    return {
        "enquiries": enquiries,
        "total": total,
        "page": page,
        "limit": limit,
        "pages": (total + limit - 1) // limit
    }

@api_router.put("/enquiries/{enquiry_id}")
async def update_enquiry(enquiry_id: str, status: str, admin: dict = Depends(require_admin)):
    """Update enquiry status"""
    result = await db.enquiries.update_one(
        {"id": enquiry_id},
        {"$set": {"status": status, "updated_at": datetime.now(timezone.utc).isoformat()}}
    )
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Enquiry not found")
    return {"message": "Enquiry updated"}


# ==================== CONTENT MANAGEMENT ENDPOINTS ====================

@api_router.get("/content/experiences")
async def get_experiences():
    """Get all lifestyle experiences for the website"""
    experiences = await db.experiences.find({}, {"_id": 0}).sort("order", 1).to_list(100)
    
    # Return default experiences if none exist
    if not experiences:
        default_experiences = [
            {"id": "exp_hotels", "title": "Luxury Hotels", "image_url": "https://images.unsplash.com/photo-1566073771259-6a8506099945?w=800&q=80", "discount": "Up to 40% Off", "description": "Exclusive rates at 5-star properties", "icon": "Hotel", "order": 1, "is_active": True},
            {"id": "exp_dining", "title": "Fine Dining", "image_url": "https://images.unsplash.com/photo-1414235077428-338989a2e8c0?w=800&q=80", "discount": "Up to 25% Off", "description": "Premium restaurants worldwide", "icon": "UtensilsCrossed", "order": 2, "is_active": True},
            {"id": "exp_spa", "title": "Spa & Wellness", "image_url": "https://images.unsplash.com/photo-1544161515-4ab6ce6db874?w=800&q=80", "discount": "Up to 35% Off", "description": "Rejuvenate at luxury spas", "icon": "Sparkles", "order": 3, "is_active": True},
            {"id": "exp_gym", "title": "Premium Gyms", "image_url": "https://images.unsplash.com/photo-1534438327276-14e5300c3a48?w=800&q=80", "discount": "Up to 30% Off", "description": "Access elite fitness centers", "icon": "Dumbbell", "order": 4, "is_active": True},
            {"id": "exp_pool", "title": "Swimming Pool", "image_url": "https://images.unsplash.com/photo-1576013551627-0cc20b96c2a7?w=800&q=80", "discount": "Complimentary", "description": "Premium pool facilities", "icon": "Waves", "order": 5, "is_active": True},
            {"id": "exp_party", "title": "Party Hall", "image_url": "https://images.unsplash.com/photo-1519671482749-fd09be7ccebf?w=800&q=80", "discount": "Up to 20% Off", "description": "Celebrate in style", "icon": "Music", "order": 6, "is_active": True},
            {"id": "exp_wedding", "title": "Marriage Venue", "image_url": "https://images.unsplash.com/photo-1519741497674-611481863552?w=800&q=80", "discount": "Special Packages", "description": "Dream wedding destinations", "icon": "PartyPopper", "order": 7, "is_active": True},
            {"id": "exp_corporate", "title": "Corporate Day Out", "image_url": "https://images.unsplash.com/photo-1511578314322-379afb476865?w=800&q=80", "discount": "Corporate Rates", "description": "Team building at its finest", "icon": "Building2", "order": 8, "is_active": True}
        ]
        # Insert default experiences
        await db.experiences.insert_many(default_experiences)
        return default_experiences
    
    return experiences

@api_router.post("/content/experiences")
async def create_experience(experience: ExperienceContent, admin: dict = Depends(require_admin)):
    """Create a new lifestyle experience"""
    exp_id = f"exp_{uuid.uuid4().hex[:8]}"
    exp_doc = {
        "id": exp_id,
        "title": experience.title,
        "image_url": experience.image_url,
        "discount": experience.discount,
        "description": experience.description,
        "icon": experience.icon,
        "order": experience.order,
        "is_active": experience.is_active,
        "created_at": datetime.now(timezone.utc).isoformat()
    }
    await db.experiences.insert_one(exp_doc)
    exp_doc.pop("_id", None)
    return exp_doc

@api_router.put("/content/experiences/{experience_id}")
async def update_experience(experience_id: str, experience: ExperienceContent, admin: dict = Depends(require_admin)):
    """Update a lifestyle experience"""
    update_data = {k: v for k, v in experience.dict().items() if v is not None and k != "id"}
    update_data["updated_at"] = datetime.now(timezone.utc).isoformat()
    
    result = await db.experiences.update_one({"id": experience_id}, {"$set": update_data})
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Experience not found")
    
    updated = await db.experiences.find_one({"id": experience_id}, {"_id": 0})
    return updated

@api_router.delete("/content/experiences/{experience_id}")
async def delete_experience(experience_id: str, admin: dict = Depends(require_admin)):
    """Delete a lifestyle experience"""
    result = await db.experiences.delete_one({"id": experience_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Experience not found")
    return {"message": "Experience deleted successfully"}

@api_router.get("/content/gallery")
async def get_gallery_images():
    """Get all gallery images"""
    images = await db.gallery.find({}, {"_id": 0}).sort("order", 1).to_list(100)
    return images

@api_router.post("/content/gallery")
async def create_gallery_image(image: GalleryImage, admin: dict = Depends(require_admin)):
    """Add a new gallery image"""
    img_id = f"img_{uuid.uuid4().hex[:8]}"
    img_doc = {
        "id": img_id,
        "title": image.title,
        "image_url": image.image_url,
        "category": image.category,
        "order": image.order,
        "is_active": image.is_active,
        "created_at": datetime.now(timezone.utc).isoformat()
    }
    await db.gallery.insert_one(img_doc)
    img_doc.pop("_id", None)
    return img_doc

@api_router.put("/content/gallery/{image_id}")
async def update_gallery_image(image_id: str, image: GalleryImage, admin: dict = Depends(require_admin)):
    """Update a gallery image"""
    update_data = {k: v for k, v in image.dict().items() if v is not None and k != "id"}
    update_data["updated_at"] = datetime.now(timezone.utc).isoformat()
    
    result = await db.gallery.update_one({"id": image_id}, {"$set": update_data})
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Image not found")
    
    updated = await db.gallery.find_one({"id": image_id}, {"_id": 0})
    return updated

@api_router.delete("/content/gallery/{image_id}")
async def delete_gallery_image(image_id: str, admin: dict = Depends(require_admin)):
    """Delete a gallery image"""
    result = await db.gallery.delete_one({"id": image_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Image not found")
    return {"message": "Image deleted successfully"}

@api_router.get("/content/settings")
async def get_website_settings():
    """Get website settings"""
    settings = await db.settings.find_one({"type": "website"}, {"_id": 0})
    if not settings:
        # Return default settings
        return {
            "hero_title": "Elevate Your Lifestyle",
            "hero_subtitle": "Join BITZ Club and unlock exclusive access to luxury hotels, fine dining, spas, premium gyms and a world of privileges.",
            "hero_image": "",
            "contact_phone": "+91 78129 01118",
            "contact_email": "hello@bitzclub.com",
            "contact_address": "Chennai, Tamil Nadu, India"
        }
    return settings

@api_router.put("/content/settings")
async def update_website_settings(settings: WebsiteSettings, admin: dict = Depends(require_admin)):
    """Update website settings"""
    update_data = {k: v for k, v in settings.dict().items() if v is not None}
    update_data["type"] = "website"
    update_data["updated_at"] = datetime.now(timezone.utc).isoformat()
    
    await db.settings.update_one({"type": "website"}, {"$set": update_data}, upsert=True)
    updated = await db.settings.find_one({"type": "website"}, {"_id": 0})
    return updated

# ==================== EVENTS ENDPOINTS ====================

@api_router.get("/events")
async def get_events(is_active: Optional[bool] = True, is_members_only: Optional[bool] = None):
    """Get all events"""
    query = {}
    if is_active is not None:
        query["is_active"] = is_active
    if is_members_only is not None:
        query["is_members_only"] = is_members_only
    
    events = await db.events.find(query, {"_id": 0}).sort("event_date", 1).to_list(100)
    return events

@api_router.post("/events")
async def create_event(event: EventCreate, admin: dict = Depends(require_admin)):
    """Create a new event (admin only)"""
    event_dict = event.dict()
    event_dict["id"] = str(uuid.uuid4())
    event_dict["created_at"] = datetime.now(timezone.utc).isoformat()
    event_dict["created_by"] = admin["id"]
    event_dict["attendees"] = []
    event_dict["attendee_count"] = 0
    
    await db.events.insert_one(event_dict)
    logger.info(f"[EVENT] Created event: {event.title} by {admin.get('name', admin['id'])}")
    
    return {k: v for k, v in event_dict.items() if k != "_id"}

@api_router.get("/events/{event_id}")
async def get_event(event_id: str):
    """Get event details"""
    event = await db.events.find_one({"id": event_id}, {"_id": 0})
    if not event:
        raise HTTPException(status_code=404, detail="Event not found")
    return event

@api_router.put("/events/{event_id}")
async def update_event(event_id: str, event: EventUpdate, admin: dict = Depends(require_admin)):
    """Update event (admin only)"""
    existing = await db.events.find_one({"id": event_id})
    if not existing:
        raise HTTPException(status_code=404, detail="Event not found")
    
    update_data = {k: v for k, v in event.dict().items() if v is not None}
    update_data["updated_at"] = datetime.now(timezone.utc).isoformat()
    
    await db.events.update_one({"id": event_id}, {"$set": update_data})
    
    updated = await db.events.find_one({"id": event_id}, {"_id": 0})
    return updated

@api_router.delete("/events/{event_id}")
async def delete_event(event_id: str, admin: dict = Depends(require_admin)):
    """Delete event (admin only)"""
    result = await db.events.delete_one({"id": event_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Event not found")
    return {"message": "Event deleted successfully"}

@api_router.post("/events/{event_id}/register")
async def register_for_event(event_id: str, registration: EventRegistration, user: dict = Depends(get_current_user)):
    """Register a member for an event"""
    event = await db.events.find_one({"id": event_id})
    if not event:
        raise HTTPException(status_code=404, detail="Event not found")
    
    if not event.get("is_active", True):
        raise HTTPException(status_code=400, detail="Event is not active")
    
    # Check if member is already registered
    existing_reg = await db.event_registrations.find_one({
        "event_id": event_id,
        "member_id": registration.member_id
    })
    if existing_reg:
        raise HTTPException(status_code=400, detail="Already registered for this event")
    
    # Check max attendees
    if event.get("max_attendees"):
        current_count = await db.event_registrations.count_documents({"event_id": event_id})
        if current_count >= event["max_attendees"]:
            raise HTTPException(status_code=400, detail="Event is full")
    
    reg_dict = {
        "id": str(uuid.uuid4()),
        "event_id": event_id,
        "member_id": registration.member_id,
        "guests": registration.guests,
        "notes": registration.notes,
        "status": "registered",
        "registered_at": datetime.now(timezone.utc).isoformat()
    }
    
    await db.event_registrations.insert_one(reg_dict)
    
    # Update attendee count
    await db.events.update_one(
        {"id": event_id},
        {"$inc": {"attendee_count": 1 + registration.guests}}
    )
    
    return {"message": "Successfully registered for event", "registration_id": reg_dict["id"]}

@api_router.get("/events/{event_id}/registrations")
async def get_event_registrations(event_id: str, admin: dict = Depends(require_admin)):
    """Get all registrations for an event (admin only)"""
    registrations = await db.event_registrations.find(
        {"event_id": event_id}, {"_id": 0}
    ).to_list(500)
    
    # Enrich with member details
    for reg in registrations:
        member = await db.members.find_one({"member_id": reg["member_id"]}, {"_id": 0, "name": 1, "mobile": 1, "email": 1})
        if member:
            reg["member_name"] = member.get("name")
            reg["member_mobile"] = member.get("mobile")
            reg["member_email"] = member.get("email")
    
    return registrations

@api_router.get("/member/events")
async def get_member_events(user: dict = Depends(get_current_user)):
    """Get events the member has registered for"""
    if user["role"] != UserRole.MEMBER:
        raise HTTPException(status_code=403, detail="Members only")
    
    member = await db.members.find_one({"user_id": user["id"]}, {"_id": 0})
    if not member:
        raise HTTPException(status_code=404, detail="Member not found")
    
    registrations = await db.event_registrations.find(
        {"member_id": member["member_id"]}, {"_id": 0}
    ).to_list(100)
    
    # Get event details
    events = []
    for reg in registrations:
        event = await db.events.find_one({"id": reg["event_id"]}, {"_id": 0})
        if event:
            event["registration"] = reg
            events.append(event)
    
    return events

# ==================== OFFERS ENDPOINTS ====================

@api_router.get("/offers")
async def get_offers(is_active: Optional[bool] = True, is_featured: Optional[bool] = None):
    """Get all offers for the website"""
    query = {}
    if is_active is not None:
        query["is_active"] = is_active
    if is_featured is not None:
        query["is_featured"] = is_featured
    
    offers = await db.offers.find(query, {"_id": 0}).sort("created_at", -1).to_list(100)
    
    # If no offers, create sample offers
    if not offers and is_active:
        sample_offers = [
            {
                "id": "offer_hotel_1",
                "title": "Weekend Getaway Special",
                "description": "Enjoy 30% off on weekend stays at premium hotels",
                "image_url": "https://images.unsplash.com/photo-1566073771259-6a8506099945?w=800&q=80",
                "discount_text": "30% OFF",
                "category": "Hotel",
                "is_active": True,
                "is_featured": True,
                "created_at": datetime.now(timezone.utc).isoformat()
            },
            {
                "id": "offer_dining_1",
                "title": "Fine Dining Experience",
                "description": "25% discount on dinner for two at partner restaurants",
                "image_url": "https://images.unsplash.com/photo-1414235077428-338989a2e8c0?w=800&q=80",
                "discount_text": "25% OFF",
                "category": "Restaurant",
                "is_active": True,
                "is_featured": True,
                "created_at": datetime.now(timezone.utc).isoformat()
            },
            {
                "id": "offer_spa_1",
                "title": "Spa & Wellness Retreat",
                "description": "Complimentary spa session with any hotel booking",
                "image_url": "https://images.unsplash.com/photo-1544161515-4ab6ce6db874?w=800&q=80",
                "discount_text": "FREE SPA",
                "category": "Spa",
                "is_active": True,
                "is_featured": False,
                "created_at": datetime.now(timezone.utc).isoformat()
            }
        ]
        await db.offers.insert_many(sample_offers)
        return sample_offers
    
    return offers

@api_router.get("/offers/{offer_id}")
async def get_offer(offer_id: str):
    """Get single offer details"""
    offer = await db.offers.find_one({"id": offer_id}, {"_id": 0})
    if not offer:
        raise HTTPException(status_code=404, detail="Offer not found")
    return offer

@api_router.post("/offers")
async def create_offer(offer: OfferCreate, admin: dict = Depends(require_admin)):
    """Create a new offer"""
    offer_id = f"offer_{uuid.uuid4().hex[:8]}"
    offer_doc = {
        "id": offer_id,
        **offer.dict(),
        "created_at": datetime.now(timezone.utc).isoformat(),
        "created_by": admin["id"]
    }
    await db.offers.insert_one(offer_doc)
    offer_doc.pop("_id", None)
    return offer_doc

@api_router.put("/offers/{offer_id}")
async def update_offer(offer_id: str, offer: OfferUpdate, admin: dict = Depends(require_admin)):
    """Update an offer"""
    update_data = {k: v for k, v in offer.dict().items() if v is not None}
    update_data["updated_at"] = datetime.now(timezone.utc).isoformat()
    
    result = await db.offers.update_one({"id": offer_id}, {"$set": update_data})
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Offer not found")
    
    updated = await db.offers.find_one({"id": offer_id}, {"_id": 0})
    return updated

@api_router.delete("/offers/{offer_id}")
async def delete_offer(offer_id: str, admin: dict = Depends(require_admin)):
    """Delete an offer"""
    result = await db.offers.delete_one({"id": offer_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Offer not found")
    return {"message": "Offer deleted successfully"}

# ==================== SEED DATA ====================

@api_router.post("/seed")
async def seed_data():
    """Seed initial data including admin user and sample plans"""
    
    # Create super admin if not exists
    admin_exists = await db.users.find_one({"role": UserRole.SUPER_ADMIN})
    if not admin_exists:
        admin = {
            "id": str(uuid.uuid4()),
            "mobile": "9999999999",
            "password": pwd_context.hash("admin123"),
            "name": "Super Admin",
            "email": "admin@bitzclub.com",
            "role": UserRole.SUPER_ADMIN,
            "is_active": True,
            "created_at": datetime.now(timezone.utc).isoformat()
        }
        await db.users.insert_one(admin)
    
    # Create sample plans if not exist
    plans_count = await db.plans.count_documents({})
    if plans_count == 0:
        sample_plans = [
            {
                "id": str(uuid.uuid4()),
                "name": "Silver",
                "description": "Basic membership with essential benefits",
                "duration_months": 6,
                "price": 5000,
                "features": ["Access to partner facilities", "Member discounts", "Digital membership card"],
                "is_active": True,
                "created_at": datetime.now(timezone.utc).isoformat()
            },
            {
                "id": str(uuid.uuid4()),
                "name": "Gold",
                "description": "Premium membership with enhanced benefits",
                "duration_months": 12,
                "price": 10000,
                "features": ["All Silver benefits", "Priority support", "Exclusive events access", "Higher discounts"],
                "is_active": True,
                "created_at": datetime.now(timezone.utc).isoformat()
            },
            {
                "id": str(uuid.uuid4()),
                "name": "Platinum",
                "description": "Elite membership with all benefits",
                "duration_months": 24,
                "price": 20000,
                "features": ["All Gold benefits", "Personal concierge", "VIP events", "Maximum discounts", "Family benefits"],
                "is_active": True,
                "created_at": datetime.now(timezone.utc).isoformat()
            }
        ]
        await db.plans.insert_many(sample_plans)
    
    # Create sample partners if not exist
    partners_count = await db.partners.count_documents({})
    if partners_count == 0:
        sample_partners = [
            {
                "id": str(uuid.uuid4()),
                "name": "Luxury Spa & Wellness",
                "description": "Premium spa and wellness center",
                "logo_url": "https://images.unsplash.com/photo-1544161515-4ab6ce6db874?w=200",
                "contact_email": "info@luxuryspa.com",
                "contact_phone": "9876543210",
                "address": "123 Wellness Street",
                "facilities": [
                    {"facility_name": "Spa Treatment", "discount_percentage": 20, "description": "All spa services"},
                    {"facility_name": "Gym Access", "discount_percentage": 15, "description": "Monthly gym membership"}
                ],
                "is_active": True,
                "created_at": datetime.now(timezone.utc).isoformat()
            },
            {
                "id": str(uuid.uuid4()),
                "name": "Fine Dining Restaurant",
                "description": "Exclusive fine dining experience",
                "logo_url": "https://images.unsplash.com/photo-1517248135467-4c7edcad34c4?w=200",
                "contact_email": "reservations@finedining.com",
                "contact_phone": "9876543211",
                "address": "456 Gourmet Avenue",
                "facilities": [
                    {"facility_name": "Dining", "discount_percentage": 15, "description": "All menu items"},
                    {"facility_name": "Private Events", "discount_percentage": 10, "description": "Private dining rooms"}
                ],
                "is_active": True,
                "created_at": datetime.now(timezone.utc).isoformat()
            }
        ]
        await db.partners.insert_many(sample_partners)
    
    return {"message": "Seed data created successfully", "admin_credentials": {"mobile": "9999999999", "password": "admin123"}}


# =====================
# COUPON ENDPOINTS
# =====================

# CouponCreate model is defined at line 270, reusing it here

@api_router.get("/coupons")
async def get_coupons(admin: dict = Depends(require_admin)):
    coupons = await db.coupons.find({}, {"_id": 0}).to_list(100)
    return coupons

@api_router.post("/coupons")
async def create_coupon(coupon: CouponCreate, admin: dict = Depends(require_admin)):
    # Check if code already exists
    existing = await db.coupons.find_one({"code": coupon.code.upper()})
    if existing:
        raise HTTPException(status_code=400, detail="Coupon code already exists")
    
    coupon_doc = {
        "id": str(uuid.uuid4()),
        "code": coupon.code.upper(),
        "discount_type": coupon.discount_type,
        "discount_value": coupon.discount_value,
        "min_amount": coupon.min_amount,
        "max_uses": coupon.max_uses,
        "valid_from": coupon.valid_from,
        "valid_until": coupon.valid_until,
        "applicable_plans": coupon.applicable_plans,
        "is_active": coupon.is_active,
        "times_used": 0,
        "created_at": datetime.now(timezone.utc).isoformat()
    }
    await db.coupons.insert_one(coupon_doc)
    coupon_doc.pop("_id", None)
    return coupon_doc

@api_router.put("/coupons/{coupon_id}")
async def update_coupon(coupon_id: str, coupon: CouponCreate, admin: dict = Depends(require_admin)):
    result = await db.coupons.update_one(
        {"id": coupon_id},
        {"$set": {
            "code": coupon.code.upper(),
            "discount_type": coupon.discount_type,
            "discount_value": coupon.discount_value,
            "min_amount": coupon.min_amount,
            "max_uses": coupon.max_uses,
            "valid_from": coupon.valid_from,
            "valid_until": coupon.valid_until,
            "applicable_plans": coupon.applicable_plans,
            "is_active": coupon.is_active,
            "updated_at": datetime.now(timezone.utc).isoformat()
        }}
    )
    if result.modified_count == 0:
        raise HTTPException(status_code=404, detail="Coupon not found")
    return {"message": "Coupon updated"}

@api_router.delete("/coupons/{coupon_id}")
async def delete_coupon(coupon_id: str, admin: dict = Depends(require_admin)):
    result = await db.coupons.delete_one({"id": coupon_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Coupon not found")
    return {"message": "Coupon deleted"}

@api_router.post("/coupons/validate")
async def validate_coupon(code: str, amount: float, plan_id: Optional[str] = None):
    coupon = await db.coupons.find_one({"code": code.upper(), "is_active": True})
    if not coupon:
        raise HTTPException(status_code=404, detail="Invalid coupon code")
    
    # Check validity dates
    now = datetime.now(timezone.utc)
    if coupon.get("valid_from") and datetime.fromisoformat(coupon["valid_from"].replace("Z", "+00:00")) > now:
        raise HTTPException(status_code=400, detail="Coupon not yet valid")
    if coupon.get("valid_until") and datetime.fromisoformat(coupon["valid_until"].replace("Z", "+00:00")) < now:
        raise HTTPException(status_code=400, detail="Coupon has expired")
    
    # Check usage limit
    if coupon.get("max_uses") and coupon.get("times_used", 0) >= coupon["max_uses"]:
        raise HTTPException(status_code=400, detail="Coupon usage limit reached")
    
    # Check min amount
    if coupon.get("min_amount") and amount < coupon["min_amount"]:
        raise HTTPException(status_code=400, detail=f"Minimum order amount is ₹{coupon['min_amount']}")
    
    # Calculate discount
    if coupon["discount_type"] == "percentage":
        discount = amount * (coupon["discount_value"] / 100)
    else:
        discount = coupon["discount_value"]
    
    discount = min(discount, amount)  # Discount cannot exceed amount
    
    return {
        "valid": True,
        "code": coupon["code"],
        "discount_type": coupon["discount_type"],
        "discount_value": coupon["discount_value"],
        "discount_amount": discount,
        "final_amount": amount - discount
    }

# =====================
# RENEWAL ENDPOINTS
# =====================

class RenewalRequest(BaseModel):
    plan_id: str
    payment_method: str = "cash"
    amount: Optional[float] = None
    notes: Optional[str] = None
    coupon_code: Optional[str] = None

@api_router.post("/members/{member_id}/renew")
async def renew_membership(member_id: str, renewal: RenewalRequest, admin: dict = Depends(require_admin)):
    member = await db.members.find_one({"member_id": member_id})
    if not member:
        member = await db.members.find_one({"id": member_id})
    if not member:
        raise HTTPException(status_code=404, detail="Member not found")
    
    plan = await db.plans.find_one({"id": renewal.plan_id})
    if not plan:
        raise HTTPException(status_code=404, detail="Plan not found")
    
    # Calculate amount with GST
    base_amount = renewal.amount if renewal.amount else plan["price"]
    gst_amount = round(base_amount * GST_RATE, 2)
    discount_amount = 0
    
    # Apply coupon if provided
    if renewal.coupon_code:
        coupon = await db.coupons.find_one({"code": renewal.coupon_code.upper(), "is_active": True})
        if coupon:
            if coupon["discount_type"] == "percentage":
                discount_amount = round(base_amount * (coupon["discount_value"] / 100), 2)
                if coupon.get("max_discount"):
                    discount_amount = min(discount_amount, coupon["max_discount"])
            else:
                discount_amount = coupon["discount_value"]
    
    total_amount = base_amount + gst_amount - discount_amount
    
    # Calculate new membership dates
    current_end = member.get("membership_end")
    if current_end:
        try:
            start_date = datetime.fromisoformat(current_end.replace("Z", "+00:00"))
            if start_date < datetime.now(timezone.utc):
                start_date = datetime.now(timezone.utc)
        except Exception:
            start_date = datetime.now(timezone.utc)
    else:
        start_date = datetime.now(timezone.utc)
    
    end_date = start_date + timedelta(days=plan.get("duration_months", 12) * 30)
    
    # Update member
    await db.members.update_one(
        {"member_id": member.get("member_id", member_id)},
        {"$set": {
            "plan_id": plan["id"],
            "plan_name": plan["name"],
            "membership_start": start_date.isoformat(),
            "membership_end": end_date.isoformat(),
            "status": "active",
            "updated_at": datetime.now(timezone.utc).isoformat()
        }}
    )
    
    # Create payment record with GST
    payment_doc = {
        "id": str(uuid.uuid4()),
        "order_id": f"RENEW-{datetime.now().strftime('%Y%m%d%H%M%S')}",
        "member_id": member.get("member_id", member_id),
        "member_name": member.get("name", ""),
        "plan_id": plan["id"],
        "plan_name": plan["name"],
        "base_amount": base_amount,
        "gst_amount": gst_amount,
        "gst_rate": GST_RATE * 100,
        "discount_amount": discount_amount,
        "coupon_code": renewal.coupon_code,
        "amount": total_amount,
        "payment_type": "renewal",
        "payment_method": renewal.payment_method,
        "notes": renewal.notes,
        "status": "completed",
        "created_at": datetime.now(timezone.utc).isoformat()
    }
    await db.payments.insert_one(payment_doc)
    
    return {
        "message": "Membership renewed successfully", 
        "new_end_date": end_date.isoformat(),
        "payment": {
            "base_amount": base_amount,
            "gst_amount": gst_amount,
            "discount_amount": discount_amount,
            "total_amount": total_amount
        }
    }


# =====================
# MAINTENANCE FEE ENDPOINTS
# =====================

@api_router.get("/maintenance-fees")
async def get_maintenance_fees(
    member_id: Optional[str] = None,
    status: Optional[str] = None,
    admin: dict = Depends(require_admin)
):
    """Get all maintenance fees with optional filters"""
    query = {}
    if member_id:
        query["member_id"] = member_id
    if status:
        query["status"] = status
    
    fees = await db.maintenance_fees.find(query, {"_id": 0}).sort("due_date", -1).to_list(500)
    return fees

@api_router.post("/maintenance-fees")
async def create_maintenance_fee(fee: MaintenanceFeeCreate, admin: dict = Depends(require_admin)):
    """Create a new maintenance fee record with discount and tax support"""
    member = await db.members.find_one({"member_id": fee.member_id})
    if not member:
        raise HTTPException(status_code=404, detail="Member not found")
    
    # Get plan for category-based maintenance amount
    plan = None
    if fee.plan_id:
        plan = await db.plans.find_one({"id": fee.plan_id}, {"_id": 0})
    elif member.get("plan_id"):
        plan = await db.plans.find_one({"id": member.get("plan_id")}, {"_id": 0})
    
    # Calculate amounts
    base_amount = fee.amount
    discount_amount = fee.discount_amount or 0
    tax_rate = min(fee.tax_rate or 0, 5) / 100  # Max 5% tax
    
    amount_after_discount = base_amount - discount_amount
    tax_amount = amount_after_discount * tax_rate
    final_amount = amount_after_discount + tax_amount
    
    fee_doc = {
        "id": str(uuid.uuid4()),
        "member_id": fee.member_id,
        "member_name": member.get("name", ""),
        "plan_id": plan.get("id") if plan else None,
        "plan_name": plan.get("name") if plan else member.get("plan_name", ""),
        "base_amount": base_amount,
        "discount_amount": discount_amount,
        "discount_reason": fee.discount_reason,
        "tax_rate": fee.tax_rate or 0,
        "tax_amount": round(tax_amount, 2),
        "amount": round(final_amount, 2),  # Final amount after discount and tax
        "fee_type": fee.fee_type,
        "due_date": fee.due_date,
        "status": "pending",
        "payment_method": fee.payment_method,
        "transaction_id": fee.transaction_id,
        "notes": fee.notes,
        "created_at": datetime.now(timezone.utc).isoformat(),
        "created_by": admin["id"]
    }
    await db.maintenance_fees.insert_one(fee_doc)
    
    logger.info(f"[MAINTENANCE] Created fee for {fee.member_id}: ₹{final_amount} (base: ₹{base_amount}, discount: ₹{discount_amount}, tax: ₹{round(tax_amount, 2)})")
    return {"message": "Maintenance fee created", "id": fee_doc["id"], "amount": round(final_amount, 2)}

@api_router.post("/maintenance-fees/{fee_id}/pay")
async def pay_maintenance_fee(
    fee_id: str,
    payment_method: str,
    transaction_id: Optional[str] = None,
    notes: Optional[str] = None,
    admin: dict = Depends(require_admin)
):
    """Mark maintenance fee as paid and create payment record"""
    fee = await db.maintenance_fees.find_one({"id": fee_id})
    if not fee:
        raise HTTPException(status_code=404, detail="Maintenance fee not found")
    
    if fee["status"] == "paid":
        raise HTTPException(status_code=400, detail="Fee already paid")
    
    paid_date = datetime.now(timezone.utc).isoformat()
    
    # Update maintenance fee
    await db.maintenance_fees.update_one(
        {"id": fee_id},
        {"$set": {
            "status": "paid",
            "paid_date": paid_date,
            "payment_method": payment_method,
            "transaction_id": transaction_id,
            "notes": notes,
            "updated_at": paid_date,
            "updated_by": admin["id"]
        }}
    )
    
    # Create payment record for member's payment history
    payment_doc = {
        "id": str(uuid.uuid4()),
        "order_id": f"MAINT-{fee_id[:8].upper()}",
        "member_id": fee["member_id"],
        "member_name": fee["member_name"],
        "payment_type": "maintenance",
        "fee_type": fee["fee_type"],
        "base_amount": fee.get("base_amount", fee["amount"]),
        "discount_amount": fee.get("discount_amount", 0),
        "tax_amount": fee.get("tax_amount", 0),
        "amount": fee["amount"],
        "payment_method": payment_method,
        "transaction_id": transaction_id,
        "status": "completed",
        "notes": notes or f"Maintenance fee ({fee['fee_type']})",
        "maintenance_fee_id": fee_id,
        "created_at": paid_date,
        "completed_at": paid_date,
        "created_by": admin["id"]
    }
    await db.payments.insert_one(payment_doc)
    
    logger.info(f"[MAINTENANCE] Paid fee {fee_id} for {fee['member_id']}: ₹{fee['amount']}")
    return {"message": "Maintenance fee paid and recorded in payment history", "payment_id": payment_doc["id"]}

@api_router.put("/maintenance-fees/{fee_id}")
async def update_maintenance_fee(fee_id: str, update: MaintenanceFeeUpdate, admin: dict = Depends(require_admin)):
    """Update maintenance fee (mark as paid, etc.)"""
    update_data = {k: v for k, v in update.dict().items() if v is not None}
    if update_data.get("status") == "paid" and not update_data.get("paid_date"):
        update_data["paid_date"] = datetime.now(timezone.utc).isoformat()
    update_data["updated_at"] = datetime.now(timezone.utc).isoformat()
    
    result = await db.maintenance_fees.update_one({"id": fee_id}, {"$set": update_data})
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Maintenance fee not found")
    return {"message": "Maintenance fee updated"}

@api_router.delete("/maintenance-fees/{fee_id}")
async def delete_maintenance_fee(fee_id: str, admin: dict = Depends(require_admin)):
    """Delete a maintenance fee record"""
    result = await db.maintenance_fees.delete_one({"id": fee_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Maintenance fee not found")
    return {"message": "Maintenance fee deleted"}

# =====================
# ENHANCED PAYMENT ENDPOINTS
# =====================

@api_router.post("/payments/manual")
async def create_manual_payment(payment: PaymentCreate, admin: dict = Depends(require_admin)):
    """Create a manual/offline payment record"""
    member = await db.members.find_one({"member_id": payment.member_id})
    if not member:
        raise HTTPException(status_code=404, detail="Member not found")
    
    plan = None
    if payment.plan_id:
        plan = await db.plans.find_one({"id": payment.plan_id})
    
    # Calculate GST
    base_amount = payment.amount
    gst_amount = payment.gst_amount if payment.gst_amount else round(base_amount * GST_RATE, 2)
    total_amount = base_amount + gst_amount
    
    # Apply discount if coupon provided
    discount_amount = 0
    if payment.coupon_code:
        coupon = await db.coupons.find_one({"code": payment.coupon_code.upper(), "is_active": True})
        if coupon:
            if coupon["discount_type"] == "percentage":
                discount_amount = round(base_amount * (coupon["discount_value"] / 100), 2)
                if coupon.get("max_discount"):
                    discount_amount = min(discount_amount, coupon["max_discount"])
            else:
                discount_amount = coupon["discount_value"]
            total_amount -= discount_amount
    
    payment_doc = {
        "id": str(uuid.uuid4()),
        "order_id": f"OFFLINE-{datetime.now().strftime('%Y%m%d%H%M%S')}",
        "member_id": payment.member_id,
        "member_name": member.get("name", ""),
        "plan_id": payment.plan_id,
        "plan_name": plan["name"] if plan else None,
        "base_amount": base_amount,
        "gst_amount": gst_amount,
        "gst_rate": GST_RATE * 100,
        "discount_amount": discount_amount,
        "coupon_code": payment.coupon_code,
        "amount": total_amount,
        "payment_type": payment.payment_type,
        "payment_method": payment.payment_method or "cash",
        "transaction_id": payment.transaction_id,
        "notes": payment.notes,
        "status": "completed",
        "created_at": datetime.now(timezone.utc).isoformat(),
        "completed_at": datetime.now(timezone.utc).isoformat()
    }
    await db.payments.insert_one(payment_doc)
    
    # Update member status if plan provided
    if plan:
        start_date = datetime.now(timezone.utc)
        end_date = start_date + timedelta(days=plan.get("duration_months", 12) * 30)
        await db.members.update_one(
            {"member_id": payment.member_id},
            {"$set": {
                "plan_id": plan["id"],
                "plan_name": plan["name"],
                "membership_start": start_date.isoformat(),
                "membership_end": end_date.isoformat(),
                "status": "active",
                "updated_at": datetime.now(timezone.utc).isoformat()
            }}
        )
    
    return {
        "message": "Payment recorded successfully",
        "payment_id": payment_doc["id"],
        "base_amount": base_amount,
        "gst_amount": gst_amount,
        "discount_amount": discount_amount,
        "total_amount": total_amount
    }

@api_router.put("/payments/{payment_id}")
async def update_payment(payment_id: str, update: PaymentUpdate, admin: dict = Depends(require_admin)):
    """Update a payment record"""
    update_data = {k: v for k, v in update.dict().items() if v is not None}
    update_data["updated_at"] = datetime.now(timezone.utc).isoformat()
    
    result = await db.payments.update_one({"id": payment_id}, {"$set": update_data})
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Payment not found")
    return {"message": "Payment updated"}

# =====================
# ENHANCED REPORTS ENDPOINTS
# =====================

@api_router.get("/reports/payments")
async def get_payment_reports(
    start_date: Optional[str] = None,
    end_date: Optional[str] = None,
    payment_type: Optional[str] = None,
    payment_method: Optional[str] = None,
    plan_id: Optional[str] = None,
    min_amount: Optional[float] = None,
    max_amount: Optional[float] = None,
    status: Optional[str] = None,
    admin: dict = Depends(require_admin)
):
    """Get payment reports with comprehensive filters"""
    query = {}
    
    if start_date:
        query["created_at"] = {"$gte": start_date}
    if end_date:
        if "created_at" in query:
            query["created_at"]["$lte"] = end_date
        else:
            query["created_at"] = {"$lte": end_date}
    if payment_type:
        query["payment_type"] = payment_type
    if payment_method:
        query["payment_method"] = payment_method
    if plan_id:
        query["plan_id"] = plan_id
    if status:
        query["status"] = status
    if min_amount:
        query["amount"] = {"$gte": min_amount}
    if max_amount:
        if "amount" in query:
            query["amount"]["$lte"] = max_amount
        else:
            query["amount"] = {"$lte": max_amount}
    
    payments = await db.payments.find(query, {"_id": 0}).sort("created_at", -1).to_list(1000)
    
    # Calculate totals
    total_amount = sum(p.get("amount", 0) for p in payments)
    total_gst = sum(p.get("gst_amount", 0) for p in payments)
    total_discount = sum(p.get("discount_amount", 0) for p in payments)
    
    return {
        "payments": payments,
        "summary": {
            "total_payments": len(payments),
            "total_amount": round(total_amount, 2),
            "total_gst": round(total_gst, 2),
            "total_discount": round(total_discount, 2),
            "online_count": len([p for p in payments if p.get("payment_type") == "online"]),
            "offline_count": len([p for p in payments if p.get("payment_type") == "offline"]),
            "renewal_count": len([p for p in payments if p.get("payment_type") == "renewal"])
        }
    }

@api_router.get("/reports/maintenance")
async def get_maintenance_reports(
    start_date: Optional[str] = None,
    end_date: Optional[str] = None,
    status: Optional[str] = None,
    fee_type: Optional[str] = None,
    admin: dict = Depends(require_admin)
):
    """Get maintenance fee reports with filters"""
    query = {}
    
    if start_date:
        query["due_date"] = {"$gte": start_date}
    if end_date:
        if "due_date" in query:
            query["due_date"]["$lte"] = end_date
        else:
            query["due_date"] = {"$lte": end_date}
    if status:
        query["status"] = status
    if fee_type:
        query["fee_type"] = fee_type
    
    fees = await db.maintenance_fees.find(query, {"_id": 0}).sort("due_date", -1).to_list(1000)
    
    total_amount = sum(f.get("amount", 0) for f in fees)
    paid_amount = sum(f.get("amount", 0) for f in fees if f.get("status") == "paid")
    pending_amount = sum(f.get("amount", 0) for f in fees if f.get("status") == "pending")
    
    return {
        "fees": fees,
        "summary": {
            "total_records": len(fees),
            "total_amount": round(total_amount, 2),
            "paid_amount": round(paid_amount, 2),
            "pending_amount": round(pending_amount, 2),
            "paid_count": len([f for f in fees if f.get("status") == "paid"]),
            "pending_count": len([f for f in fees if f.get("status") == "pending"])
        }
    }

@api_router.get("/reports/referrals")
async def get_referral_reports(
    referral_id: Optional[str] = None,
    start_date: Optional[str] = None,
    end_date: Optional[str] = None,
    admin: dict = Depends(require_admin)
):
    """Get referral reports showing members by referral ID"""
    query = {}
    
    if referral_id:
        query["referral_id"] = referral_id
    else:
        query["referral_id"] = {"$exists": True, "$nin": [None, ""]}
    
    if start_date:
        query["created_at"] = {"$gte": start_date}
    if end_date:
        if "created_at" in query:
            query["created_at"]["$lte"] = end_date
        else:
            query["created_at"] = {"$lte": end_date}
    
    members = await db.members.find(query, {"_id": 0}).to_list(1000)
    
    # Group by referral ID
    referral_stats = {}
    for member in members:
        ref_id = member.get("referral_id", "Unknown")
        if ref_id not in referral_stats:
            referral_stats[ref_id] = {"count": 0, "members": []}
        referral_stats[ref_id]["count"] += 1
        referral_stats[ref_id]["members"].append({
            "member_id": member.get("member_id"),
            "name": member.get("name"),
            "plan_name": member.get("plan_name"),
            "created_at": member.get("created_at")
        })
    
    return {
        "total_referrals": len(members),
        "referral_breakdown": referral_stats
    }

@api_router.get("/reports/telecallers")
async def get_telecaller_reports(
    telecaller_id: Optional[str] = None,
    start_date: Optional[str] = None,
    end_date: Optional[str] = None,
    admin: dict = Depends(require_admin)
):
    """Get telecaller performance reports"""
    telecallers = await db.users.find({"role": "telecaller"}, {"_id": 0, "password": 0}).to_list(100)
    
    reports = []
    for tc in telecallers:
        if telecaller_id and tc["id"] != telecaller_id:
            continue
        
        # Get assigned members
        query = {"assigned_telecaller": tc["id"]}
        if start_date:
            query["created_at"] = {"$gte": start_date}
        if end_date:
            if "created_at" in query:
                query["created_at"]["$lte"] = end_date
            else:
                query["created_at"] = {"$lte": end_date}
        
        assigned_members = await db.members.count_documents(query)
        
        # Get follow-ups
        followups = await db.followups.find({"telecaller_id": tc["id"]}, {"_id": 0}).to_list(500)
        completed_followups = len([f for f in followups if f.get("status") == "completed"])
        pending_followups = len([f for f in followups if f.get("status") == "pending"])
        
        reports.append({
            "telecaller_id": tc["id"],
            "name": tc.get("name"),
            "mobile": tc.get("mobile"),
            "email": tc.get("email"),
            "assigned_members": assigned_members,
            "total_followups": len(followups),
            "completed_followups": completed_followups,
            "pending_followups": pending_followups
        })
    
    return reports

@api_router.get("/reports/general")
async def get_general_report(admin: dict = Depends(require_admin)):
    """Get general overview report with all key metrics"""
    
    # Member stats
    total_members = await db.members.count_documents({})
    active_members = await db.members.count_documents({"status": MembershipStatus.ACTIVE})
    expired_members = await db.members.count_documents({"status": MembershipStatus.EXPIRED})
    
    # Revenue stats
    payments = await db.payments.find({"status": "completed"}, {"_id": 0}).to_list(10000)
    total_revenue = sum(p.get("amount", 0) for p in payments)
    membership_revenue = sum(p.get("amount", 0) for p in payments if p.get("payment_type") != "maintenance")
    maintenance_revenue = sum(p.get("amount", 0) for p in payments if p.get("payment_type") == "maintenance")
    
    # Lead stats
    total_leads = await db.leads.count_documents({})
    converted_leads = await db.leads.count_documents({"status": LeadStatus.CONVERTED})
    conversion_rate = round((converted_leads / total_leads * 100), 2) if total_leads > 0 else 0
    
    # Plan distribution
    pipeline = [
        {"$group": {"_id": "$plan_name", "count": {"$sum": 1}}},
        {"$sort": {"count": -1}}
    ]
    plan_distribution = []
    async for doc in db.members.aggregate(pipeline):
        plan_distribution.append({"plan_name": doc["_id"] or "Unknown", "count": doc["count"]})
    
    # Monthly revenue trend (last 6 months)
    monthly_revenue = []
    for i in range(5, -1, -1):
        month_start = (datetime.now(timezone.utc).replace(day=1) - timedelta(days=i*30)).replace(day=1)
        month_end = (month_start + timedelta(days=32)).replace(day=1)
        
        month_payments = [p for p in payments if month_start.isoformat() <= p.get("created_at", "") < month_end.isoformat()]
        monthly_revenue.append({
            "month": month_start.strftime("%b %Y"),
            "revenue": sum(p.get("amount", 0) for p in month_payments),
            "count": len(month_payments)
        })
    
    # Telecaller stats
    telecallers = await db.users.count_documents({"role": UserRole.TELECALLER})
    
    # Family members
    family_members = await db.family_members.count_documents({"is_active": True})
    
    return {
        "members": {
            "total": total_members,
            "active": active_members,
            "expired": expired_members,
            "family_members": family_members
        },
        "revenue": {
            "total": round(total_revenue, 2),
            "membership": round(membership_revenue, 2),
            "maintenance": round(maintenance_revenue, 2)
        },
        "leads": {
            "total": total_leads,
            "converted": converted_leads,
            "conversion_rate": conversion_rate
        },
        "plan_distribution": plan_distribution,
        "monthly_revenue": monthly_revenue,
        "telecallers": telecallers
    }

@api_router.get("/reports/transactions")
async def get_transaction_report(
    start_date: Optional[str] = None,
    end_date: Optional[str] = None,
    payment_type: Optional[str] = None,  # membership, maintenance, renewal
    payment_method: Optional[str] = None,  # cash, upi, card, razorpay
    member_id: Optional[str] = None,
    page: int = Query(1, ge=1),
    limit: int = Query(50, ge=1, le=200),
    admin: dict = Depends(require_admin)
):
    """Get detailed transaction report with filters"""
    query = {"status": "completed"}
    
    if start_date:
        query["created_at"] = {"$gte": start_date}
    if end_date:
        if "created_at" in query:
            query["created_at"]["$lte"] = end_date
        else:
            query["created_at"] = {"$lte": end_date}
    if payment_type:
        query["payment_type"] = payment_type
    if payment_method:
        query["payment_method"] = payment_method
    if member_id:
        query["member_id"] = member_id
    
    skip = (page - 1) * limit
    total = await db.payments.count_documents(query)
    
    payments = await db.payments.find(query, {"_id": 0}).sort("created_at", -1).skip(skip).limit(limit).to_list(limit)
    
    # Calculate totals
    all_payments = await db.payments.find(query, {"_id": 0, "amount": 1, "payment_method": 1}).to_list(10000)
    total_amount = sum(p.get("amount", 0) for p in all_payments)
    
    # Group by payment method
    method_breakdown = {}
    for p in all_payments:
        method = p.get("payment_method", "unknown")
        method_breakdown[method] = method_breakdown.get(method, 0) + p.get("amount", 0)
    
    return {
        "transactions": payments,
        "total": total,
        "page": page,
        "limit": limit,
        "pages": (total + limit - 1) // limit,
        "summary": {
            "total_amount": round(total_amount, 2),
            "total_transactions": len(all_payments),
            "by_payment_method": {k: round(v, 2) for k, v in method_breakdown.items()}
        }
    }

@api_router.get("/reports/transactions/export")
async def export_transaction_report(
    start_date: Optional[str] = None,
    end_date: Optional[str] = None,
    payment_type: Optional[str] = None,
    admin: dict = Depends(require_admin)
):
    """Export transactions to Excel"""
    query = {"status": "completed"}
    
    if start_date:
        query["created_at"] = {"$gte": start_date}
    if end_date:
        if "created_at" in query:
            query["created_at"]["$lte"] = end_date
        else:
            query["created_at"] = {"$lte": end_date}
    if payment_type:
        query["payment_type"] = payment_type
    
    payments = await db.payments.find(query, {"_id": 0}).sort("created_at", -1).to_list(10000)
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Transaction Report"
    
    headers = ["Date", "Member ID", "Member Name", "Type", "Amount", "Payment Method", "Transaction ID", "Status"]
    ws.append(headers)
    
    for p in payments:
        ws.append([
            p.get("created_at", "")[:10],
            p.get("member_id", ""),
            p.get("member_name", ""),
            p.get("payment_type", ""),
            p.get("amount", 0),
            p.get("payment_method", ""),
            p.get("transaction_id", p.get("razorpay_payment_id", "")),
            p.get("status", "")
        ])
    
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename=transactions_{datetime.now().strftime('%Y%m%d')}.xlsx"}
    )


# Include router
app.include_router(api_router)

app.add_middleware(
    CORSMiddleware,
    allow_credentials=True,
    allow_origins=os.environ.get('CORS_ORIGINS', '*').split(','),
    allow_methods=["*"],
    allow_headers=["*"],
)

@app.on_event("shutdown")
async def shutdown_db_client():
    client.close()
